import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# Crear workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Clientes"

# Estilos
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = [
    "Nombre",
    "Cédula",
    "Teléfono",
    "Email",
    "Dirección",
    "Estado",
    "Ocupación",
    "Empresa"
]

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Datos de prueba
datos = [
    ["Juan Carlos Pérez", "12345678", "3001234567", "juan@email.com", "Calle 1 #10", "ACTIVO", "Ingeniero", "TechCorp"],
    ["María González López", "87654321", "3007654321", "maria@email.com", "Calle 2 #20", "ACTIVO", "Abogada", "LawFirm"],
    ["Carlos Mendoza Silva", "11111111", "3009999999", "carlos@email.com", "Calle 3 #30", "ACTIVO", "Contador", "FinanceGo"],
    ["Ana Martínez Rodríguez", "22222222", "3008888888", "ana@email.com", "Calle 4 #40", "ACTIVO", "Doctora", "ClinicaSalud"],
    ["Roberto Jiménez Díaz", "33333333", "3007777777", "roberto@email.com", "Calle 5 #50", "ACTIVO", "Electricista", "ElectroServ"],
    ["Claudia Fernández", "44444444", "3006666666", "claudia@email.com", "Calle 6 #60", "INACTIVO", "Psicóloga", "MindCare"],
    ["Diego Ramírez Vega", "55555555", "3005555555", "diego@email.com", "Calle 7 #70", "ACTIVO", "Plomero", "PlomeriaExcel"],
    ["Laura García López", "66666666", "3004444444", "laura@email.com", "Calle 8 #80", "ACTIVO", "Diseñadora", "DesignStudio"],
    ["Felipe Moreno Ruiz", "77777777", "3003333333", "felipe@email.com", "Calle 9 #90", "ACTIVO", "Chofer", "TransportCo"],
    ["Mónica Torres Sánchez", "88888888", "3002222222", "monica@email.com", "Calle 10 #100", "ACTIVO", "Vendedora", "RetailPlus"],
]

# Agregar datos
for row_num, row_data in enumerate(datos, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

# Ajustar ancho de columnas
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 18
ws.column_dimensions['H'].width = 18

# Crear hoja de instrucciones
ws_instr = wb.create_sheet("Instrucciones")
ws_instr.column_dimensions['A'].width = 80

instrucciones = [
    ["INSTRUCCIONES DE IMPORTACIÓN"],
    [""],
    ["1. Formato de archivo: Excel (.xlsx)"],
    ["2. Hoja requerida: 'Clientes'"],
    ["3. Columnas obligatorias:"],
    ["   - Nombre: Nombre completo del cliente"],
    ["   - Cédula: Número de cédula (único)"],
    ["   - Teléfono: Número de contacto"],
    ["   - Email: Correo electrónico"],
    ["   - Dirección: Domicilio del cliente"],
    ["   - Estado: ACTIVO o INACTIVO"],
    ["   - Ocupación: Profesión o actividad"],
    ["   - Empresa: Empresa donde labora"],
    [""],
    ["4. Estados válidos: ACTIVO, INACTIVO"],
    ["5. Los datos de ejemplo ya incluidos pueden ser modificados o eliminados"],
    ["6. No incluir encabezados adicionales"],
    ["7. Asegurar que cédulas sean únicas (no duplicadas)"],
    ["8. Campos obligatorios: Nombre, Cédula, Estado"],
    [""],
    ["PASOS PARA IMPORTAR:"],
    ["1. Ir a: Más > Centro de Exportaciones"],
    ["2. Seleccionar: Importar Clientes"],
    ["3. Subir este archivo Excel"],
    ["4. Confirmar importación"],
    [""],
    ["VALIDACIONES:"],
    ["✓ Cédula única (no puede haber duplicadas)"],
    ["✓ Email válido si se incluye"],
    ["✓ Estado debe ser ACTIVO o INACTIVO"],
]

for row_num, row_data in enumerate(instrucciones, 1):
    cell = ws_instr.cell(row=row_num, column=1)
    cell.value = row_data[0] if row_data else ""
    if row_num == 1:
        cell.font = Font(bold=True, size=14)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=14)

# Guardar archivo
output_path = "C:\\Users\\Juancho\\Desktop\\Importar_Clientes.xlsx"
wb.save(output_path)

print("✅ ARCHIVO GENERADO EXITOSAMENTE")
print(f"\n📁 Ubicación: {output_path}")
print(f"📊 Clientes de prueba: {len(datos)}")
print(f"📋 Hojas incluidas: 'Clientes' e 'Instrucciones'")
print("\n✓ Listo para importar en el sistema")
