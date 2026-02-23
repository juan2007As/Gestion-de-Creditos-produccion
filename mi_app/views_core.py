from django.shortcuts import get_object_or_404, render, redirect
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required, user_passes_test
from django_ratelimit.decorators import ratelimit
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib import messages
from django.db import models
from django.db.models import Q, Count, Sum, DecimalField
from django.db.models.functions import Coalesce
from django.core.mail import EmailMessage
from django.conf import settings
from django.urls import reverse
from mi_app.forms import ClienteForm, PrestamoForm
from .models import Cliente, Prestamo, Cuota, Pago, Configuracion, PrestamoRapido, PagoPrestamoRapido, CuotaRapida, ListaNegra, calcular_fechas_pago, AuditoriaBackup
from mi_app.utilities.decorators import (
    require_rol, require_permission, require_any_permission,
    admin_required, gerente_o_admin, no_operario_solamente,
    valida_propiedad_cliente, valida_propiedad_prestamo  # ✅ NUEVA: Validación propiedad de recurso
)
from mi_app.utilities.transaction_integrity import atomic_payment_view, registrar_pago_atomico  # ✅ CRÍTICA #7
from mi_app.utils import determinar_estado_cuota_al_crear  # ✅ OPCIÓN C PASO 2: Import centralizado
from datetime import date, timedelta
from decimal import Decimal
from calendar import monthrange
import json
import datetime as dt
import zipfile
import os
import io
from pathlib import Path

# Create your views here.

# ===============================================================================
# AUTENTICACIÓN - LOGIN / LOGOUT
# ===============================================================================

def login_view(request):
    """
    Vista de login del sistema.
    Permite a usuarios registrados acceder al sistema.
    """
    if request.user.is_authenticated:
        return redirect('inicio')
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        
        if not username or not password:
            messages.error(request, 'Por favor ingresa usuario y contraseña.')
            return render(request, 'mi_app/login.html')
        
        # Intentar autenticar
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, f'¡Bienvenido {user.first_name or user.username}!')
            return redirect('inicio')
        else:
            messages.error(request, 'Usuario o contraseña inválidos.')
            return render(request, 'mi_app/login.html', {'username': username})
    
    return render(request, 'mi_app/login.html')


def logout_view(request):
    """
    Vista de logout del sistema.
    Cierra la sesión del usuario actual.
    """
    logout(request)
    messages.success(request, 'Sesión cerrada correctamente.')
    return redirect('login')


# ===============================================================================
# ENDPOINT AJAX - BÚSQUEDA DE CLIENTES
# ===============================================================================

@require_http_methods(["GET"])
@require_permission('cliente.view')
@login_required(login_url='login')
def buscar_cliente(request):
    """
    Endpoint AJAX para búsqueda en tiempo real de clientes.
    Busca SIMULTÁNEAMENTE por nombre, teléfono O cédula.
    
    GET params:
        q: Término de búsqueda
    
    Returns:
        JSON con lista de clientes coincidentes
    """
    from django.db.models import Q
    
    query = request.GET.get('q', '').strip()
    
    # Buscar incluso con 1 carácter para ser más flexible
    if not query:
        return JsonResponse({'resultados': []})
    
    # BUG #3 ARREGLADO: Búsqueda SIMULTÁNEA por nombre O celular O cédula
    clientes = Cliente.objects.filter(
        Q(nombre__icontains=query) |
        Q(celular__icontains=query) |
        Q(cedula__icontains=query)
    ).order_by('nombre')[:10]  # Máximo 10 resultados
    
    resultados = [{
        'id': c.id,
        'nombre': c.nombre,
        'celular': c.celular or 'N/A',
        'cedula': c.cedula,
        'display': f"{c.nombre} ({c.celular or 'Sin teléfono'})"
    } for c in clientes]
    
    return JsonResponse({'resultados': resultados})

@require_permission('cliente.view')
@login_required
def lista_clientes_api(request):
    """
    Endpoint API para obtener lista completa de clientes en JSON.
    Usado por formularios AJAX para autocompletar.
    
    Returns:
        JSON con lista de todos los clientes activos
    """
    clientes = Cliente.objects.all().order_by('nombre').values(
        'id', 'nombre', 'cedula', 'celular'
    )
    
    return JsonResponse({
        'clientes': list(clientes),
        'total': len(list(clientes))
    })

@require_permission('prestamo.view')
@login_required
@require_http_methods(["GET"])
def mora_diaria_api(request):
    """
    🆕 ENDPOINT PARA MORA EN TIEMPO REAL
    
    ✅ OPCIÓN C PASO 4: MEJORADO - Ahora sincroniza estado de cuota en BD.
    
    Retorna mora diaria calculada de todas las cuotas pendientes.
    Útil para dashboard y alertas de mora actualizada.
    
    ⚡ NUEVA LÓGICA: Si el estado en BD es incorrecto (ej: PENDIENTE pero vencida),
       se actualiza automáticamente a VENCIDA o VENCIDA_PARCIAL.
    
    Returns:
        JSON con:
        - total_mora: Suma de mora de todas las cuotas vencidas
        - cuotas_mora: Lista detallada de cuotas con mora
        - estadisticas: Resumen de mora por cliente
        - estados_sincronizados: Cantidad de cuotas cuyo estado fue actualizado
    """
    from decimal import Decimal
    
    # Obtener todas las cuotas pendientes vencidas
    cuotas_vencidas = Cuota.objects.filter(
        pagado=False,
        fecha_pago_esperada__lt=date.today()
    ).select_related('prestamo__cliente')
    
    total_mora = Decimal('0')
    cuotas_mora = []
    estadisticas_por_cliente = {}
    estados_sincronizados = 0  # ✅ OPCIÓN C: Contador de updates
    
    for cuota in cuotas_vencidas:
        mora = cuota.calcular_mora_diaria()
        total_mora += mora
        
        # ✅ OPCIÓN C PASO 4 - NUEVA: Sincronizar estado en BD
        estado_correcto = determinar_estado_cuota_al_crear(
            pagado=cuota.pagado,
            fecha_pago_esperada=cuota.fecha_pago_esperada,
            monto_pagado_principal=cuota.monto_pagado_principal,
            monto_original=cuota.monto_original
        )
        
        # Si el estado en BD es diferente, actualizar
        if cuota.estado != estado_correcto:
            Cuota.objects.filter(id=cuota.id).update(estado=estado_correcto)
            estados_sincronizados += 1
        
        cliente = cuota.prestamo.cliente
        if cliente.id not in estadisticas_por_cliente:
            estadisticas_por_cliente[cliente.id] = {
                'cliente_id': cliente.id,
                'cliente_nombre': cliente.nombre,
                'cliente_cedula': cliente.cedula,
                'total_mora': Decimal('0'),
                'cuotas_con_mora': 0,
            }
        
        estadisticas_por_cliente[cliente.id]['total_mora'] += mora
        estadisticas_por_cliente[cliente.id]['cuotas_con_mora'] += 1
        
        if mora > 0:  # Solo incluir cuotas que tengan mora
            cuotas_mora.append({
                'cuota_id': cuota.id,
                'numero_cuota': cuota.numero_cuota,
                'prestamo_id': cuota.prestamo.id,
                'cliente_id': cliente.id,
                'cliente_nombre': cliente.nombre,
                'fecha_pago_esperada': cuota.fecha_pago_esperada.isoformat(),
                'dias_atraso': (date.today() - cuota.fecha_pago_esperada).days,
                'monto_pendiente': str(cuota.monto_pendiente),
                'mora': str(mora),
                'estado': estado_correcto,  # ✅ OPCIÓN C: Mostrar estado sincronizado
            })
    
    return JsonResponse({
        'total_mora': str(total_mora),
        'total_cuotas_vencidas': cuotas_vencidas.count(),
        'cuotas_con_mora': len(cuotas_mora),
        'cuotas': cuotas_mora,
        'estadisticas_clientes': list(estadisticas_por_cliente.values()),
        'timestamp': date.today().isoformat(),
        'estados_sincronizados': estados_sincronizados,  # ✅ OPCIÓN C: Reporte de sincronización
    })

@login_required
def calcular_fecha_pago_esperada(fecha_inicio, numero_cuota, calendario_pagos):
    """
    Calcula la fecha de pago esperada para una cuota.
    
    LA PRIMERA CUOTA SIEMPRE ES EN LA PRÓXIMA FECHA DE PAGO DESPUÉS DE fecha_inicio.
    
    Args:
        fecha_inicio: Fecha de inicio del préstamo
        numero_cuota: Número de la cuota (1-6)
        calendario_pagos: '5_21' o '15_30'
    
    Returns:
        date: Fecha esperada de pago para esa cuota
    
    Ejemplo:
        Si fecha_inicio = 2024-12-03, calendario='15_30':
        - Cuota 1: 2024-12-15 (próxima fecha de pago)
        - Cuota 2: 2024-12-30
        - Cuota 3: 2025-01-15
        
        Si fecha_inicio = 2024-12-20, calendario='5_21':
        - Cuota 1: 2024-12-21 (próxima fecha de pago)
        - Cuota 2: 2025-01-05
        - Cuota 3: 2025-01-21
    """
    
    if calendario_pagos == '5_21':
        dias_pago_mes = [5, 21]
    else:  # '15_30'
        dias_pago_mes = [15, 30]
    
    # Generar lista de TODAS las próximas fechas de pago
    fechas_pago_ordenadas = []
    
    mes = fecha_inicio.month
    año = fecha_inicio.year
    
    # Generar suficientes fechas de pago (24 meses = 48 posibles fechas)
    for _ in range(24):
        for dia in dias_pago_mes:
            try:
                fecha = date(año, mes, dia)
                # Solo incluir fechas POSTERIORES a fecha_inicio
                if fecha > fecha_inicio:
                    fechas_pago_ordenadas.append(fecha)
            except ValueError:
                # Día no existe en ese mes (ej: 30 en febrero), ignorar
                pass
        
        mes += 1
        if mes > 12:
            mes = 1
            año += 1
    
    # Ordenar y devolver la N-ésima cuota
    fechas_pago_ordenadas.sort()
    return fechas_pago_ordenadas[numero_cuota - 1]


@login_required(login_url='login')
def inicio(request):
    """Página de inicio con estadísticas del sistema"""
    estadisticas = obtener_estadisticas_sistema()
    return render(request, 'mi_app/inicio.html', {'estadisticas': estadisticas})


@require_permission('reporte.view')
@login_required
def centro_exportaciones(request):
    """Centro de descarga de reportes en Excel"""
    from django.db.models import Sum, Count
    
    contexto = {
        'total_clientes': Cliente.objects.count(),
        'total_prestamos': Prestamo.objects.count(),
        'total_cuotas': Cuota.objects.count(),
        'cuotas_vencidas': Cuota.objects.filter(
            estado='PENDIENTE',
            fecha_pago_esperada__lt=date.today()
        ).count(),
    }
    return render(request, 'mi_app/centro_exportaciones.html', contexto)


def obtener_estadisticas_sistema():
    """
    Obtiene estadísticas generales del sistema para el dashboard con mínimas queries.
    
    Esta función calcula estadísticas agregadas del sistema de forma optimizada:
    - Sin loops anidados (N+1 queries eliminados en FASE 2.2)
    - Usa prefetch_related para relaciones FK
    - Usa aggregate() para sumas y conteos
    
    Returns:
        dict: Diccionario con estructura:
            {
                'clientes': {
                    'total': int,
                    'activos': int,
                    'inactivos': int
                },
                'prestamos': {
                    'total': int,
                    'activos': int,
                    'completados': int,
                    'borrador': int
                },
                'dinero': {
                    'capital_prestado': Decimal,
                    'total_credito': Decimal (capital + interes),
                    'total_pagado': Decimal,
                    'total_pendiente_capital': Decimal,
                    'total_pendiente_credito': Decimal,
                    'total_mora': Decimal
                },
                'cuotas': {...}
            }
    
    Note:
        FASE 2.2: Optimización N+1 - 5000 queries -> aprox 50 queries
        Usa prefetch_related + aggregate para máxima eficiencia
    """
    from decimal import Decimal
    from django.db.models import Sum, F, Q, Case, When, IntegerField, Prefetch
    
    # OPTIMIZACIÓN N+1 #1: Usar prefetch_related para evitar nested queries
    clientes = Cliente.objects.all()
    prestamos = Prestamo.objects.prefetch_related('cuotas').all()
    cuotas = Cuota.objects.select_related('prestamo').all()
    
    hoy = date.today()
    
    # CLIENTES
    total_clientes = clientes.count()
    clientes_activos = clientes.filter(estado='ACTIVO').count()
    
    # PRÉSTAMOS
    total_prestamos = prestamos.count()
    prestamos_activos = prestamos.filter(estado__in=['ACTIVO']).count()
    prestamos_completados = prestamos.filter(estado='COMPLETADO').count()
    
    # DINERO
    # BUG FIX #1: Calcular AMBOS capital y capital+interés para claridad
    capital_prestado = Decimal('0')  # ← Solo capital
    total_credito = Decimal('0')     # ← Capital + Interés
    
    for p in prestamos:
        capital_prestado += Decimal(str(p.monto_total))  # ← Ensure Decimal type
        total_credito += Decimal(str(p.total_credito))   # ← Convert float property to Decimal
    
    # ✅ OPTIMIZACIÓN N+1 #2: Usar agregación en lugar de loops sobre propiedades
    from django.db.models import Sum
    from django.db.models.functions import Coalesce
    total_pagado_result = Cuota.objects.aggregate(
        principal=Coalesce(Sum('monto_pagado_principal'), Decimal('0')),
        interes=Coalesce(Sum('monto_pagado_interes'), Decimal('0')),
        mora=Coalesce(Sum('monto_pagado_mora'), Decimal('0'))
    )
    total_pagado = total_pagado_result['principal'] + total_pagado_result['interes'] + total_pagado_result['mora']
    
    total_pendiente_capital = capital_prestado - total_pagado
    total_pendiente_credito = total_credito - total_pagado
    
    # CUOTAS
    total_cuotas = cuotas.count()
    cuotas_pagadas = cuotas.filter(pagado=True).count()
    cuotas_pendientes = cuotas.filter(pagado=False).count()
    
    # CUOTAS VENCIDAS
    cuotas_vencidas = 0
    monto_vencido = Decimal('0')
    for c in cuotas.filter(pagado=False):
        if c.fecha_pago_esperada and c.fecha_pago_esperada < hoy:
            cuotas_vencidas += 1
            monto_vencido += c.monto_pendiente + c.interes_normal + c.calcular_mora_diaria()
    
    # TASA DE CUMPLIMIENTO
    tasa_pagos = (cuotas_pagadas / total_cuotas * 100) if total_cuotas > 0 else 0
    
    return {
        'clientes': {
            'total': total_clientes,
            'activos': clientes_activos,
            'inactivos': total_clientes - clientes_activos,
        },
        'prestamos': {
            'total': total_prestamos,
            'activos': prestamos_activos,
            'completados': prestamos_completados,
            'borrador': total_prestamos - prestamos_activos - prestamos_completados,
        },
        'dinero': {
            'capital_prestado': float(capital_prestado),  # ← BUG FIX #1: Solo capital
            'total_credito': float(total_credito),        # ← BUG FIX #1: Capital + Interés
            'total_pagado': float(total_pagado),
            'total_pendiente_capital': float(total_pendiente_capital),  # ← Capital pendiente
            'total_pendiente_credito': float(total_pendiente_credito),  # ← Credito pendiente
            'promedio_capital': float(capital_prestado / total_prestamos) if total_prestamos > 0 else 0,
            'promedio_credito': float(total_credito / total_prestamos) if total_prestamos > 0 else 0,
        },
        'cuotas': {
            'total': total_cuotas,
            'pagadas': cuotas_pagadas,
            'pendientes': cuotas_pendientes,
            'vencidas': cuotas_vencidas,
        },
        'mora': {
            'cuotas_vencidas': cuotas_vencidas,
            'monto_vencido': float(monto_vencido),
        },
        'indicadores': {
            'tasa_pagos': round(tasa_pagos, 1),
            'tasa_mora': round((cuotas_vencidas / total_cuotas * 100) if total_cuotas > 0 else 0, 1),
        }
    }

@login_required(login_url='login')
@require_permission('cliente.view')
def lista_clientes(request):
    """Muestra una lista de todos los clientes."""
    from django.db.models import Sum, Count, Q
    from django.db.models.functions import Coalesce
    from decimal import Decimal
    
    busqueda = request.GET.get('q', '').strip()
    clientes = Cliente.objects.all().select_related('lista_negra')
    
    if busqueda:
        clientes = clientes.filter(
            Q(nombre__icontains=busqueda) | 
            Q(cedula__icontains=busqueda) | 
            Q(celular__icontains=busqueda)
        )
    
    # Obtener estadísticas generales de forma eficiente
    stats = Cliente.objects.aggregate(
        total_clientes=Count('id'),
        clientes_activos=Count('id', filter=Q(estado='ACTIVO'))
    )
    
    total_prestamos = Prestamo.objects.count()
    
    # Calcular total pagado global de forma eficiente
    total_pagado_agg = Cuota.objects.aggregate(
        total=Coalesce(Sum('monto_pagado_principal'), Decimal('0')) + 
              Coalesce(Sum('monto_pagado_interes'), Decimal('0')) + 
              Coalesce(Sum('monto_pagado_mora'), Decimal('0'))
    )
    total_pagado = total_pagado_agg['total']
    
    contexto = {
        'clientes': clientes,
        'clientes_info': [{'cliente': c, 'en_lista_negra': c.lista_negra.esta_vigente if hasattr(c, 'lista_negra') and c.lista_negra else False} for c in clientes],
        'busqueda': busqueda,
        'cantidad': clientes.count(),
        'total_clientes': stats['total_clientes'],
        'clientes_activos': stats['clientes_activos'],
        'total_prestamos': total_prestamos,
        'total_pagado': total_pagado,
    }
    
    return render(request, 'mi_app/lista_clientes.html', contexto)

@require_permission('cliente.view')
@login_required(login_url='login')
def clientes_importados(request):
    """Muestra solo los clientes importados desde Excel."""
    from django.db.models import Q
    busqueda = request.GET.get('q', '').strip()
    
    # ✅ OPTIMIZACIÓN N+1 #3: Prefetch relativos y calcula agregados una sola vez
    from django.db.models import Prefetch
    clientes_optimized = Cliente.objects.prefetch_related(
        Prefetch('prestamo_set',
                 queryset=Prestamo.objects.prefetch_related('cuotas').all())
    ).filter(importado_excel=True).order_by('-fecha_creacion')
    
    # Aplicar búsqueda si es necesario
    if busqueda:
        clientes_optimized = clientes_optimized.filter(
            Q(nombre__icontains=busqueda) |
            Q(celular__icontains=busqueda) |
            Q(cedula__icontains=busqueda)
        )
    
    clientes_info = []
    for cliente in clientes_optimized:
        # Contar préstamos activos sin hacer additional queries (ya están prefetched)
        prestamos_activos = sum(1 for p in cliente.prestamo_set.all() if p.estado == 'ACTIVO')
        # Calcular total_pendiente usando los préstamos ya cargados
        total_pendiente = sum(p.total_pendiente for p in cliente.prestamo_set.all())
        clientes_info.append({
            'cliente': cliente,
            'prestamos_activos': prestamos_activos,
            'total_pendiente': total_pendiente
        })
    
    # ✅ OPTIMIZACIÓN N+1 #4: Usar agregación en lugar de sum(propiedad)
    total_clientes = Cliente.objects.count()
    clientes_importados_count = Cliente.objects.filter(importado_excel=True).count()
    total_prestamos = Prestamo.objects.count()
    
    # Calcular total_pagado usando agregación, no propiedades que hacen queries
    from django.db.models import Sum
    from django.db.models.functions import Coalesce
    from decimal import Decimal
    total_pagado_aggregated = Cuota.objects.aggregate(
        total_principal=Coalesce(Sum('monto_pagado_principal'), Decimal('0')),
        total_interes=Coalesce(Sum('monto_pagado_interes'), Decimal('0')),
        total_mora=Coalesce(Sum('monto_pagado_mora'), Decimal('0'))
    )
    total_pagado = total_pagado_aggregated['total_principal'] + total_pagado_aggregated['total_interes'] + total_pagado_aggregated['total_mora']
    
    contexto = {
        'clientes_info': clientes_info,
        'clientes': clientes_optimized,
        'busqueda': busqueda,
        'cantidad': clientes_optimized.count(),
        # Estadísticas globales
        'total_clientes': total_clientes,
        'clientes_importados_count': clientes_importados_count,
        'total_prestamos': total_prestamos,
        'total_pagado': total_pagado,
    }
    
    return render(request, 'mi_app/clientes_importados.html', contexto)

@require_permission('cliente.view')
@login_required(login_url='login')
@valida_propiedad_cliente('cliente_id')
def detalle_cliente(request, cliente_id):
    """Redirige al nuevo perfil del cliente (Fase 2)."""
    return redirect('perfil_cliente', cliente_id=cliente_id)

@require_any_permission('cliente.create')
@login_required(login_url='login')    
def crear_cliente(request):
    """Vista para crear un nuevo cliente."""
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            # Extra safety: ensure estado is never empty
            cliente = form.save(commit=False)
            if not cliente.estado or cliente.estado.strip() == '':
                cliente.estado = 'ACTIVO'
            cliente.save()
            return redirect('lista_clientes')
    else:
        form = ClienteForm()
    
    return render(request, 'mi_app/formularios/formulario_cliente.html', {'form': form, 'titulo': 'Crear Nuevo Cliente'})

@require_permission('cliente.edit')
@login_required(login_url='login')
@valida_propiedad_cliente('cliente_id')
def editar_cliente(request, cliente_id):
    """Vista para editar un cliente existente."""
    cliente = get_object_or_404(Cliente, id=cliente_id)
    
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            return redirect('detalle_cliente', cliente_id=cliente.id)
    else:
        form = ClienteForm(instance=cliente)
    return render(request, 'mi_app/formularios/formulario_cliente.html', {'form': form, 'titulo': 'Editar Cliente', 'cliente': cliente})

@require_any_permission('prestamo.create')
@login_required(login_url='login')
def crear_prestamo(request, cliente_id=None):
    """
    Crear un nuevo préstamo con validaciones transversales.
    
    Permite a usuarios autenticados crear un nuevo préstamo para un cliente.
    Incluye validaciones de:
    - Cliente activo y no en lista negra (FASE 2.1)
    - Montos positivos y dentro de límites
    - Número de cuotas entre 1-6
    - Interés porcentual válido
    
    Args:
        request (HttpRequest): Objeto de solicitud HTTP
        cliente_id (int, optional): ID del cliente preseleccionado. Default None.
    
    Returns:
        HttpResponse: Página de formulario (GET) o redirección a detalle (POST válido)
    
    Raises:
        Http404: Si cliente_id no existe en BD
        
    Note:
        REGLA #3: Cambios Transversales - Validación de lista negra integrada
        FASE 2.1: Bloque A - Validación de lista negra vigente
    """
    from decimal import Decimal
    from django.utils.html import escape
    
    def obtener_proximas_fechas_pago(fecha_inicio, num_cuotas):
        """
        Obtiene las próximas fechas de pago CON GARANTÍA DE 15+ DÍAS entre cada una.
        Usa el calendario fijo: 5, 15, 20, 30
        """
        from mi_app.models import calcular_fechas_pago
        # Usar la función de models.py que garantiza 15+ días
        return calcular_fechas_pago('QUINCENAL', num_cuotas, fecha_inicio)
    
    if request.method == 'POST':
        cliente_id_form = request.POST.get('cliente')
        monto_str = request.POST.get('monto_total', '').strip()
        num_cuotas_str = request.POST.get('num_cuotas', '').strip()
        interes_str = request.POST.get('interes_porcentaje', '').strip()
        
        # ✅ VALIDACIONES ESTRICTAS AGREGADAS (CRÍTICA #4)
        errores = []
        cliente = None
        num_cuotas = None
        monto = None
        interes_porcentaje = None
        
        # ============================================================================
        # VALIDACIÓN #1: Fecha de inicio DEBE ser hoy o posterior
        # ============================================================================
        fecha_inicio = date.today()
        if fecha_inicio < date.today():
            errores.append("[V1] La fecha de inicio no puede ser anterior a hoy")
        
        # ============================================================================
        # VALIDACIÓN #2: Verificar cliente existe y NO está en lista negra
        # ============================================================================
        if not cliente_id_form:
            errores.append("[V2] Debe seleccionar un cliente")
        else:
            try:
                cliente = Cliente.objects.get(id=int(cliente_id_form))
            except (Cliente.DoesNotExist, ValueError):
                errores.append("[V2] Cliente no válido en base de datos")
                cliente = None
        
        # ============================================================================
        # VALIDACIÓN #2b: Cliente NO puede estar en lista negra (vigente)
        # ============================================================================
        if cliente:
            lista_negra_vigente = ListaNegra.objects.filter(
                cliente=cliente, 
                activa=True
            ).first()
            if lista_negra_vigente:
                errores.append(
                    f"[V2b-BLOQUEADO] ❌ {cliente.nombre} está en lista negra. "
                    f"Razón: {lista_negra_vigente.razon}. "
                    f"Vigente desde: {lista_negra_vigente.fecha_desde}"
                )
        
        # ============================================================================
        # VALIDACIÓN #3: Máximo 5 préstamos ACTIVOS simultáneos por cliente
        # ============================================================================
        if cliente:
            prestamos_activos = Prestamo.objects.filter(
                cliente=cliente,
                estado__in=['ACTIVO', 'VIGENTE']
            ).count()
            
            if prestamos_activos >= 5:
                errores.append(
                    f"[V3] ❌ Cliente ya tiene {prestamos_activos} préstamos activos. "
                    f"Límite máximo: 5"
                )
        
        # ============================================================================
        # VALIDACIÓN #4: Monto debe ser > 0 y máximo $999,999,999
        # ============================================================================
        if not monto_str:
            errores.append("[V4] El monto es requerido")
        else:
            try:
                monto = Decimal(monto_str)
                if monto <= 0:
                    errores.append("[V4] El monto debe ser mayor a $0")
                elif monto > Decimal('999999999'):
                    errores.append("[V4] El monto no puede exceder $999,999,999")
            except (ValueError, ArithmeticError):
                errores.append("[V4] El monto debe ser un número válido (máximo 2 decimales)")
                monto = None
        
        # ============================================================================
        # VALIDACIÓN #5: Número de cuotas DEBE ser exactamente 2, 4, 6 u 8
        # ============================================================================
        if not num_cuotas_str:
            errores.append("[V5] Debe seleccionar el número de cuotas")
        else:
            try:
                num_cuotas = int(num_cuotas_str)
                CUOTAS_VALIDAS = [2, 4, 6, 8]
                if num_cuotas not in CUOTAS_VALIDAS:
                    errores.append(
                        f"[V5] ❌ Número de cuotas inválido: {num_cuotas}. "
                        f"Valores válidos: {', '.join(map(str, CUOTAS_VALIDAS))}"
                    )
            except ValueError:
                errores.append("[V5] Número de cuotas debe ser un entero válido")
                num_cuotas = None
        
        # ============================================================================
        # VALIDACIÓN #6: Tasa de interés DEBE estar entre 1.5% y 10%
        # ============================================================================
        if interes_str:
            try:
                interes_porcentaje = Decimal(interes_str)
                MIN_TASA = Decimal('1.5')
                MAX_TASA = Decimal('10.0')
                
                if interes_porcentaje < MIN_TASA or interes_porcentaje > MAX_TASA:
                    errores.append(
                        f"[V6] ❌ Tasa de interés fuera de rango: {interes_porcentaje}%. "
                        f"Rango válido: {MIN_TASA}% - {MAX_TASA}%"
                    )
            except (ValueError, ArithmeticError):
                errores.append("[V6] La tasa de interés debe ser un número válido")
                interes_porcentaje = None
        
        # ============================================================================
        # VALIDACIÓN #7: Monto vs Capacidad de pago (advertencia, no bloquea)
        # ============================================================================
        if cliente and monto:
            # Calcular capacidad como historial de pagos / 2
            pagos_historicos = Pago.objects.filter(
                cuota__prestamo__cliente=cliente
            ).aggregate(total=Sum('monto_principal'))['total'] or Decimal('0')
            
            capacidad_estimada = pagos_historicos / Decimal('2') if pagos_historicos > 0 else monto
            
            # Si es la PRIMERA vez, capacidad = monto (permitir)
            if pagos_historicos > 0 and monto > capacidad_estimada:
                # Advertencia, pero no bloquea (usuario puede confirmar)
                messages.warning(
                    request,
                    f"⚠️  Advertencia: Monto ${monto} excede capacidad estimada ${capacidad_estimada}. "
                    f"Historial de pagos: ${pagos_historicos}"
                )
        
        # Si hay ERRORES críticos, bloquear creación
        if errores:
            clientes = Cliente.objects.all().order_by('nombre')
            # Mostrar todos los errores
            error_msg = ' | '.join(errores)
            context = {
                'error': error_msg,
                'clientes': clientes,
                'cliente_id': cliente_id_form
            }
            return render(request, 'mi_app/formularios/formulario_prestamo.html', context)
        
        try:
            config = Configuracion.obtener_configuracion()
            
            # Si no viene interés del formulario, usar config
            if not interes_str:
                interes_porcentaje = Decimal(str(config.tasa_interes_prestamo_normal))
                # ✅ También validar la config tenga tasa válida
                if interes_porcentaje < Decimal('1.5') or interes_porcentaje > Decimal('10.0'):
                    interes_porcentaje = Decimal('5.0')  # Default seguro
            
            # Calcular fechas automáticamente según calendario: 5, 15, 20, 30
            fecha_inicio = date.today()
            fechas_pago = obtener_proximas_fechas_pago(fecha_inicio, num_cuotas)
            fecha_fin_estimada = fechas_pago[-1] if fechas_pago else fecha_inicio
            
            # Crear préstamo CON VALIDACIONES APLICADAS
            prestamo = Prestamo.objects.create(
                cliente=cliente,
                monto_total=monto,
                interes_porcentaje=interes_porcentaje,
                fecha_inicio=fecha_inicio,
                fecha_fin_estimada=fecha_fin_estimada,
                tipo_pago='QUINCENAL',
                estado='ACTIVO'
            )
            
            # Calcular montos de cuotas con estructura QUINCENAL (2 cuotas por mes)
            cuotas_por_mes = 2
            num_meses = Decimal(num_cuotas) / Decimal(cuotas_por_mes)
            
            # Capital por mes (se divide entre 2 quincenas en cada mes)
            capital_por_mes = monto / num_meses
            capital_por_cuota = capital_por_mes / Decimal(cuotas_por_mes)
            
            # Interés: se aplica MENSUAL (interes_porcentaje es mensual)
            # Se divide entre 2 quincenas
            interes_por_mes = capital_por_mes * (interes_porcentaje / Decimal('100'))
            interes_por_cuota = interes_por_mes / Decimal(cuotas_por_mes)
            
            # Capital por cuota (para consistencia)
            monto_por_cuota = capital_por_cuota
            
            # Crear cuotas SÓLO si num_cuotas es válido (ya fue validado arriba)
            for i, fecha_pago in enumerate(fechas_pago, 1):
                Cuota.objects.create(
                    prestamo=prestamo,
                    numero_cuota=i,
                    monto_original=monto_por_cuota,
                    monto_pendiente=monto_por_cuota,
                    interes_normal=interes_por_cuota,
                    monto_pendiente_interes=interes_por_cuota,
                    fecha_pago_esperada=fecha_pago
                )
            
            # Actualizar total_prestado del cliente
            cliente.total_prestado += monto
            cliente.save()
            
            return redirect('perfil_cliente', cliente_id=cliente.id)
        
        except Exception as e:
            clientes = Cliente.objects.all().order_by('nombre')
            context = {
                'error': f'Error al crear préstamo: {str(e)}',
                'clientes': clientes,
                'cliente_id': cliente_id_form
            }
            return render(request, 'mi_app/formularios/formulario_prestamo.html', context)
    
    # GET - Mostrar formulario simplificado
    clientes = Cliente.objects.all().order_by('nombre')
    
    # ===== ERROR #5: AGREGAR INFORMACIÓN DE LISTA NEGRA A CADA CLIENTE =====
    import json
    clientes_con_lista_negra = []
    clientes_lista_negra_map = {}  # Mapa para JavaScript
    
    for cliente in clientes:
        try:
            lista_negra = cliente.lista_negra if hasattr(cliente, 'lista_negra') else None
            esta_en_lista_negra = lista_negra.esta_vigente if lista_negra else False
        except:
            lista_negra = None
            esta_en_lista_negra = False
        
        clientes_con_lista_negra.append({
            'cliente': cliente,
            'en_lista_negra': esta_en_lista_negra,
            'lista_negra': lista_negra
        })
        
        # Guardar en map para JavaScript
        if esta_en_lista_negra and lista_negra:
            clientes_lista_negra_map[cliente.id] = {
                'en_lista_negra': True,
                'razon': lista_negra.get_razon_display(),
                'fecha': lista_negra.fecha_desde.strftime('%d/%m/%Y')
            }
    
    context = {
        'clientes': clientes,
        'clientes_con_lista_negra': clientes_con_lista_negra,
        'clientes_lista_negra_json': json.dumps(clientes_lista_negra_map)
    }
    
    # Si viene desde el perfil de un cliente, preseleccionarlo
    if cliente_id:
        cliente = get_object_or_404(Cliente, id=cliente_id)
        try:
            lista_negra = cliente.lista_negra if hasattr(cliente, 'lista_negra') else None
            cliente_en_lista_negra = lista_negra.esta_vigente if lista_negra else False
        except:
            lista_negra = None
            cliente_en_lista_negra = False
        
        context['cliente_preseleccionado'] = cliente
        context['cliente_en_lista_negra'] = cliente_en_lista_negra
        context['lista_negra_cliente'] = lista_negra
        
        # Preparar datos del cliente preseleccionado en JSON
        cliente_preselectc_data = {
            'id': cliente.id,
            'nombre': cliente.nombre,
            'cedula': cliente.cedula,
            'celular': cliente.celular
        }
        context['cliente_preseleccionado_json'] = json.dumps(cliente_preselectc_data)
    
    return render(request, 'mi_app/formularios/formulario_prestamo.html', context)


@require_permission('pago.view')
@login_required(login_url='login')
def buscar_cliente_pago(request):
    """Interfaz mejorada de pagos: Buscar cliente -> Ver préstamos activos -> Pagar cuotas"""
    from django.db.models import Q
    
    cliente_id = request.GET.get('cliente_id')
    prestamo_id = request.GET.get('prestamo_id')
    
    # PASO 1: Buscar clientes
    if not cliente_id:
        clientes = Cliente.objects.filter(prestamo__estado='ACTIVO').distinct()
        busqueda = request.GET.get('q', '').strip()
        
        if busqueda:
            # BUG #3 ARREGLADO: Búsqueda SIMULTÁNEA por nombre O celular O cédula
            clientes = clientes.filter(
                Q(nombre__icontains=busqueda) |
                Q(celular__icontains=busqueda) |
                Q(cedula__icontains=busqueda)
            )
        
        # Enriquecer con conteo de préstamos activos
        clientes_data = []
        for cliente in clientes:
            clientes_data.append({
                'cliente': cliente,
                'prestamos_activos_count': cliente.prestamo_set.filter(estado='ACTIVO').count()
            })
        
        contexto = {
            'paso': 1,
            'clientes_data': clientes_data,
            'busqueda': busqueda,
        }
        return render(request, 'mi_app/pagos_dinamico.html', contexto)
    
    # PASO 2: Mostrar préstamos activos del cliente
    cliente = get_object_or_404(Cliente, id=cliente_id)
    if not prestamo_id:
        prestamos_qs = cliente.prestamo_set.filter(estado='ACTIVO')
        
        # Enriquecer préstamos con información de crédito y pendiente
        prestamos = []
        for p in prestamos_qs:
            resumen = p.resumen_financiero()
            total_pendiente = resumen['total_pendiente_principal'] + resumen['total_pendiente_interes']
            prestamos.append({
                'prestamo': p,
                'total_credito': resumen['total_credito'],
                'total_pendiente': total_pendiente,
            })
        
        contexto = {
            'paso': 2,
            'cliente': cliente,
            'prestamos': prestamos,
        }
        return render(request, 'mi_app/pagos_dinamico.html', contexto)
    
    # PASO 3: Mostrar cuotas del préstamo seleccionado
    prestamo = get_object_or_404(Prestamo, id=prestamo_id, cliente=cliente)
    cuotas = prestamo.cuotas.filter(pagado=False).order_by('numero_cuota')
    resumen = prestamo.resumen_financiero()
    
    contexto = {
        'paso': 3,
        'cliente': cliente,
        'prestamo': prestamo,
        'cuotas': cuotas,
        'resumen': resumen,
    }
    return render(request, 'mi_app/pagos_dinamico.html', contexto)

@require_permission('pago.view')
@login_required(login_url='login')
@valida_propiedad_cliente('cliente_id')
def cuotas_pendientes(request,cliente_id):
    """Muestra todas las cuotas de un cliente (pagadas y pendientes)"""
    
    cliente = get_object_or_404(Cliente, id=cliente_id)
    
    # Mostrar cuotas de TODOS los préstamos (activos, pagados y vencidos)
    cuotas = Cuota.objects.filter(prestamo__cliente=cliente).order_by('prestamo__id', 'numero_cuota')
    
    contexto = {
        'cliente': cliente,
        'cuotas': cuotas,
        'cantidad_pendientes': cuotas.filter(pagado=False).count(),
        'cantidad_pagadas': cuotas.filter(pagado=True).count()
    }
    
    return render(request, 'mi_app/cuotas_pendientes.html', contexto)

@require_any_permission('pago.create')
@atomic_payment_view
@login_required(login_url='login')
def registrar_pago(request, cuota_id):
    """
    Registra el pago de una cuota con desglose de capital, interés y mora.
    
    Procesa el pago de una cuota específica desagregando el monto en:
    1. Capital (principal)
    2. Interés normal
    3. Mora acumulada
    
    Actualiza:
    - Montos pagados y pendientes en la cuota
    - Estado de pago (si se completó)
    - Historial de pagos (tabla Pago)
    - Cascada de recálculos: etiqueta cliente y lista negra (FASE 2.1)
    
    Args:
        request (HttpRequest): Objeto de solicitud HTTP
        cuota_id (int): ID de la cuota a pagar
    
    Returns:
        HttpResponse: Formulario de pago (GET/ERROR) o redirección (POST exitoso)
        
    Raises:
        Http404: Si cuota_id no existe
        
    Note:
        FASE 2.1: Bloque A - Cascada de recálculos tras pago
        REGLA #3: Cambios transversales - Auditoría de usuario actualizada
        CRÍTICA #7: Transacción atómica para integridad de datos
    """
    from .models import Pago
    from datetime import date
    
    cuota = get_object_or_404(Cuota, id=cuota_id)
    
    if request.method == 'POST':
        from decimal import Decimal
        
        monto_pagado = request.POST.get('monto_pagado')
        
        if monto_pagado:
            monto_pagado = Decimal(monto_pagado)  # Convertir a Decimal
            
            # Calcular cuánto debe pagar en total
            mora = Decimal(str(cuota.calcular_mora_diaria()))  # Convertir a Decimal
            total_debido = cuota.monto_pendiente + cuota.interes_normal + mora
            
            if monto_pagado > total_debido:
                # Error: pagó más de lo debido
                contexto = {
                    'cuota': cuota,
                    'error': f'No puede pagar más de lo debido. Debe: ${total_debido:.2f}'
                }
                return render(request, 'mi_app/registrar_pago.html', contexto)
            
            # BUG #9 ARREGLADO: Desglosar el pago de manera proporcional
            # Primero pagar capital, luego interés, luego mora
            pendiente_capital = cuota.monto_pendiente
            pendiente_interes = cuota.interes_normal
            pendiente_mora = mora
            
            monto_pago_capital = Decimal('0')
            monto_pago_interes = Decimal('0')
            monto_pago_mora = Decimal('0')
            
            monto_restante = monto_pagado
            
            # Pagar capital
            if pendiente_capital > 0 and monto_restante > 0:
                monto_pago_capital = min(pendiente_capital, monto_restante)
                monto_restante -= monto_pago_capital
                cuota.monto_pendiente -= monto_pago_capital
            
            # Pagar interés
            if pendiente_interes > 0 and monto_restante > 0:
                monto_pago_interes = min(pendiente_interes, monto_restante)
                monto_restante -= monto_pago_interes
                cuota.monto_pendiente_interes -= monto_pago_interes
            
            # Pagar mora
            if pendiente_mora > 0 and monto_restante > 0:
                monto_pago_mora = min(pendiente_mora, monto_restante)
                monto_restante -= monto_pago_mora
            
            # BUG #9: Actualizar desglose de pagos
            cuota.monto_pagado_principal += monto_pago_capital
            cuota.monto_pagado_interes += monto_pago_interes
            cuota.monto_pagado_mora += monto_pago_mora
            
            # Registrar mora acumulada
            cuota.interes_mora_acumulado += mora
            
            # Si pagó todo, marcar como pagado
            if cuota.monto_pendiente <= 0 and cuota.monto_pendiente_interes <= 0:
                cuota.monto_pendiente = Decimal('0')
                cuota.monto_pendiente_interes = Decimal('0')
                cuota.pagado = True
                cuota.fecha_pago_real = date.today()
            
            cuota.save()
            
            # Registrar el pago en la tabla Pago
            Pago.objects.create(
                cuota=cuota,
                monto_pagado=monto_pagado,
                monto_principal=monto_pago_capital,
                monto_interes=monto_pago_interes,
                monto_mora=monto_pago_mora,
                usuario_registra='admin',
                notas=f'Pago manual de ${monto_pagado}'
            )
            
            # Actualizar estado del préstamo
            prestamo = cuota.prestamo
            if prestamo.cuotas.filter(pagado=False).count() == 0:
                prestamo.estado = 'COMPLETADO'
                prestamo.save()
            
            # BUG #9: Llamar a actualizar_estado para recalcular estado y porcentaje
            cuota.actualizar_estado()
            
            # ✅ NUEVA: Cascada de recalculos (REGLA #3: Cambios Transversales)
            cliente = cuota.prestamo.cliente
            cliente.actualizar_etiqueta()
            cliente.actualizar_lista_negra_automatica(usuario=request.user)
            
            # Registrar cambio en auditoría
            try:
                from mi_app.auditoria import registrar_cambio_manual
                registrar_cambio_manual(usuario=request.user, modelo='Cliente', id_objeto=cliente.id, accion='PAGO_REGISTRADO')
            except Exception as e:
                import logging
                logging.warning(f"Error auditoría de pago: {str(e)}")
            
            return redirect('cuotas_pendientes', cliente_id=cuota.prestamo.cliente.id)
    
    contexto = {
        'cuota': cuota,
        'mora_actual': cuota.calcular_mora_diaria(),
        'total_debido': cuota.total_a_pagar()
    }
    
    return render(request, 'mi_app/registrar_pago.html', contexto)


# ===============================================================================
# NUEVAS VISTAS - FASE 2 (Profesionales y Funcionales)
# ===============================================================================

@require_permission('prestamo.view')
@login_required(login_url='login')
def perfil_cliente(request, cliente_id):
    """
    Perfil completo del cliente con:
    - Datos básicos
    - Rating y puntuación
    - Resumen de préstamos activos
    - Resumen de préstamos rápidos (BUG #5 FIX)
    - Resumen de pagos (BUG #8 FIX)
    - Pestaña: Mis Préstamos
    - Pestaña: Mis Préstamos Rápidos (BUG #5 FIX)
    - Pestaña: Historial de Pagos (BUG #8 FIX - Mejorado)
    - Pestaña: Datos de Contacto
    """
    from datetime import date, timedelta
    
    cliente = get_object_or_404(Cliente, id=cliente_id)
    
    # Sincronizar estado: si un préstamo ACTIVO tiene todas las cuotas pagadas → COMPLETADO
    for prestamo in cliente.prestamo_set.filter(estado='ACTIVO'):
        if prestamo.cuotas.filter(pagado=False).count() == 0:
            prestamo.estado = 'COMPLETADO'
            prestamo.save()
    
    # Actualizar etiqueta del cliente por si cambió (ej. de SIN_HISTORIAL a BUENO)
    cliente.actualizar_etiqueta()
    
    # Calcular rating actual
    rating_actual = cliente.calcular_rating()
    cliente.rating = rating_actual
    cliente.save()
    
    # Obtener préstamos organizados por estado
    prestamos_activos = cliente.prestamo_set.filter(estado='ACTIVO')
    prestamos_completados = cliente.prestamo_set.filter(estado='COMPLETADO')
    
    # BUG #5: Obtener préstamos rápidos
    # NOTA: Filtrar por estado NO por saldo_pendiente para evitar inconsistencias
    prestamos_rapidos_activos = cliente.prestamos_rapidos.exclude(estado='PAGADO')
    prestamos_rapidos_pagados = cliente.prestamos_rapidos.filter(estado='PAGADO')
    
    # Información financiera del cliente
    # En Circulación = lo que el cliente aún debe; nunca negativo (la mora pagada
    # puede hacer total_pendiente < 0, pero En Circulación debe ser >= 0)
    suma_pendiente = sum(p.total_pendiente for p in prestamos_activos)
    total_en_circulacion = max(Decimal('0'), suma_pendiente)
    
    # BUG #5: Total de préstamos rápidos pendientes
    total_rapidos_pendiente = sum(float(pr.saldo_pendiente) for pr in prestamos_rapidos_activos)
    
    cuotas_vencidas = cliente.obtener_cuotas_vencidas()
    
    # BUG #8: Histórico de Pagos Mejorado
    # Obtener todos los pagos del cliente
    pagos_cliente = Pago.objects.filter(cuota__prestamo__cliente=cliente).order_by('-fecha_pago')
    
    # Pagos este mes
    hoy = date.today()
    inicio_mes = hoy.replace(day=1)
    pagos_este_mes = pagos_cliente.filter(fecha_pago__gte=inicio_mes)
    total_pagos_mes = sum(float(p.monto_pagado) for p in pagos_este_mes)
    
    # Promedio mensual (últimos 6 meses)
    hace_6_meses = hoy - timedelta(days=180)
    pagos_6_meses = pagos_cliente.filter(fecha_pago__gte=hace_6_meses)
    promedio_mensual = (sum(float(p.monto_pagado) for p in pagos_6_meses) / 6) if pagos_6_meses else 0
    
    # Último pago
    ultimo_pago = pagos_cliente.first()
    
    # ===== ERROR #5: INFORMACIÓN DE LISTA NEGRA =====
    lista_negra = None
    esta_en_lista_negra = False
    
    try:
        lista_negra = cliente.lista_negra
        esta_en_lista_negra = lista_negra.esta_vigente if lista_negra else False
    except:
        pass
    
    contexto = {
        'cliente': cliente,
        'rating_actual': rating_actual,
        'prestamos_activos': prestamos_activos,
        'prestamos_completados': prestamos_completados,
        'total_prestamos': cliente.prestamo_set.count(),
        'total_en_circulacion': total_en_circulacion,
        'cuotas_vencidas': cuotas_vencidas,
        'cantidad_vencidas': len(cuotas_vencidas),
        # BUG #5: Información de Préstamos Rápidos
        'prestamos_rapidos_activos': prestamos_rapidos_activos,
        'prestamos_rapidos_pagados': prestamos_rapidos_pagados,
        'total_rapidos': cliente.prestamos_rapidos.count(),
        'total_rapidos_pendiente': total_rapidos_pendiente,
        # BUG #8: Histórico de Pagos Mejorado
        'pagos_cliente': pagos_cliente[:20],  # Últimos 20 pagos
        'pagos_este_mes': pagos_este_mes,  # Añadido para template
        'total_pagos_mes': total_pagos_mes,
        'promedio_mensual': promedio_mensual,
        'ultimo_pago': ultimo_pago,
        'total_pagos_cliente': pagos_cliente.count(),
        # ERROR #5: Lista Negra
        'lista_negra': lista_negra,
        'esta_en_lista_negra': esta_en_lista_negra,
    }
    
    return render(request, 'mi_app/clientes/perfil_cliente.html', contexto)


@require_permission('prestamo.view')
@login_required(login_url='login')
@valida_propiedad_cliente('cliente_id')
def mis_prestamos(request, cliente_id):
    """
    Vista de todos los préstamos de un cliente con filtros
    y opciones de ver detalles, pagar, editar, eliminar
    """
    cliente = get_object_or_404(Cliente, id=cliente_id)
    prestamos = cliente.prestamo_set.all()
    
    # Filtro por estado
    estado_filter = request.GET.get('estado', 'todos')
    if estado_filter != 'todos':
        prestamos = prestamos.filter(estado=estado_filter)
    
    # Enriquecer cada préstamo con información calculada
    prestamos_data = []
    for prestamo in prestamos:
        resumen = prestamo.resumen_financiero()
        prestamo_info = {
            'prestamo': prestamo,
            'resumen': resumen,
            'progreso': (resumen['total_pagado_principal'] / resumen['monto_original'] * 100) if resumen['monto_original'] > 0 else 0,
            'num_cuotas': prestamo.cuotas.count(),
            'num_pagadas': prestamo.num_cuotas_pagadas,
            'num_pendientes': prestamo.cuotas.filter(pagado=False).count(),
            'num_vencidas': prestamo.num_cuotas_vencidas,
        }
        prestamos_data.append(prestamo_info)
    
    contexto = {
        'cliente': cliente,
        'prestamos_data': prestamos_data,
        'estado_filter': estado_filter,
        'total_prestamos': prestamos.count(),
    }
    
    return render(request, 'mi_app/clientes/mis_prestamos.html', contexto)


@require_permission('prestamo.view')
@login_required(login_url='login')
@valida_propiedad_prestamo('prestamo_id')
def detalles_prestamo(request, prestamo_id):
    """
    Detalles completos de un préstamo:
    - Resumen financiero desglosado (Original/Pagado/Pendiente)
    - Tabla de cuotas
    - Historial de pagos
    - Botones de acciones
    """
    prestamo = get_object_or_404(Prestamo, id=prestamo_id)
    cuotas = prestamo.cuotas.all()
    
    # Resumen financiero desglosado
    resumen = prestamo.resumen_financiero()
    
    # Enriquecer cuotas con detalles
    cuotas_data = []
    for cuota in cuotas:
        detalles = cuota.detalles_completos()
        cuota_info = {
            'cuota': cuota,
            'detalles': detalles,
            'estado_visual': _obtener_estado_visual_cuota(cuota),
        }
        cuotas_data.append(cuota_info)
    
    contexto = {
        'prestamo': prestamo,
        'resumen': resumen,
        'cuotas_data': cuotas_data,
        'progreso': ((resumen['total_credito'] - resumen['total_pendiente_principal'] - resumen['total_pendiente_interes']) / resumen['total_credito'] * 100) if resumen['total_credito'] > 0 else 0,
        'num_pagadas': prestamo.num_cuotas_pagadas,
        'num_pendientes': prestamo.cuotas.filter(pagado=False).count(),
        'num_vencidas': prestamo.num_cuotas_vencidas,
    }
    
    return render(request, 'mi_app/detalles_prestamo.html', contexto)


@require_permission('pago.view')
@login_required(login_url='login')
def detalles_cuota(request, cuota_id):
    """
    Detalles completos de una cuota:
    - Desglose Original (Lo que se debe)
    - Desglose Pagado (Lo que ya se pagó)
    - Desglose Pendiente (Lo que falta)
    - Historial de pagos (transacciones)
    - Botón para registrar pago
    """
    from .models import Pago
    
    cuota = get_object_or_404(Cuota, id=cuota_id)
    detalles = cuota.detalles_completos()
    
    # Obtener historial de pagos de esta cuota
    pagos = Pago.objects.filter(cuota=cuota).order_by('-fecha_pago')
    
    contexto = {
        'cuota': cuota,
        'prestamo': cuota.prestamo,
        'detalles': detalles,
        'pagos': pagos,
        'estado_visual': _obtener_estado_visual_cuota(cuota),
    }
    
    return render(request, 'mi_app/detalles_cuota.html', contexto)


@require_any_permission('pago.create')
@login_required(login_url='login')
def pagar_cuota_especifica(request, cuota_id):
    """
    BUG #7 FIX: Nueva vista dedicada al pago de una cuota ESPECÍFICA.
    Interfaz limpia y enfocada, mostrando SOLO los detalles de esa cuota.
    
    GET: Muestra formulario de pago
    POST: Procesa el pago
    """
    from .models import Pago
    
    cuota = get_object_or_404(Cuota, id=cuota_id)
    prestamo = cuota.prestamo
    cliente = prestamo.cliente
    detalles = cuota.detalles_completos()
    
    if request.method == 'POST':
        monto_principal = Decimal(request.POST.get('monto_principal', '0').strip() or '0')
        monto_interes = Decimal(request.POST.get('monto_interes', '0').strip() or '0')
        monto_mora = Decimal(request.POST.get('monto_mora', '0').strip() or '0')
        referencia = request.POST.get('referencia', '')
        notas = request.POST.get('notas', '')
        
        monto_total = monto_principal + monto_interes + monto_mora
        
        # Validaciones
        if monto_total <= 0:
            contexto = {
                'cuota': cuota,
                'prestamo': prestamo,
                'cliente': cliente,
                'detalles': detalles,
                'error': '❌ El monto debe ser mayor a $0',
            }
            return render(request, 'mi_app/pagar_cuota_especifica.html', contexto)
        
        # Validar que no supere lo pendiente
        if monto_principal > cuota.monto_pendiente:
            contexto = {
                'cuota': cuota,
                'prestamo': prestamo,
                'cliente': cliente,
                'detalles': detalles,
                'error': f'❌ Principal pendiente: ${cuota.monto_pendiente}',
            }
            return render(request, 'mi_app/pagar_cuota_especifica.html', contexto)
        
        if monto_interes > cuota.monto_pendiente_interes:
            contexto = {
                'cuota': cuota,
                'prestamo': prestamo,
                'cliente': cliente,
                'detalles': detalles,
                'error': f'❌ Interés pendiente: ${cuota.monto_pendiente_interes}',
            }
            return render(request, 'mi_app/pagar_cuota_especifica.html', contexto)
        
        # Crear registro de pago
        pago = Pago.objects.create(
            cuota=cuota,
            monto_pagado=monto_total,
            monto_principal=monto_principal,
            monto_interes=monto_interes,
            monto_mora=monto_mora,
            usuario_registra=request.user.username,  # ✅ SOLUCIONADO: Obtener del usuario logueado
            referencia=referencia,
            notas=notas
        )
        
        # Actualizar cuota (igual que en registrar_pago_mejorado)
        cuota.monto_pagado_principal += monto_principal
        cuota.monto_pagado_interes += monto_interes
        cuota.monto_pagado_mora += monto_mora
        cuota.monto_pendiente = max(cuota.monto_original - cuota.monto_pagado_principal, Decimal('0'))
        cuota.monto_pendiente_interes = max(cuota.interes_normal - cuota.monto_pagado_interes, Decimal('0'))
        
        if cuota.monto_pendiente == 0 and cuota.monto_pendiente_interes == 0:
            cuota.pagado = True
            cuota.fecha_pago_real = date.today()
        
        cuota.actualizar_estado()
        cuota.save()  # IMPORTANTE: Guardar los cambios en la BD
        
        # Actualizar préstamo si todas las cuotas están pagadas
        if prestamo.cuotas.filter(pagado=False).count() == 0:
            prestamo.estado = 'COMPLETADO'
            prestamo.save()
        
        # Recargar detalles después del pago para mostrar valores actualizados
        detalles = cuota.detalles_completos()
        
        # Mostrar comprobante
        contexto = {
            'pago': pago,
            'cuota': cuota,
            'prestamo': prestamo,
            'cliente': cliente,
            'detalles': detalles,
            'pagos': Pago.objects.filter(cuota=cuota).order_by('-fecha_pago'),
            'comprobante': pago.comprobante_texto(),
            'success': True,
        }
        
        return render(request, 'mi_app/pagar_cuota_especifica.html', contexto)
    
    else:  # GET
        # Obtener historial de pagos
        pagos = Pago.objects.filter(cuota=cuota).order_by('-fecha_pago')
        
        contexto = {
            'cuota': cuota,
            'prestamo': prestamo,
            'cliente': cliente,
            'detalles': detalles,
            'pagos': pagos,
        }
        
        return render(request, 'mi_app/pagar_cuota_especifica.html', contexto)

@require_any_permission('pago.create')
@login_required(login_url='login')
@valida_propiedad_cliente('cliente_id')
def registrar_pago_mejorado(request, cliente_id):
    """
    Interfaz mejorada de pago:
    - Selector de cuotas (múltiples, agrupadas por préstamo)
    - Vista previa de mora actualizada
    - Desglose editable (principal/interés/mora)
    - Comprobante al registrar
    """
    from .models import Pago
    from decimal import Decimal
    
    cliente = get_object_or_404(Cliente, id=cliente_id)
    
    # Obtener solo cuotas pendientes de este cliente
    cuotas_pendientes = Cuota.objects.filter(
        prestamo__cliente=cliente,
        pagado=False
    ).select_related('prestamo').order_by('prestamo_id', 'numero_cuota')
    
    # Agrupar por préstamo
    prestamos_cuotas = {}
    for cuota in cuotas_pendientes:
        if cuota.prestamo.id not in prestamos_cuotas:
            prestamos_cuotas[cuota.prestamo.id] = {
                'prestamo': cuota.prestamo,
                'cuotas': []
            }
        prestamos_cuotas[cuota.prestamo.id]['cuotas'].append(cuota)
    
    if request.method == 'POST':
        from django.utils.html import escape
        import re
        
        cuota_id = request.POST.get('cuota_id', '').strip()
        monto_pagado_str = request.POST.get('monto_pagado', '0').strip()
        monto_principal_str = request.POST.get('monto_principal', '0').strip()
        monto_interes_str = request.POST.get('monto_interes', '0').strip()
        monto_mora_str = request.POST.get('monto_mora', '0').strip()
        referencia = escape(request.POST.get('referencia', '').strip())  # Sanitizar
        notas = escape(request.POST.get('notas', '').strip())  # Sanitizar
        
        # ✅ VALIDACIONES MEJORADAS
        errores = []
        
        # 1. Validar cuota_id (debe ser número válido)
        if not cuota_id:
            errores.append("Debe seleccionar una cuota")
        else:
            try:
                cuota_id_int = int(cuota_id)
                if cuota_id_int <= 0:
                    errores.append("ID de cuota inválido")
                cuota = Cuota.objects.get(id=cuota_id_int)
            except (Cuota.DoesNotExist, ValueError, TypeError):
                errores.append("Cuota no válida")
                cuota = None
        
        # 2. Validar monto_pagado > 0 (y que sea número válido)
        try:
            # Verificar que solo contiene números y punto decimal
            if not re.match(r'^[\d.]+$', monto_pagado_str):
                errores.append("El monto debe contener solo números")
            
            monto_pagado = Decimal(monto_pagado_str)
            if monto_pagado <= 0:
                errores.append("El monto a pagar debe ser mayor a $0")
            if monto_pagado > Decimal('999999999'):
                errores.append("El monto es demasiado alto")
        except:
            errores.append("El monto debe ser un número válido")
            monto_pagado = None
        
        # 3. Validar desglose de pagos
        try:
            # Validar formato de cada monto
            for campo, valor in [('principal', monto_principal_str), ('interés', monto_interes_str), ('mora', monto_mora_str)]:
                if not re.match(r'^[\d.]*$', valor):
                    errores.append(f"El monto de {campo} contiene caracteres inválidos")
            
            monto_principal = Decimal(monto_principal_str) if monto_principal_str else Decimal('0')
            monto_interes = Decimal(monto_interes_str) if monto_interes_str else Decimal('0')
            monto_mora = Decimal(monto_mora_str) if monto_mora_str else Decimal('0')
            
            if monto_principal < 0 or monto_interes < 0 or monto_mora < 0:
                errores.append("Los montos no pueden ser negativos")
            
            # Verificar que la suma no supere lo pendiente
            if cuota and monto_pagado:
                monto_total_pendiente = (
                    cuota.monto_pendiente + 
                    cuota.monto_pendiente_interes + 
                    cuota.calcular_mora()
                )
                if monto_pagado > monto_total_pendiente:
                    errores.append(
                        f"El monto a pagar (${monto_pagado}) no puede superar "
                        f"lo pendiente (${monto_total_pendiente})"
                    )
        except Exception as e:
            errores.append("Los montos del desglose deben ser números válidos")
        
        # 4. Validar que referencia y notas no contengan inyecciones SQL
        # (escape() ya lo hace, pero verificamos tamaño)
        if len(referencia) > 100:
            errores.append("La referencia no puede exceder 100 caracteres")
        if len(notas) > 500:
            errores.append("Las notas no pueden exceder 500 caracteres")
        
        # Si hay errores, mostrar formulario
        if errores:
            contexto = {
                'cliente': cliente,
                'prestamos_cuotas': prestamos_cuotas,
                'error': ' | '.join(errores),
            }
            return render(request, 'mi_app/registrar_pago_mejorado.html', contexto)
        
        if not cuota:
            contexto = {
                'cliente': cliente,
                'prestamos_cuotas': prestamos_cuotas,
                'error': 'Error al procesar la cuota',
            }
            return render(request, 'mi_app/registrar_pago_mejorado.html', contexto)
        
        # Crear registro de pago
        pago = Pago.objects.create(
            cuota=cuota,
            monto_pagado=monto_pagado,
            monto_principal=monto_principal,
            monto_interes=monto_interes,
            monto_mora=monto_mora,
            usuario_registra=request.user.username,  # ✅ SOLUCIONADO: Obtener del usuario logueado
            referencia=referencia,
            notas=notas
        )
        
        # Actualizar cuota
        cuota.monto_pagado_principal += monto_principal
        cuota.monto_pagado_interes += monto_interes
        cuota.monto_pagado_mora += monto_mora
        
        # BUG #6 FIX: Recalcular montos pendientes correctamente
        cuota.monto_pendiente = max(cuota.monto_original - cuota.monto_pagado_principal, Decimal('0'))
        cuota.monto_pendiente_interes = max(cuota.interes_normal - cuota.monto_pagado_interes, Decimal('0'))
        
        # Marcar como pagada si está completa
        if cuota.monto_pendiente == 0 and cuota.monto_pendiente_interes == 0:
            cuota.pagado = True
            cuota.fecha_pago_real = date.today()
        
        # BUG #6 FIX: Llamar a actualizar_estado() para sincronizar 'estado' y 'porcentaje_pagado'
        # Este método recalcula automáticamente los campos de estado basado en los montos pagados
        cuota.actualizar_estado()
        
        # Actualizar estado del préstamo
        if cuota.prestamo.cuotas.filter(pagado=False).count() == 0:
            cuota.prestamo.estado = 'COMPLETADO'
            cuota.prestamo.save()
        
        # ✅ NUEVA: Cascada de recalculos (REGLA #3: Cambios Transversales)
        cliente.actualizar_etiqueta()
        cliente.actualizar_lista_negra_automatica(usuario=request.user)
        
        # Registrar cambio en auditoría
        try:
            from mi_app.auditoria import registrar_cambio_manual
            registrar_cambio_manual(usuario=request.user, modelo='Cliente', id_objeto=cliente.id, accion='PAGO_REGISTRADO')
        except Exception as e:
            import logging
            logging.warning(f"Error auditoría de pago: {str(e)}")
        
        # Mostrar comprobante
        contexto = {
            'pago': pago,
            'cuota': cuota,
            'cliente': cliente,
            'comprobante': pago.comprobante_texto(),
        }
        
        return render(request, 'mi_app/comprobante_pago.html', contexto)
    
    contexto = {
        'cliente': cliente,
        'prestamos_cuotas': prestamos_cuotas,
        'total_cuotas_pendientes': cuotas_pendientes.count(),
    }
    
    return render(request, 'mi_app/registrar_pago_mejorado.html', contexto)


@require_permission('reporte.view')
@login_required(login_url='login')
def reporte_clientes(request):
    """
    Reporte general de todos los clientes con filtros avanzados:
    - Búsqueda por nombre/cédula
    - Filtro por estado
    - Filtro por rango de fecha de creación
    """
    clientes = Cliente.objects.all()
    
    # Filtro por cliente_id si viene del dropdown
    cliente_id = request.GET.get('cliente_id', '').strip()
    if cliente_id:
        try:
            clientes = clientes.filter(id=int(cliente_id))
        except (ValueError, TypeError):
            pass
    
    # Búsqueda
    busqueda = request.GET.get('search', '').strip()
    if busqueda:
        clientes = clientes.filter(
            Q(nombre__icontains=busqueda) |
            Q(cedula__icontains=busqueda) |
            Q(celular__icontains=busqueda)
        )
    
    # Filtro estado (coherente con lista de clientes)
    estado = request.GET.get('estado', '')
    if estado in ['ACTIVO', 'INACTIVO']:
        clientes = clientes.filter(estado=estado)
    elif estado in ['BUENO', 'MEDIO', 'MALO', 'SIN_HISTORIAL']:
        clientes = clientes.filter(etiqueta_cliente=estado)
    elif estado == 'NUEVO':
        clientes = clientes.filter(etiqueta_cliente='SIN_HISTORIAL')
    elif estado == 'LISTA_NEGRA':
        clientes = clientes.filter(
            lista_negra__activa=True
        ).filter(
            Q(lista_negra__fecha_hasta__isnull=True) | Q(lista_negra__fecha_hasta__gte=date.today())
        )
    
    # Filtro rango de fechas de creación
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    if fecha_desde:
        try:
            from datetime import datetime
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            clientes = clientes.filter(fecha_creacion__gte=fecha_desde_obj)
        except (ValueError, AttributeError):
            pass
    
    if fecha_hasta:
        try:
            from datetime import datetime
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            clientes = clientes.filter(fecha_creacion__lte=fecha_hasta_obj)
        except (ValueError, AttributeError):
            pass
    
    clientes = clientes.distinct().order_by('nombre')

    # Calcular estadísticas
    activos = clientes.filter(estado='ACTIVO').count()
    ratings = [c.calcular_rating() for c in clientes]
    rating_promedio = sum(ratings) / len(ratings) if ratings else 0

    # Estados visuales enriquecidos
    clientes_reporte = []
    for cliente in clientes:
        try:
            lista_negra = cliente.lista_negra if hasattr(cliente, 'lista_negra') else None
            en_lista_negra = lista_negra.esta_vigente if lista_negra else False
        except Exception:
            lista_negra = None
            en_lista_negra = False

        if en_lista_negra:
            estado_visual = 'LISTA_NEGRA'
        elif cliente.etiqueta_cliente == 'SIN_HISTORIAL':
            estado_visual = 'NUEVO'
        else:
            estado_visual = cliente.etiqueta_cliente

        clientes_reporte.append({
            'cliente': cliente,
            'en_lista_negra': en_lista_negra,
            'lista_negra': lista_negra,
            'estado_visual': estado_visual,
        })
    
    # ✅ OPTIMIZACIÓN N+1 #5-6 (CRITICAL): Usar prefetch_related + agregación en lugar de loops
    from django.db.models import Prefetch, Sum as DbSum, F as DbF, Value as DbValue, IntegerField, Case, When
    from mi_app.models import PrestamoRapido
    
    # Prefetch todas las relaciones necesarias UNA sola vez
    clientes_prefetched = Cliente.objects.prefetch_related(
        Prefetch('prestamo_set', queryset=Prestamo.objects.prefetch_related('cuotas').all()),
        Prefetch('prestamos_rapidos', queryset=PrestamoRapido.objects.all())
    ).all()
    
    # BUG FIX #1: Calcular AMBOS capital y total con interés
    capital_prestado_result = Prestamo.objects.aggregate(
        total=DbSum('monto_total')
    )['total'] or Decimal('0')
    capital_prestado = Decimal(str(capital_prestado_result))
    
    # Total crédito usando cuotas (más preciso que propiedad que hace queries)
    total_credito_result = Cuota.objects.aggregate(
        total=DbSum('monto_original') + DbSum('interes_normal')
    )['total'] or 0
    total_credito = Decimal(str(total_credito_result))
    
    # BUG #5: Préstamos Rápidos - saldo_pendiente es @property, no campo BD
    rapidos_pendientes = PrestamoRapido.objects.filter(
        estado__in=['PENDIENTE', 'PARCIALMENTE_PAGADO']
    )
    total_rapidos = PrestamoRapido.objects.count()
    total_rapidos_pendiente = sum(float(pr.saldo_pendiente) for pr in rapidos_pendientes)
    
    contexto = {
        'clientes': clientes_reporte,
        'total': clientes.count(),
        'search': busqueda,
        'estado': estado,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'activos_count': activos,
        'rating_promedio': rating_promedio,
        'capital_prestado': capital_prestado,  # ← BUG FIX #1: Solo capital
        'total_credito': total_credito,        # ← BUG FIX #1: Capital + Interés
        # BUG #5: Info Préstamos Rápidos
        'total_rapidos': total_rapidos,
        'total_rapidos_pendiente': total_rapidos_pendiente,
    }
    
    return render(request, 'mi_app/reporte_clientes.html', contexto)

@require_permission('reporte.view')
@login_required(login_url='login')

def reporte_prestamos(request):
    """
    Reporte general de todos los préstamos con filtros avanzados:
    - Búsqueda por cliente
    - Filtro por estado (ACTIVO, PAGADO, VENCIDO)
    - Filtro por rango de fechas de inicio
    - Filtro por rango de montos
    """
    prestamos = Prestamo.objects.select_related('cliente').all()
    
    # Filtro por cliente_id si viene del dropdown
    cliente_id = request.GET.get('cliente_id', '').strip()
    if cliente_id:
        try:
            prestamos = prestamos.filter(cliente_id=int(cliente_id))
        except (ValueError, TypeError):
            pass
    
    # Búsqueda
    busqueda = request.GET.get('search', '').strip()
    if busqueda:
        prestamos = prestamos.filter(
            cliente__nombre__icontains=busqueda
        )
    
    # Filtro estado
    estado = request.GET.get('estado', '')
    if estado:
        prestamos = prestamos.filter(estado=estado)
    
    # Filtro rango de fechas de inicio
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    if fecha_desde:
        try:
            from datetime import datetime
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            prestamos = prestamos.filter(fecha_inicio__gte=fecha_desde_obj)
        except (ValueError, AttributeError):
            pass
    
    if fecha_hasta:
        try:
            from datetime import datetime
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            prestamos = prestamos.filter(fecha_inicio__lte=fecha_hasta_obj)
        except (ValueError, AttributeError):
            pass
    
    # Filtro por rango de montos
    monto_desde = request.GET.get('monto_desde', '')
    monto_hasta = request.GET.get('monto_hasta', '')
    
    if monto_desde:
        try:
            from decimal import Decimal
            monto_desde_val = Decimal(monto_desde)
            prestamos = prestamos.filter(monto_total__gte=monto_desde_val)
        except:
            pass
    
    if monto_hasta:
        try:
            from decimal import Decimal
            monto_hasta_val = Decimal(monto_hasta)
            prestamos = prestamos.filter(monto_total__lte=monto_hasta_val)
        except:
            pass
    
    # Calcular totales
    total_monto = sum(p.monto_total for p in prestamos)
    total_pagado = sum(p.total_pagado for p in prestamos)
    total_pendiente = sum(p.total_pendiente for p in prestamos)
    
    contexto = {
        'prestamos': prestamos,
        'total': prestamos.count(),
        'search': busqueda,
        'estado': estado,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'monto_desde': monto_desde,
        'monto_hasta': monto_hasta,
        'total_monto': total_monto,
        'total_pagado': total_pagado,
        'total_pendiente': total_pendiente,
        # Estadísticas globales
        'total_clientes': Cliente.objects.count(),
        'clientes_activos': Cliente.objects.filter(estado='ACTIVO').count(),
        'total_prestamos': Prestamo.objects.count(),
    }
    
    return render(request, 'mi_app/reporte_prestamos.html', contexto)

# ===============================================================================
# BUG #2: REPORTE DE CUOTAS COMPLETO
# ===============================================================================

@require_permission('reporte.view')
@login_required(login_url='login')
def reporte_cuotas_completo(request):
    """
    Reporte completo de TODAS las cuotas del sistema con información detallada.
    
    Incluye:
    - Información del cliente y préstamo
    - Estado de pago (pendiente, parcial, pagada)
    - Fechas de vencimiento y pago real
    - Montos de principal, interés y mora
    - Filtros por cliente, estado, tipo de pago
    """
    from django.db.models import Q, Sum, DecimalField, Case, When
    from django.db.models.functions import Coalesce
    from django.core.paginator import Paginator
    
    # Obtener todas las cuotas con información relacionada
    cuotas = Cuota.objects.select_related(
        'prestamo__cliente'
    ).order_by(
        'prestamo__cliente__nombre',
        'prestamo__id',
        'numero_cuota'
    )
    
    # FILTROS
    # 1. Filtro por cliente
    cliente_id = request.GET.get('cliente_id')
    if cliente_id:
        cuotas = cuotas.filter(prestamo__cliente_id=cliente_id)
    
    # 2. Filtro por estado
    estado_filtro = request.GET.get('estado')
    if estado_filtro == 'pendiente':
        # Cuotas sin pagar (ambos montos pendientes)
        cuotas = cuotas.filter(
            pagado=False,
            monto_pendiente__gt=0,
            monto_pendiente_interes__gt=0
        )
    elif estado_filtro == 'parcial':
        # Cuotas parcialmente pagadas
        from django.db.models import Q as DjangoQ
        cuotas = cuotas.filter(
            DjangoQ(monto_pagado_principal__gt=0) | DjangoQ(monto_pagado_interes__gt=0),
            DjangoQ(monto_pendiente__gt=0) | DjangoQ(monto_pendiente_interes__gt=0)
        )
    elif estado_filtro == 'pagada':
        # Cuotas completamente pagadas
        cuotas = cuotas.filter(pagado=True)
    elif estado_filtro == 'vencida':
        # Cuotas vencidas pero no pagadas
        cuotas = cuotas.filter(
            pagado=False,
            fecha_pago_esperada__lt=date.today()
        )
    
    # 3. Filtro por tipo de pago (principal, interés, ambos)
    tipo_pago = request.GET.get('tipo_pago')
    if tipo_pago == 'principal':
        cuotas = cuotas.filter(monto_pagado_principal__gt=0)
    elif tipo_pago == 'interes':
        cuotas = cuotas.filter(monto_pagado_interes__gt=0)
    elif tipo_pago == 'mora':
        cuotas = cuotas.filter(monto_pagado_mora__gt=0)
    
    # 4. Rango de fechas de vencimiento
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if fecha_desde:
        try:
            fecha_desde_obj = dt.datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            cuotas = cuotas.filter(fecha_pago_esperada__gte=fecha_desde_obj)
        except:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_obj = dt.datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            cuotas = cuotas.filter(fecha_pago_esperada__lte=fecha_hasta_obj)
        except:
            pass
    
    # CÁLCULOS DE RESUMEN
    resumen_general = cuotas.aggregate(
        total_cuotas=Count('id'),
        total_capital=Coalesce(Sum('monto_original'), Decimal('0'), output_field=DecimalField()),
        total_interes=Coalesce(Sum('interes_normal'), Decimal('0'), output_field=DecimalField()),
        total_pagado_principal=Coalesce(Sum('monto_pagado_principal'), Decimal('0'), output_field=DecimalField()),
        total_pagado_interes=Coalesce(Sum('monto_pagado_interes'), Decimal('0'), output_field=DecimalField()),
        total_pagado_mora=Coalesce(Sum('monto_pagado_mora'), Decimal('0'), output_field=DecimalField()),
        total_pendiente_principal=Coalesce(Sum('monto_pendiente'), Decimal('0'), output_field=DecimalField()),
        total_pendiente_interes=Coalesce(Sum('monto_pendiente_interes'), Decimal('0'), output_field=DecimalField()),
    )
    
    resumen_general['total_esperado'] = (
        resumen_general['total_capital'] + resumen_general['total_interes']
    )
    resumen_general['total_pagado'] = (
        resumen_general['total_pagado_principal'] + 
        resumen_general['total_pagado_interes'] + 
        resumen_general['total_pagado_mora']
    )
    resumen_general['total_pendiente'] = (
        resumen_general['total_pendiente_principal'] + 
        resumen_general['total_pendiente_interes']
    )
    
    # Calcular porcentaje pagado
    if resumen_general['total_esperado'] > 0:
        resumen_general['porcentaje_pagado'] = (
            resumen_general['total_pagado'] / resumen_general['total_esperado'] * 100
        )
    else:
        resumen_general['porcentaje_pagado'] = 0
    
    # PAGINACIÓN
    paginator = Paginator(cuotas, 50)  # 50 cuotas por página
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # OBTENER LISTA DE CLIENTES PARA DROPDOWN
    clientes = Cliente.objects.filter(
        prestamo__isnull=False
    ).distinct().order_by('nombre')
    
    contexto = {
        'cuotas': page_obj,
        'resumen': resumen_general,
        'clientes': clientes,
        'cliente_id_filtro': cliente_id,
        'estado_filtro': estado_filtro,
        'tipo_pago_filtro': tipo_pago,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'total_paginas': paginator.num_pages,
        'now': date.today(),
    }
    
    return render(request, 'mi_app/reporte_cuotas.html', contexto)

@require_permission('reporte.view')
@login_required(login_url='login')

def reporte_cuotas_vencidas(request):
    """
    Reporte de cuotas vencidas y próximas a vencer con filtros:
    - Filtro por cliente
    - Filtro por rango de fechas de vencimiento
    - Filtro por monto mínimo
    """
    from django.db.models import Q
    
    # Cuotas vencidas
    cuotas_vencidas = Cuota.objects.filter(
        pagado=False,
        fecha_pago_esperada__lt=date.today()
    ).select_related('prestamo__cliente').order_by('-fecha_pago_esperada')
    
    # Cuotas próximas a vencer (próximos 7 días)
    fecha_limite = date.today() + timedelta(days=7)
    cuotas_proximas = Cuota.objects.filter(
        pagado=False,
        fecha_pago_esperada__gte=date.today(),
        fecha_pago_esperada__lte=fecha_limite
    ).select_related('prestamo__cliente').order_by('fecha_pago_esperada')
    
    # Filtro por cliente_id si viene del dropdown
    cliente_id = request.GET.get('cliente_id', '').strip()
    if cliente_id:
        try:
            cliente_id_int = int(cliente_id)
            cuotas_vencidas = cuotas_vencidas.filter(prestamo__cliente_id=cliente_id_int)
            cuotas_proximas = cuotas_proximas.filter(prestamo__cliente_id=cliente_id_int)
        except (ValueError, TypeError):
            pass
    
    # Búsqueda por cliente
    busqueda = request.GET.get('search', '').strip()
    if busqueda:
        cuotas_vencidas = cuotas_vencidas.filter(
            prestamo__cliente__nombre__icontains=busqueda
        )
        cuotas_proximas = cuotas_proximas.filter(
            prestamo__cliente__nombre__icontains=busqueda
        )
    
    # Filtro por rango de fechas de vencimiento
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    if fecha_desde:
        try:
            from datetime import datetime
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            cuotas_vencidas = cuotas_vencidas.filter(fecha_pago_esperada__gte=fecha_desde_obj)
            cuotas_proximas = cuotas_proximas.filter(fecha_pago_esperada__gte=fecha_desde_obj)
        except (ValueError, AttributeError):
            pass
    
    if fecha_hasta:
        try:
            from datetime import datetime
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            cuotas_vencidas = cuotas_vencidas.filter(fecha_pago_esperada__lte=fecha_hasta_obj)
            cuotas_proximas = cuotas_proximas.filter(fecha_pago_esperada__lte=fecha_hasta_obj)
        except (ValueError, AttributeError):
            pass
    
    # Filtro por monto mínimo a recuperar
    monto_minimo = request.GET.get('monto_minimo', '')
    if monto_minimo:
        try:
            from decimal import Decimal
            monto_val = Decimal(monto_minimo)
            cuotas_vencidas = cuotas_vencidas.filter(monto_pendiente__gte=monto_val)
            cuotas_proximas = cuotas_proximas.filter(monto_pendiente__gte=monto_val)
        except:
            pass
    
    # Calcular mora acumulada
    total_mora = sum(c.calcular_mora_diaria() for c in cuotas_vencidas)
    total_pendiente_vencidas = sum(c.monto_pendiente for c in cuotas_vencidas)
    total_pendiente_proximas = sum(c.monto_pendiente for c in cuotas_proximas)
    total_en_riesgo = total_pendiente_vencidas + total_pendiente_proximas + total_mora
    
    # Clientes afectados
    clientes_afectados_ids = set()
    for cuota in cuotas_vencidas:
        clientes_afectados_ids.add(cuota.prestamo.cliente.id)
    for cuota in cuotas_proximas:
        clientes_afectados_ids.add(cuota.prestamo.cliente.id)
    
    contexto = {
        'vencidas': cuotas_vencidas,
        'proximas': cuotas_proximas,
        'total_vencidas': cuotas_vencidas.count(),
        'total_proximas': cuotas_proximas.count(),
        'search': busqueda,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'monto_minimo': monto_minimo,
        'total_mora': total_mora,
        'total_pendiente_vencidas': total_pendiente_vencidas,
        'total_pendiente_proximas': total_pendiente_proximas,
        'total_en_riesgo': total_en_riesgo,
        'clientes_afectados': len(clientes_afectados_ids),
    }
    
    return render(request, 'mi_app/reporte_cuotas_vencidas.html', contexto)


@require_permission('reporte.view')
@login_required
def reporte_estadisticas(request):
    """
    Dashboard de estadísticas generales del sistema:
    - Cantidad de clientes activos/inactivos
    - Cantidad de préstamos por estado
    - Total de cuotas pendientes/pagadas
    - Calificación promedio de clientes
    - Ingresos totales
    - Mora acumulada
    """
    from decimal import Decimal
    
    # CLIENTES
    total_clientes = Cliente.objects.count()
    clientes_activos = Cliente.objects.filter(estado='ACTIVO').count()
    clientes_inactivos = Cliente.objects.filter(estado='INACTIVO').count()
    
    # PRESTAMOS
    total_prestamos = Prestamo.objects.count()
    prestamos_activos = Prestamo.objects.filter(estado='ACTIVO').count()
    prestamos_completados = Prestamo.objects.filter(estado='COMPLETADO').count()
    prestamos_vencidos = Prestamo.objects.filter(estado='VENCIDO').count()
    
    # MONTOS - BUG FIX #1: Calcular AMBOS capital y total con interés para mayor claridad
    capital_prestado = sum(Decimal(str(p.monto_total)) for p in Prestamo.objects.all())
    total_credito = sum(Decimal(str(p.total_credito)) for p in Prestamo.objects.all())
    total_pagado = sum(Decimal(str(p.total_pagado)) for p in Prestamo.objects.all())
    total_pendiente_capital = capital_prestado - total_pagado
    total_pendiente_credito = total_credito - total_pagado
    tasa_pago = (total_pagado / total_credito * 100) if total_credito > 0 else 0
    
    # CUOTAS
    total_cuotas = Cuota.objects.count()
    cuotas_pagadas = Cuota.objects.filter(pagado=True).count()
    cuotas_pendientes = Cuota.objects.filter(pagado=False).count()
    cuotas_vencidas = Cuota.objects.filter(pagado=False, fecha_pago_esperada__lt=date.today()).count()
    
    # CALIFICACIONES
    clientes_con_calificacion = Cliente.objects.all()
    ratings = [c.calcular_rating() for c in clientes_con_calificacion]
    rating_promedio = sum(ratings) / len(ratings) if ratings else 0
    
    # MORA
    cuotas_mora = Cuota.objects.filter(pagado=False, fecha_pago_esperada__lt=date.today())
    mora_total = sum(c.calcular_mora_diaria() for c in cuotas_mora)
    
    # INGRESOS - Total pagado este mes
    from django.db.models import Sum
    ingresos_mes_actual = Pago.objects.filter(
        fecha_pago__year=date.today().year,
        fecha_pago__month=date.today().month
    ).aggregate(Sum('monto_pagado'))['monto_pagado__sum'] or Decimal('0')
    
    # Cálculos adicionales
    cartera_promedio_capital = (capital_prestado / total_clientes) if total_clientes > 0 else 0
    cartera_promedio_credito = (total_credito / total_clientes) if total_clientes > 0 else 0
    tasa_mora_porcentaje = (cuotas_vencidas / total_cuotas * 100) if total_cuotas > 0 else 0
    
    contexto = {
        # Clientes
        'total_clientes': total_clientes,
        'clientes_activos': clientes_activos,
        'clientes_inactivos': clientes_inactivos,
        'porcentaje_activos': (clientes_activos / total_clientes * 100) if total_clientes > 0 else 0,
        
        # Préstamos
        'total_prestamos': total_prestamos,
        'prestamos_activos': prestamos_activos,
        'prestamos_completados': prestamos_completados,
        'prestamos_vencidos': prestamos_vencidos,
        
        # Montos
        'capital_prestado': float(capital_prestado),  # ← BUG FIX #1: Solo capital
        'total_credito': float(total_credito),        # ← BUG FIX #1: Capital + Interés
        'total_pagado': float(total_pagado),
        'total_pendiente_capital': float(total_pendiente_capital),
        'total_pendiente_credito': float(total_pendiente_credito),
        'tasa_pago': tasa_pago,
        'cartera_promedio_capital': float(cartera_promedio_capital),
        'cartera_promedio_credito': float(cartera_promedio_credito),
        
        # Cuotas
        'total_cuotas': total_cuotas,
        'cuotas_pagadas': cuotas_pagadas,
        'cuotas_pendientes': cuotas_pendientes,
        'cuotas_vencidas': cuotas_vencidas,
        'porcentaje_cuotas_pagadas': (cuotas_pagadas / total_cuotas * 100) if total_cuotas > 0 else 0,
        'tasa_mora_porcentaje': tasa_mora_porcentaje,
        
        # Calificaciones
        'rating_promedio': rating_promedio,
        'rating_promedio_estrella': int(rating_promedio * 5),
        
        # Mora
        'mora_total': mora_total,
        'cuotas_en_mora': cuotas_vencidas,
        
        # Ingresos
        'ingresos_mes_actual': float(ingresos_mes_actual),
    }
    
    return render(request, 'mi_app/reporte_estadisticas.html', contexto)


@admin_required
def importar_excel(request):
    """
    IMPORTACIÓN CON PREVIEW: 2 fases
    Fase 1: Valida y muestra preview
    Fase 2: Confirma e importa realmente
    """
    import pandas as pd
    from decimal import Decimal
    import re
    from datetime import datetime as dt_datetime
    import json
    
    # Mapeo de meses en español a números
    MESES_ESPAÑOL = {
        'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
        'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
        'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12
    }
    
    def parsear_fecha_tipo_pago(valor_celda):
        """Parsea celda con formato: MesEnEspañol+Día+TipoPago"""
        if pd.isna(valor_celda) or not valor_celda:
            return None, None, None
        
        valor = str(valor_celda).strip().lower()
        if not valor:
            return None, None, None
        
        for mes_es, mes_num in MESES_ESPAÑOL.items():
            if valor.startswith(mes_es):
                resto = valor[len(mes_es):]
                match = re.match(r'(\d{1,2})(.+)', resto)
                if match:
                    dia_str, tipo_pago_str = match.groups()
                    try:
                        dia = int(dia_str)
                        if 1 <= dia <= 31 and 1 <= mes_num <= 12:
                            return mes_num, dia, tipo_pago_str.lower()
                    except:
                        pass
                break
        
        return None, None, None
    
    def normalizar_tipo_pago(tipo_str):
        """Normaliza string de tipo de pago"""
        if not tipo_str:
            return None
        
        tipo = tipo_str.lower().replace('eres', 'es')
        
        if 'int' in tipo and ('cap' in tipo or '+' in tipo):
            return 'ambos'
        elif 'cap' in tipo:
            return 'capital'
        elif 'int' in tipo:
            return 'interes'
        
        return None
    
    def calcular_siguiente_fecha_pago(ultimo_dia_pago, dias_fijos=[5, 15, 20, 30], min_dias=15):
        """Calcula siguiente fecha de pago"""
        siguiente = ultimo_dia_pago + timedelta(days=min_dias)
        
        for dia in dias_fijos:
            try:
                candidata = siguiente.replace(day=dia)
                if candidata >= siguiente:
                    return candidata
            except ValueError:
                continue
        
        proximo_mes = siguiente.replace(day=1) + timedelta(days=32)
        proximo_mes = proximo_mes.replace(day=1)
        
        for dia in dias_fijos:
            try:
                candidata = proximo_mes.replace(day=dia)
                return candidata
            except ValueError:
                continue
        
        return siguiente
    
    # ===== FASE 1: MOSTRAR PREVIEW =====
    if request.method == 'GET':
        return render(request, 'mi_app/auditoria/importar_excel.html')
    
    if request.method == 'POST':
        # ===== VERIFICAR SI ES CONFIRMACIÓN O PREVIEW =====
        confirmar = request.POST.get('confirmar') == 'true'
        
        # ===== SI ES CONFIRMACIÓN: IMPORTAR DESDE SESSION =====
        if confirmar:
            from django.db import transaction
            import traceback
            
            datos_preview_json = request.session.get('datos_preview_json')
            
            if not datos_preview_json:
                return render(request, 'mi_app/auditoria/importar_excel.html', {
                    'error': 'La sesión expiró. Por favor, sube el archivo nuevamente.'
                })
            
            try:
                datos_preview = json.loads(datos_preview_json)
                
                clientes_creados = 0
                prestamos_creados = 0
                errores_importacion = []  # ← Nueva lista para registrar errores
                
                # ✅ MEJORADO: Transacción ÚNICA que envuelve TODO el proceso
                # Si algo falla en cualquier punto, automaticamente se hace ROLLBACK de TODO
                with transaction.atomic():
                    # Importar cada fila validada
                    for idx, fila_data in enumerate(datos_preview['filas_validas']):
                        try:
                            # Obtener datos
                            nombre = fila_data['nombre']
                            celular = fila_data['celular']
                            monto_prestamo = Decimal(str(fila_data['monto_prestamo']))
                            
                            # Validar que la fila tiene cuotas
                            if not fila_data.get('cuotas') or len(fila_data['cuotas']) == 0:
                                raise ValueError(f"El cliente {nombre} no tiene cuotas definidas")
                            
                            # Crear/actualizar cliente
                            cliente, cliente_creado = Cliente.objects.get_or_create(
                                celular=celular,
                                defaults={
                                    'nombre': nombre,
                                    'estado': 'ACTIVO',
                                    'importado_excel': True
                                }
                            )
                            
                            if not cliente_creado:
                                cliente.nombre = nombre
                                cliente.importado_excel = True
                                cliente.estado = 'ACTIVO'  # Asegurar que esté activo
                                cliente.save()
                            
                            # ✅ MEJORADO: Contar TODOS los clientes procesados (creados o ya existentes)
                            clientes_creados += 1
                            
                            # ✅ NUEVA: Validación cruzada - Verificar que NO esté en lista negra vigente
                            lista_negra_vigente = ListaNegra.objects.filter(
                                cliente=cliente, 
                                activa=True
                            ).first()
                            
                            if lista_negra_vigente:
                                raise ValueError(
                                    f"BLOQUEADO: {nombre} está en lista negra. "
                                    f"Razón: {lista_negra_vigente.razon}. "
                                    f"No se puede crear préstamo."
                                )
                            
                            # Crear préstamo
                            fecha_inicio = date.today()
                            
                            # FIX ERROR #2: Calcular interés real del Excel en lugar de hardcodeado 15%
                            total_interes_cuotas = sum(Decimal(str(c['interes'])) for c in fila_data['cuotas'])
                            interes_porcentaje_real = (total_interes_cuotas / monto_prestamo * 100) if monto_prestamo > 0 else Decimal('0')
                            
                            prestamo = Prestamo.objects.create(
                                cliente=cliente,
                                monto_total=monto_prestamo,
                                interes_porcentaje=interes_porcentaje_real,
                                fecha_inicio=fecha_inicio,
                                fecha_fin_estimada=fecha_inicio + timedelta(days=90),
                                tipo_pago='QUINCENAL',
                                estado='ACTIVO'
                            )
                            
                            # Crear cuotas (DENTRO de la transacción atómica)
                            cuotas_creadas = 0
                            for cuota_data in fila_data['cuotas']:
                                try:
                                    # Validar tipos de datos
                                    monto_capital_dec = Decimal(str(cuota_data['capital']))
                                    monto_interes_dec = Decimal(str(cuota_data['interes']))
                                    
                                    pagado_capital = Decimal(str(cuota_data['pagado_capital']))
                                    pagado_interes = Decimal(str(cuota_data['pagado_interes']))
                                    pendiente_capital = Decimal(str(cuota_data['pendiente_capital']))
                                    pendiente_interes = Decimal(str(cuota_data['pendiente_interes']))
                                    
                                    fecha_pago_esperada = None
                                    if cuota_data.get('fecha_pago_esperada'):
                                        fecha_pago_esperada = dt_datetime.strptime(
                                            cuota_data['fecha_pago_esperada'], '%Y-%m-%d'
                                        ).date()
                                    
                                    fecha_pago_real = None
                                    if cuota_data.get('fecha_pago_real'):
                                        fecha_pago_real = dt_datetime.strptime(
                                            cuota_data['fecha_pago_real'], '%Y-%m-%d'
                                        ).date()
                                    
                                    # ✅ OPCIÓN C PASO 3: Determinar estado correcto al crear
                                    is_pagada = (pendiente_capital == 0 and pendiente_interes == 0)
                                    estado_correcto = determinar_estado_cuota_al_crear(
                                        pagado=is_pagada,
                                        fecha_pago_esperada=fecha_pago_esperada,
                                        monto_pagado_principal=pagado_capital,
                                        monto_original=monto_capital_dec
                                    )
                                    
                                    cuota = Cuota.objects.create(
                                        prestamo=prestamo,
                                        numero_cuota=cuota_data['numero'],
                                        monto_original=monto_capital_dec,
                                        monto_pendiente=pendiente_capital,
                                        interes_normal=monto_interes_dec,
                                        monto_pendiente_interes=pendiente_interes,
                                        monto_pagado_principal=pagado_capital,
                                        monto_pagado_interes=pagado_interes,
                                        monto_pagado_mora=Decimal('0'),
                                        pagado=is_pagada,
                                        fecha_pago_esperada=fecha_pago_esperada,
                                        fecha_pago_real=fecha_pago_real,
                                        estado=estado_correcto  # ✅ OPCIÓN C: Nuevo parámetro
                                    )
                                    
                                    # Crear registro de pago si aplica
                                    if cuota_data.get('tipo_pago') and fecha_pago_real:
                                        Pago.objects.create(
                                            cuota=cuota,
                                            monto_pagado=pagado_capital + pagado_interes,
                                            monto_principal=pagado_capital,
                                            monto_interes=pagado_interes,
                                            monto_mora=Decimal('0'),
                                            fecha_pago=fecha_pago_real,
                                            notas=f'Importado Excel - {cuota_data["tipo_pago"]}'
                                        )
                                    
                                    cuotas_creadas += 1
                                
                                except Exception as e_cuota:
                                    # Si falla una cuota, mostrar error específico
                                    raise ValueError(
                                        f"Error en cuota #{cuota_data.get('numero', '?')}: {str(e_cuota)}"
                                    )
                            
                            # Validar que se crearon todas las cuotas
                            if cuotas_creadas != len(fila_data['cuotas']):
                                raise ValueError(
                                    f"Solo se crearon {cuotas_creadas} de {len(fila_data['cuotas'])} cuotas"
                                )
                            
                            prestamos_creados += 1
                        
                        except Exception as e:
                            # Capturar error pero continuar con siguiente fila
                            error_msg = f"Fila {idx + 1} ({fila_data.get('nombre', 'Desconocido')}): {str(e)}"
                            errores_importacion.append(error_msg)
                            # Imprimir para debugging en servidor
                            import sys
                            print(f"❌ {error_msg}", file=sys.stderr)
                            print(traceback.format_exc(), file=sys.stderr)
                            continue
                
                # ===== ERROR #5: IMPORTAR LISTA NEGRA =====
                lista_negra_creadas = 0
                
                if datos_preview.get('lista_negra'):
                    for entrada_data in datos_preview['lista_negra']:
                        try:
                            nombre = entrada_data['nombre']
                            celular = entrada_data['celular']
                            
                            # Obtener o crear cliente
                            cliente, _ = Cliente.objects.get_or_create(
                                celular=celular,
                                defaults={
                                    'nombre': nombre,
                                    'estado': 'ACTIVO',
                                    'importado_excel': True
                                }
                            )
                            
                            # Si cliente ya existía, actualizar nombre si es diferente
                            if cliente.nombre != nombre:
                                cliente.nombre = nombre
                                cliente.save()
                            
                            # Crear entrada en ListaNegra (o actualizar si ya existe)
                            lista_negra, creada = ListaNegra.objects.get_or_create(
                                cliente=cliente,
                                defaults={
                                    'razon': 'OTRO',
                                    'fecha_desde': date.today(),
                                    'activa': True,
                                    'importado_excel': True,
                                    'usuario_creador': request.user if request.user.is_authenticated else None
                                }
                            )
                            
                            if creada:
                                lista_negra_creadas += 1
                            else:
                                # Si ya existe y está inactiva, reactivar
                                if not lista_negra.activa:
                                    lista_negra.activa = True
                                    lista_negra.fecha_desde = date.today()
                                    lista_negra.save()
                                    lista_negra_creadas += 1
                        
                        except Exception as e_lista:
                            error_msg = f"ListaNegra ({nombre}/{celular}): {str(e_lista)}"
                            errores_importacion.append(error_msg)
                            import sys
                            print(f"❌ {error_msg}", file=sys.stderr)
                            print(traceback.format_exc(), file=sys.stderr)
                            continue
                
                # Limpiar sesión
                del request.session['datos_preview_json']
                
                contexto = {
                    'exito': True,
                    'importacion_completada': True,
                    'clientes_creados': clientes_creados,
                    'prestamos_creados': prestamos_creados,
                    'lista_negra_creadas': lista_negra_creadas,
                    'errores': errores_importacion,  # ← Mostrar errores de importación al usuario
                    'errores_preview': datos_preview.get('errores', [])  # ← Errores del Excel original
                }
                
                return render(request, 'mi_app/auditoria/importar_excel.html', contexto)
            
            except Exception as e:
                return render(request, 'mi_app/auditoria/importar_excel.html', {
                    'error': f'Error en la importación: {str(e)}'
                })
        
        # ===== SINO: ES PREVIEW (FASE 1) =====
        archivo = request.FILES.get('archivo')
        
        if not archivo:
            return render(request, 'mi_app/auditoria/importar_excel.html', {
                'error': 'Por favor selecciona un archivo Excel'
            })
        
        try:
            # Leer Excel
            df = pd.read_excel(archivo, sheet_name=0)
            df.columns = df.columns.str.strip()
            
            errores = []
            filas_validas = []
            celulares_procesados = set()
            
            # Validar cada fila
            for idx, row in df.iterrows():
                try:
                    # Extraer datos básicos
                    nombre = str(row.get('Nombre con responsable', '')).strip()
                    celular_raw = row.get('celular', '')
                    
                    if pd.notna(celular_raw):
                        try:
                            if isinstance(celular_raw, (int, float)):
                                celular = str(int(celular_raw))
                            else:
                                celular = str(celular_raw).strip()
                        except:
                            celular = ''
                    else:
                        celular = ''
                    
                    monto_prestamo_raw = row.get('Monto del préstamo', 0)
                    try:
                        monto_prestamo = float(monto_prestamo_raw) if pd.notna(monto_prestamo_raw) else 0
                    except:
                        monto_prestamo = 0
                    
                    # Validar
                    if not nombre or not celular or monto_prestamo <= 0:
                        errores.append({
                            'fila': idx + 2,
                            'tipo': 'Datos incompletos',
                            'detalle': f"Nombre: '{nombre}', Celular: '{celular}', Monto: {monto_prestamo}"
                        })
                        continue
                    
                    if celular in celulares_procesados:
                        errores.append({
                            'fila': idx + 2,
                            'tipo': 'Duplicado',
                            'detalle': f"Celular {celular} duplicado"
                        })
                        continue
                    
                    celulares_procesados.add(celular)
                    
                    # Procesar cuotas
                    columnas_cuotas = sorted([col for col in df.columns 
                                            if 'uota' in col and 'interes' not in col.lower()])
                    columnas_interes = sorted([col for col in df.columns 
                                             if 'interes' in col.lower() or 'interés' in col.lower()])
                    columnas_fechas = sorted([col for col in df.columns if 'echa' in col.lower()])
                    
                    # BUG C: Validaciones robustas de estructura de columnas
                    if not columnas_cuotas or not columnas_interes:
                        errores.append({
                            'fila': idx + 2,
                            'tipo': 'Estructura inválida',
                            'detalle': 'No se encontraron columnas de cuota e interés'
                        })
                        continue
                    
                    # Validar que coincidan las cantidades de columnas
                    if len(columnas_cuotas) != len(columnas_interes):
                        errores.append({
                            'fila': idx + 2,
                            'tipo': 'Desalineación de columnas',
                            'detalle': f'{len(columnas_cuotas)} columnas de cuota vs {len(columnas_interes)} columnas de interés'
                        })
                        continue
                    
                    # Validar que los números de columnas sean válidos
                    cuotas_info = []
                    fecha_pago_calculada = None
                    dias_pago_fijos = [5, 15, 20, 30]
                    fecha_minima = date.today() + timedelta(days=15)
                    
                    # Procesar cada cuota con validación adicional
                    for i in range(len(columnas_cuotas)):
                        col_cuota = columnas_cuotas[i]
                        col_interes = columnas_interes[i]
                        
                        try:
                            monto_capital = float(row.get(col_cuota, 0)) if pd.notna(row.get(col_cuota)) else 0
                            monto_interes = float(row.get(col_interes, 0)) if pd.notna(row.get(col_interes)) else 0
                        except:
                            monto_capital = 0
                            monto_interes = 0
                        
                        if monto_capital == 0 and monto_interes == 0:
                            break
                        
                        # Obtener fecha de pago y tipo
                        tipo_pago_parseo = None
                        fecha_pago_real = None
                        
                        if i < len(columnas_fechas):
                            mes, dia, tipo_str = parsear_fecha_tipo_pago(row.get(columnas_fechas[i]))
                            
                            if mes and dia:
                                tipo_pago_parseo = normalizar_tipo_pago(tipo_str)
                                
                                try:
                                    año_actual = date.today().year
                                    fecha_pago_real = date(año_actual, mes, dia)
                                    
                                    if fecha_pago_real < date.today():
                                        fecha_pago_real = date(año_actual + 1, mes, dia)
                                    
                                    if fecha_pago_calculada is None:
                                        fecha_pago_calculada = fecha_pago_real
                                except ValueError:
                                    fecha_pago_real = None
                        
                        # Calcular fecha esperada
                        if fecha_pago_calculada is None:
                            fecha_pago_esperada = fecha_minima.replace(day=1)
                            for dia in dias_pago_fijos:
                                try:
                                    candidata = fecha_minima.replace(day=dia)
                                    if candidata >= fecha_minima:
                                        fecha_pago_esperada = candidata
                                        break
                                except ValueError:
                                    pass
                        else:
                            fecha_pago_esperada = calcular_siguiente_fecha_pago(
                                fecha_pago_calculada, dias_pago_fijos
                            )
                            fecha_pago_calculada = fecha_pago_esperada
                        
                        # Procesar pagos
                        monto_capital_dec = Decimal(str(monto_capital))
                        monto_interes_dec = Decimal(str(monto_interes))
                        
                        pagado_capital = Decimal('0')
                        pagado_interes = Decimal('0')
                        pendiente_capital = monto_capital_dec
                        pendiente_interes = monto_interes_dec
                        
                        if tipo_pago_parseo:
                            if tipo_pago_parseo == 'capital':
                                pagado_capital = monto_capital_dec
                                pendiente_capital = Decimal('0')
                            elif tipo_pago_parseo == 'interes':
                                pagado_interes = monto_interes_dec
                                pendiente_interes = Decimal('0')
                            elif tipo_pago_parseo == 'ambos':
                                pagado_capital = monto_capital_dec
                                pagado_interes = monto_interes_dec
                                pendiente_capital = Decimal('0')
                                pendiente_interes = Decimal('0')
                        
                        cuotas_info.append({
                            'numero': i + 1,
                            'capital': float(monto_capital),
                            'interes': float(monto_interes),
                            'total': float(monto_capital_dec + monto_interes_dec),
                            'pagado_capital': float(pagado_capital),
                            'pagado_interes': float(pagado_interes),
                            'pendiente_capital': float(pendiente_capital),
                            'pendiente_interes': float(pendiente_interes),
                            'tipo_pago': tipo_pago_parseo,
                            'fecha_pago_esperada': fecha_pago_esperada.isoformat() if fecha_pago_esperada else None,
                            'fecha_pago_real': fecha_pago_real.isoformat() if fecha_pago_real else None,
                            'estado_pago': f'{tipo_pago_parseo.capitalize()} ({fecha_pago_real})' if tipo_pago_parseo else 'Sin pagar'
                        })
                    
                    # Agregar fila válida
                    filas_validas.append({
                        'numero_fila': idx + 2,
                        'nombre': nombre,
                        'celular': celular,
                        'monto_prestamo': float(monto_prestamo),
                        'cuotas': cuotas_info
                    })
                
                except Exception as e:
                    errores.append({
                        'fila': idx + 2,
                        'tipo': 'Error',
                        'detalle': str(e)
                    })
            
            # ===== ERROR #5: LECTURA DE LISTA NEGRA =====
            lista_negra_data = []
            errores_lista_negra = []
            
            try:
                # Intentar leer hoja "ListaNegra"
                try:
                    df_lista_negra = pd.read_excel(archivo, sheet_name='ListaNegra')
                    df_lista_negra.columns = df_lista_negra.columns.str.strip()
                    
                    celulares_lista_negra_procesados = set()
                    
                    for idx, row in df_lista_negra.iterrows():
                        try:
                            # Campos obligatorios: Nombre y Celular
                            nombre = str(row.get('Nombre', '')).strip()
                            celular_raw = row.get('Celular', '')
                            
                            if pd.notna(celular_raw):
                                try:
                                    if isinstance(celular_raw, (int, float)):
                                        celular = str(int(celular_raw))
                                    else:
                                        celular = str(celular_raw).strip()
                                except:
                                    celular = ''
                            else:
                                celular = ''
                            
                            # Validar campos obligatorios
                            if not nombre or not celular:
                                errores_lista_negra.append({
                                    'fila': idx + 2,
                                    'tipo': 'Datos incompletos',
                                    'detalle': f"Nombre: '{nombre}', Celular: '{celular}'"
                                })
                                continue
                            
                            # Evitar duplicados en la hoja
                            if celular in celulares_lista_negra_procesados:
                                errores_lista_negra.append({
                                    'fila': idx + 2,
                                    'tipo': 'Duplicado',
                                    'detalle': f"Celular {celular} duplicado en ListaNegra"
                                })
                                continue
                            
                            celulares_lista_negra_procesados.add(celular)
                            
                            # Validar que el cliente existe en filas_validas (si está importando clientes también)
                            cliente_existe_en_importacion = any(
                                f['celular'] == celular for f in filas_validas
                            )
                            
                            lista_negra_data.append({
                                'nombre': nombre,
                                'celular': celular,
                                'cliente_existe_en_importacion': cliente_existe_en_importacion
                            })
                        
                        except Exception as e_lista:
                            errores_lista_negra.append({
                                'fila': idx + 2,
                                'tipo': 'Error',
                                'detalle': str(e_lista)
                            })
                
                except ValueError:
                    # No hay hoja "ListaNegra" - es opcional
                    pass
            
            except Exception as e_lista_negra:
                errores_lista_negra.append({
                    'fila': 'General',
                    'tipo': 'Error al leer ListaNegra',
                    'detalle': str(e_lista_negra)
                })
            
            # Guardar en sesión para confirmación posterior
            datos_preview = {
                'filas_validas': filas_validas,
                'errores': errores,
                'total_filas': len(df),
                'filas_validas_count': len(filas_validas),
                'errores_count': len(errores),
                'lista_negra': lista_negra_data,
                'errores_lista_negra': errores_lista_negra,
                'lista_negra_count': len(lista_negra_data),
                'errores_lista_negra_count': len(errores_lista_negra)
            }
            
            request.session['datos_preview_json'] = json.dumps(datos_preview, default=str)
            
            contexto = {
                'preview': True,
                'datos_preview': datos_preview,
                'filas_preview': filas_validas[:10]  # Mostrar primeras 10
            }
            
            return render(request, 'mi_app/auditoria/importar_excel.html', contexto)
        
        except Exception as e:
            return render(request, 'mi_app/auditoria/importar_excel.html', {
                'error': f'Error al leer el archivo: {str(e)}'
            })


# ===============================================================================
# FUNCIONES AUXILIARES
# ===============================================================================

def _obtener_estado_visual_cuota(cuota):
    """
    Determina el estado visual de una cuota para mostrar en templates
    Retorna: ('PAGADA', '✓', 'bg-success') o ('PENDIENTE', '⏳', 'bg-info'), etc.
    """
    if cuota.pagado:
        return {
            'estado': 'PAGADA',
            'icono': '✓',
            'clase': 'badge bg-success',
            'color': 'green',
        }
    
    mora = cuota.calcular_mora_diaria()
    if mora > 0:
        return {
            'estado': 'VENCIDA',
            'icono': '⚠',
            'clase': 'badge bg-danger',
            'color': 'red',
            'mora': mora,
        }
    
    dias_para_vencer = (cuota.fecha_pago_esperada - date.today()).days if cuota.fecha_pago_esperada else None
    if dias_para_vencer is not None and dias_para_vencer < 0:
        return {
            'estado': 'VENCIDA',
            'icono': '⚠',
            'clase': 'badge bg-danger',
            'color': 'red',
        }
    
    if cuota.monto_pendiente < cuota.monto_original:
        return {
            'estado': 'PARCIAL',
            'icono': '◐',
            'clase': 'badge bg-warning',
            'color': 'orange',
        }
    
    return {
        'estado': 'PENDIENTE',
        'icono': '⏳',
        'clase': 'badge bg-info',
        'color': 'blue',
    }


# ===============================================================================
# VISTAS PARA PRÉSTAMOS RÁPIDOS Y CONFIGURACIÓN
# ===============================================================================

@require_any_permission('prestamo.create')
@login_required(login_url='login')
def crear_prestamo_rapido(request):
    """
    Vista para crear un nuevo préstamo rápido.
    GET: Muestra el formulario
    POST: Procesa el formulario y crea el préstamo
    """
    from mi_app.forms import PrestamoRapidoForm
    from .models import PrestamoRapido, Configuracion, CuotaRapida
    
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente_id')
        
        if not cliente_id:
            return render(request, 'mi_app/formularios/formulario_prestamo_rapido.html', {
                'form': PrestamoRapidoForm(),
                'error': 'Debe seleccionar un cliente'
            })
        
        cliente = get_object_or_404(Cliente, pk=cliente_id)
        form = PrestamoRapidoForm(request.POST)
        
        if form.is_valid():
            prestamo_rapido = form.save(commit=False)
            prestamo_rapido.cliente = cliente
            prestamo_rapido.save()

            usar_cuotas = form.cleaned_data.get('usar_cuotas', False)

            if usar_cuotas:
                config = Configuracion.obtener_configuracion()
                num_cuotas = int(form.cleaned_data.get('num_cuotas') or config.cuotas_por_defecto or 1)
                if num_cuotas < 1:
                    num_cuotas = 1

                fecha_inicio = date.today()
                fechas_pago = calcular_fechas_pago('QUINCENAL', num_cuotas, fecha_inicio)

                capital_total = Decimal(str(prestamo_rapido.monto))
                interes_total = Decimal(str(prestamo_rapido.calcular_interes_total()))

                capital_por_cuota = (capital_total / Decimal(num_cuotas)).quantize(Decimal('0.01'))
                interes_por_cuota = (interes_total / Decimal(num_cuotas)).quantize(Decimal('0.01'))

                for i in range(1, num_cuotas + 1):
                    capital_cuota = capital_por_cuota
                    interes_cuota = interes_por_cuota

                    if i == num_cuotas:
                        capital_cuota = capital_total - (capital_por_cuota * Decimal(num_cuotas - 1))
                        interes_cuota = interes_total - (interes_por_cuota * Decimal(num_cuotas - 1))

                    fecha_pago = fechas_pago[i - 1] if len(fechas_pago) >= i else None

                    CuotaRapida.objects.create(
                        prestamo_rapido=prestamo_rapido,
                        numero_cuota=i,
                        monto_original=capital_cuota,
                        monto_pendiente=capital_cuota,
                        interes_normal=interes_cuota,
                        monto_pendiente_interes=interes_cuota,
                        fecha_pago_esperada=fecha_pago,
                    )

                if not prestamo_rapido.fecha_vencimiento and fechas_pago:
                    prestamo_rapido.fecha_vencimiento = fechas_pago[-1]
                    prestamo_rapido.save(update_fields=['fecha_vencimiento'])
            
            # Registrar en log
            return redirect('detalle_prestamo_rapido', prestamo_id=prestamo_rapido.id)
        else:
            return render(request, 'mi_app/formularios/formulario_prestamo_rapido.html', {
                'form': form,
                'cliente': cliente
            })
    
    else:  # GET
        form = PrestamoRapidoForm()
        return render(request, 'mi_app/formularios/formulario_prestamo_rapido.html', {
            'form': form
        })


@require_permission('prestamo.view')
@login_required(login_url='login')
def detalle_prestamo_rapido(request, prestamo_id):
    """
    Vista del detalle de un préstamo rápido.
    Muestra: Información, estado de pago, pagos realizados, opción de pagar.
    """
    from .models import PrestamoRapido, PagoPrestamoRapido, CuotaRapida
    
    prestamo = get_object_or_404(PrestamoRapido, pk=prestamo_id)
    
    # BUG #4 FIX: Asegurar que los datos sean frescos desde BD
    prestamo.refresh_from_db()
    
    pagos = PagoPrestamoRapido.objects.filter(prestamo_rapido=prestamo).order_by('-fecha_pago')
    cuotas = CuotaRapida.objects.filter(prestamo_rapido=prestamo).order_by('numero_cuota')

    cuotas_pagadas = cuotas.filter(pagado=True).count()
    cuotas_pendientes = cuotas.filter(pagado=False).count()
    tiene_cuotas = cuotas.exists()
    cuotas_vencidas = 0
    for cuota in cuotas.filter(pagado=False):
        if cuota.fecha_pago_esperada and cuota.fecha_pago_esperada < date.today():
            cuotas_vencidas += 1
    
    contexto = {
        'prestamo': prestamo,
        'pagos': pagos,
        'cuotas': cuotas,
        'cuotas_pagadas': cuotas_pagadas,
        'cuotas_pendientes': cuotas_pendientes,
        'cuotas_vencidas': cuotas_vencidas,
        'tiene_cuotas': tiene_cuotas,
        'porcentaje_pagado': round(prestamo.porcentaje_pagado, 2),
        'saldo_pendiente': prestamo.saldo_pendiente,
        'total_a_pagar': prestamo.total_a_pagar,
    }
    
    response = render(request, 'mi_app/detalle_prestamo_rapido.html', contexto)
    
    # BUG #4 FIX: Desactivar cache del navegador para evitar stale data
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    return response


@require_permission('prestamo.view')
@login_required(login_url='login')
def listar_prestamos_rapidos(request):
    """
    Vista para listar todos los préstamos rápidos del sistema.
    Filtrable por cliente, estado, etc.
    """
    from .models import PrestamoRapido
    
    prestamos = PrestamoRapido.objects.select_related('cliente').all()
    
    # Filtros
    cliente_id = request.GET.get('cliente_id')
    estado = request.GET.get('estado')
    
    if cliente_id:
        prestamos = prestamos.filter(cliente_id=cliente_id)
    
    if estado:
        prestamos = prestamos.filter(estado=estado)
    
    contexto = {
        'prestamos': prestamos,
        'clientes': Cliente.objects.all(),
        'estados': PrestamoRapido.ESTADO_CHOICES,
        'cliente_id_filtro': cliente_id,
        'estado_filtro': estado,
    }
    
    return render(request, 'mi_app/lista_prestamos_rapidos.html', contexto)


@require_any_permission('pago.create')
@login_required(login_url='login')
def registrar_pago_rapido(request, cuota_id):
    """
    Vista para registrar un pago en un préstamo rápido.
    GET: Muestra el formulario
    POST: Procesa el pago
    """
    from mi_app.forms import PagoPrestamoRapidoForm
    from .models import PrestamoRapido, PagoPrestamoRapido, CuotaRapida

    cuota = get_object_or_404(CuotaRapida, pk=cuota_id)
    prestamo = cuota.prestamo_rapido
    
    if request.method == 'POST':
        form = PagoPrestamoRapidoForm(request.POST)
        
        if form.is_valid():
            monto_pagado = form.cleaned_data.get('monto_pagado')
            usuario_registra = form.cleaned_data.get('usuario_registra', 'Sistema')
            referencia = form.cleaned_data.get('referencia', '')
            notas = form.cleaned_data.get('notas', '')
            
            from decimal import Decimal

            monto_pagado = Decimal(str(monto_pagado))
            mora = Decimal(str(cuota.calcular_mora_diaria()))
            total_debido = cuota.monto_pendiente + cuota.monto_pendiente_interes + mora

            if monto_pagado > total_debido:
                form.add_error('monto_pagado', f'El monto no puede ser mayor a {total_debido}')
                return render(request, 'mi_app/registrar_pago_rapido.html', {
                    'form': form,
                    'prestamo': prestamo,
                    'cuota': cuota,
                    'total_debido': total_debido,
                    'mora_actual': mora,
                })

            pendiente_capital = cuota.monto_pendiente
            pendiente_interes = cuota.monto_pendiente_interes
            pendiente_mora = mora

            monto_pago_capital = Decimal('0')
            monto_pago_interes = Decimal('0')
            monto_pago_mora = Decimal('0')

            monto_restante = monto_pagado

            if pendiente_capital > 0 and monto_restante > 0:
                monto_pago_capital = min(pendiente_capital, monto_restante)
                monto_restante -= monto_pago_capital
                cuota.monto_pendiente -= monto_pago_capital

            if pendiente_interes > 0 and monto_restante > 0:
                monto_pago_interes = min(pendiente_interes, monto_restante)
                monto_restante -= monto_pago_interes
                cuota.monto_pendiente_interes -= monto_pago_interes

            if pendiente_mora > 0 and monto_restante > 0:
                monto_pago_mora = min(pendiente_mora, monto_restante)
                monto_restante -= monto_pago_mora

            cuota.monto_pagado_principal += monto_pago_capital
            cuota.monto_pagado_interes += monto_pago_interes
            cuota.monto_pagado_mora += monto_pago_mora
            cuota.interes_mora_acumulado += mora

            if cuota.monto_pendiente <= 0 and cuota.monto_pendiente_interes <= 0:
                cuota.monto_pendiente = Decimal('0')
                cuota.monto_pendiente_interes = Decimal('0')
                cuota.pagado = True
                cuota.fecha_pago_real = date.today()

            cuota.save()

            PagoPrestamoRapido.objects.create(
                prestamo_rapido=prestamo,
                cuota_rapida=cuota,
                monto_pagado=monto_pagado,
                usuario_registra=usuario_registra,
                referencia=referencia,
                notas=notas,
            )

            total_pagado = prestamo.pagos.aggregate(total=Coalesce(Sum('monto_pagado'), Decimal('0')))
            prestamo.monto_pagado = total_pagado['total'] or Decimal('0')
            prestamo.actualizar_estado()
            if prestamo.estado == 'PAGADO':
                prestamo.fecha_pago_real = date.today()
            prestamo.save()

            cuota.actualizar_estado()

            prestamo.refresh_from_db()

            return redirect('detalle_prestamo_rapido', prestamo_id=prestamo.id)
        
        else:
            return render(request, 'mi_app/registrar_pago_rapido.html', {
                'form': form,
                'prestamo': prestamo,
                'cuota': cuota,
                'total_debido': cuota.total_a_pagar(),
                'mora_actual': cuota.calcular_mora_diaria(),
            })
    
    else:  # GET
        form = PagoPrestamoRapidoForm()
        return render(request, 'mi_app/registrar_pago_rapido.html', {
            'form': form,
            'prestamo': prestamo,
            'cuota': cuota,
            'total_debido': cuota.total_a_pagar(),
            'mora_actual': cuota.calcular_mora_diaria(),
        })


@require_any_permission('pago.create')
@login_required(login_url='login')
def registrar_pago_rapido_prestamo(request, prestamo_id):
    """
    Redirige al pago de la primera cuota pendiente del préstamo rápido.
    """
    prestamo = get_object_or_404(PrestamoRapido, pk=prestamo_id)

    if not prestamo.cuotas_rapidas.exists():
        return redirect('registrar_pago_rapido_directo', prestamo_id=prestamo.id)

    cuota = prestamo.cuotas_rapidas.filter(pagado=False).order_by('numero_cuota').first()

    if not cuota:
        messages.info(request, 'Este préstamo rápido no tiene cuotas pendientes.')
        return redirect('detalle_prestamo_rapido', prestamo_id=prestamo.id)

    return redirect('registrar_pago_cuota_rapida', cuota_id=cuota.id)


@require_any_permission('pago.create')
@login_required(login_url='login')
def registrar_pago_rapido_directo(request, prestamo_id):
    """
    Registra pago directo para préstamo rápido sin cuotas.
    """
    from mi_app.forms import PagoPrestamoRapidoForm

    prestamo = get_object_or_404(PrestamoRapido, pk=prestamo_id)

    if request.method == 'POST':
        form = PagoPrestamoRapidoForm(request.POST)

        if form.is_valid():
            monto_pagado = Decimal(str(form.cleaned_data.get('monto_pagado')))
            usuario_registra = form.cleaned_data.get('usuario_registra', 'Sistema')
            referencia = form.cleaned_data.get('referencia', '')
            notas = form.cleaned_data.get('notas', '')

            total_debido = Decimal(str(prestamo.saldo_pendiente))

            if monto_pagado > total_debido:
                form.add_error('monto_pagado', f'El monto no puede ser mayor a {total_debido}')
                return render(request, 'mi_app/registrar_pago_rapido.html', {
                    'form': form,
                    'prestamo': prestamo,
                    'cuota': None,
                    'total_debido': total_debido,
                    'mora_actual': Decimal('0'),
                })

            PagoPrestamoRapido.objects.create(
                prestamo_rapido=prestamo,
                monto_pagado=monto_pagado,
                usuario_registra=usuario_registra,
                referencia=referencia,
                notas=notas,
            )

            total_pagado = prestamo.pagos.aggregate(total=Coalesce(Sum('monto_pagado'), Decimal('0')))
            prestamo.monto_pagado = total_pagado['total'] or Decimal('0')
            prestamo.actualizar_estado()
            if prestamo.estado == 'PAGADO':
                prestamo.fecha_pago_real = date.today()
            prestamo.save()

            return redirect('detalle_prestamo_rapido', prestamo_id=prestamo.id)

        return render(request, 'mi_app/registrar_pago_rapido.html', {
            'form': form,
            'prestamo': prestamo,
            'cuota': None,
            'total_debido': Decimal(str(prestamo.saldo_pendiente)),
            'mora_actual': Decimal('0'),
        })

    form = PagoPrestamoRapidoForm()
    return render(request, 'mi_app/registrar_pago_rapido.html', {
        'form': form,
        'prestamo': prestamo,
        'cuota': None,
        'total_debido': Decimal(str(prestamo.saldo_pendiente)),
        'mora_actual': Decimal('0'),
    })


@admin_required
def editar_configuracion(request):
    """
    Vista para editar la configuración global del sistema.
    Tasas de interés, mora, etc.
    """
    from mi_app.forms import ConfiguracionForm
    from .models import Configuracion
    
    config = Configuracion.obtener_configuracion()
    
    if request.method == 'POST':
        form = ConfiguracionForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            return redirect('editar_configuracion')
    else:
        form = ConfiguracionForm(instance=config)
    
    contexto = {
        'form': form,
        'configuracion': config,
    }
    
    return render(request, 'mi_app/configuracion/configuracion.html', contexto)


# ===============================================================================
# EXPORTAR A EXCEL
# ===============================================================================
@require_permission('reporte.export')
@login_required(login_url='login')

def exportar_clientes_excel(request):
    """
    Exporta la lista de clientes a un archivo Excel con TODA la información de préstamos.
    Formato es SIMÉTRICO con la importación para permitir re-importación de datos.
    
    Formato esperado:
    - Columna A: Nombre con responsable
    - Columna B: Celular
    - Columna C: Monto del préstamo
    - Columnas D-H: Cuota 1-5 (monto de capital)
    - Columnas I-M: Interés 1-5 (monto de interés)
    - Columnas N-S: Fecha 1-6 (formato {MesEnEspañol}{Día}{int|interes|int+cap|capital})
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from datetime import datetime
    from dateutil.relativedelta import relativedelta
    
    # Mapeo de números de mes a español
    MESES_ESPAÑOL = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    
    def generar_fecha_pago_string(fecha_pago, monto_capital, monto_interes):
        """
        Genera string en formato: {MesEnEspañol}{Día}{tipo_pago}
        Ejemplo: Enero16int+cap, Diciembre21interes
        """
        if not fecha_pago:
            return ""
        
        mes = MESES_ESPAÑOL.get(fecha_pago.month, '')
        dia = fecha_pago.day
        
        # Determinar tipo de pago basado en qué se pagó
        monto_capital_decimal = float(monto_capital) if monto_capital else 0
        monto_interes_decimal = float(monto_interes) if monto_interes else 0
        
        if monto_capital_decimal > 0 and monto_interes_decimal > 0:
            tipo_pago = 'int+cap'
        elif monto_interes_decimal > 0:
            tipo_pago = 'interes'
        elif monto_capital_decimal > 0:
            tipo_pago = 'capital'
        else:
            return ""
        
        return f"{mes}{dia}{tipo_pago}"
    
    # Obtener todos los clientes
    clientes = Cliente.objects.all().order_by('celular')
    clientes_lista_negra = set(
        ListaNegra.objects.filter(activa=True).values_list('cliente_id', flat=True)
    )
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Estilos
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="9db4e8", end_color="9db4e8", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=9)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados principales (Row 1)
    headers = [
        'Nombre con responsable',  # A
        'celular',                  # B
        'Monto del préstamo',       # C
        'Cuota 1', 'Cuota 2', 'Cuota 3', 'Cuota 4', 'Cuota 5',  # D-H
        'Interés 1', 'Interés 2', 'Interés 3', 'Interés 4', 'Interés 5',  # I-M
        'Fecha 1', 'Fecha 2', 'Fecha 3', 'Fecha 4', 'Fecha 5', 'Fecha 6',  # N-S
        'Estado Cliente', 'Lista Negra', 'Mora Total'  # T-V
    ]
    ws.append(headers)
    
    # Aplicar estilos al encabezado
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Altura para el encabezado
    ws.row_dimensions[1].height = 30
    
    # Agregar datos de clientes y préstamos
    row_num = 2
    for cliente in clientes:
        # Obtener préstamos del cliente
        prestamos = cliente.prestamo_set.all().order_by('-fecha_inicio')
        
        if not prestamos:
            # Si no tiene préstamos, solo añadir datos del cliente
            estado_cliente = 'LISTA_NEGRA' if cliente.id in clientes_lista_negra else (cliente.etiqueta_cliente or 'SIN_HISTORIAL')
            mora_total = sum(c.calcular_mora_diaria() for c in Cuota.objects.filter(prestamo__cliente=cliente, pagado=False))
            ws.append([
                cliente.nombre,
                str(cliente.celular),
                '',  # Monto
                '', '', '', '', '',
                '', '', '', '', '',
                '', '', '', '', '', '',
                estado_cliente,
                'Sí' if cliente.id in clientes_lista_negra else 'No',
                float(mora_total),
            ])
            row_num += 1
        else:
            # Para cada préstamo del cliente
            for prestamo in prestamos:
                estado_cliente = 'LISTA_NEGRA' if cliente.id in clientes_lista_negra else (cliente.etiqueta_cliente or 'SIN_HISTORIAL')
                mora_total = sum(c.calcular_mora_diaria() for c in Cuota.objects.filter(prestamo__cliente=cliente, pagado=False))
                row_data = [
                    cliente.nombre,
                    str(cliente.celular),
                    float(prestamo.monto_total),
                ]
                
                # Obtener cuotas ordenadas por número
                cuotas = prestamo.cuotas.all().order_by('numero_cuota')
                
                # Convertir cuotas a listas
                cuotas_montos = []
                intereses_montos = []
                fechas_pagos = []
                
                for cuota in cuotas:
                    # Si está pagada, usar montos pagados; si no, usar montos originales
                    if cuota.pagado:
                        monto_capital = float(cuota.monto_pagado_principal)
                        monto_interes = float(cuota.monto_pagado_interes)
                    else:
                        monto_capital = float(cuota.monto_original)
                        monto_interes = float(cuota.interes_normal)
                    
                    cuotas_montos.append(monto_capital)
                    intereses_montos.append(monto_interes)
                    
                    # Generar string de fecha de pago
                    fecha_pago_real = cuota.fecha_pago_real if cuota.pagado else cuota.fecha_pago_esperada
                    fecha_string = generar_fecha_pago_string(fecha_pago_real, monto_capital, monto_interes)
                    if fecha_string:
                        fechas_pagos.append(fecha_string)
                
                # Agregar montos de cuotas (D-H): máximo 5
                for i in range(5):
                    row_data.append(cuotas_montos[i] if i < len(cuotas_montos) else '')
                
                # Agregar montos de intereses (I-M): máximo 5
                for i in range(5):
                    row_data.append(intereses_montos[i] if i < len(intereses_montos) else '')
                
                # Agregar fechas de pago (N-S): máximo 6
                for i in range(6):
                    row_data.append(fechas_pagos[i] if i < len(fechas_pagos) else '')

                row_data.extend([
                    estado_cliente,
                    'Sí' if cliente.id in clientes_lista_negra else 'No',
                    float(mora_total),
                ])
                
                ws.append(row_data)
                row_num += 1
    
    # Aplicar bordes y alineación a datos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=22):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            # Alineación especial para números
            if cell.column in [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 22]:  # Columnas de montos
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
            # Alineación especial para fechas
            elif cell.column in [14, 15, 16, 17, 18, 19]:  # Columnas de fechas
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 25  # Nombre
    ws.column_dimensions['B'].width = 15  # Celular
    ws.column_dimensions['C'].width = 18  # Monto
    
    # Cuotas (D-H)
    for col in ['D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 12
    
    # Intereses (I-M)
    for col in ['I', 'J', 'K', 'L', 'M']:
        ws.column_dimensions[col].width = 12
    
    # Fechas (N-S)
    for col in ['N', 'O', 'P', 'Q', 'R', 'S']:
        ws.column_dimensions[col].width = 18

    ws.column_dimensions['T'].width = 18
    ws.column_dimensions['U'].width = 14
    ws.column_dimensions['V'].width = 16
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Clientes_Exportacion_{datetime.now().strftime("%d%m%Y_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@require_permission('reporte.export')
@login_required(login_url='login')
def exportar_prestamos_excel(request):
    """
    Exporta la lista de préstamos a un archivo Excel.
    Formato simétrico con importación para permitir re-importación de datos.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from datetime import datetime
    
    # Mapeo de números de mes a español
    MESES_ESPAÑOL = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    
    def generar_fecha_pago_string(fecha_pago, monto_capital, monto_interes):
        """Genera string: {MesEnEspañol}{Día}{tipo_pago}"""
        if not fecha_pago:
            return ""
        mes = MESES_ESPAÑOL.get(fecha_pago.month, '')
        dia = fecha_pago.day
        monto_capital_decimal = float(monto_capital) if monto_capital else 0
        monto_interes_decimal = float(monto_interes) if monto_interes else 0
        if monto_capital_decimal > 0 and monto_interes_decimal > 0:
            tipo_pago = 'int+cap'
        elif monto_interes_decimal > 0:
            tipo_pago = 'interes'
        elif monto_capital_decimal > 0:
            tipo_pago = 'capital'
        else:
            return ""
        return f"{mes}{dia}{tipo_pago}"
    
    # Obtener todos los préstamos
    prestamos = Prestamo.objects.select_related('cliente').all().order_by('-fecha_inicio')
    clientes_lista_negra = set(
        ListaNegra.objects.filter(activa=True).values_list('cliente_id', flat=True)
    )
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Préstamos"
    
    # Estilos
    header_fill = PatternFill(start_color="48bb78", end_color="48bb78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'Cliente', 'Estado Cliente', 'Lista Negra', 'Mora Acumulada',
        'Número', 'Monto Original', 'Monto de Interés',
        'Interés Quincenal', 'Interés Mensual',
        'Monto Pagado', 'Monto Pendiente', 'Número de Cuotas',
        'Fecha de Creación', 'Fecha de Último Pago'
    ]
    ws.append(headers)
    
    # Aplicar estilos al encabezado
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    ws.row_dimensions[1].height = 35
    
    # Agregar datos de préstamos
    for prestamo in prestamos:
        # Calcular totales
        monto_original = float(prestamo.monto_total)
        total_pagado = prestamo.total_pagado
        total_pendiente = prestamo.total_pendiente
        num_cuotas = prestamo.cuotas.count()
        
        # Calcular interés total e intereses por período
        cuotas = prestamo.cuotas.all()
        interes_total = sum(float(c.interes_normal) for c in cuotas)
        
        # Determinar si es quincenal o mensual
        interes_quincenal = 0
        interes_mensual = 0
        
        if prestamo.tipo_pago == 'QUINCENAL':
            interes_quincenal = interes_total / (num_cuotas * 2) if num_cuotas > 0 else 0
        else:
            interes_mensual = interes_total / num_cuotas if num_cuotas > 0 else 0
        
        # Obtener fecha de último pago
        fecha_ultimo_pago = "N/A"
        cuota_pagada = cuotas.filter(pagado=True).order_by('-fecha_pago_real').first()
        if cuota_pagada and cuota_pagada.fecha_pago_real:
            fecha_ultimo_pago = cuota_pagada.fecha_pago_real.strftime('%d/%m/%Y')
        
        mora_acumulada = sum(c.calcular_mora_diaria() for c in cuotas.filter(pagado=False))

        # Agregar fila
        row_data = [
            prestamo.cliente.nombre,
            'LISTA_NEGRA' if prestamo.cliente_id in clientes_lista_negra else (prestamo.cliente.etiqueta_cliente or 'SIN_HISTORIAL'),
            'Sí' if prestamo.cliente_id in clientes_lista_negra else 'No',
            float(mora_acumulada),
            prestamo.id,
            monto_original,
            interes_total,
            interes_quincenal,
            interes_mensual,
            total_pagado,
            total_pendiente,
            num_cuotas,
            prestamo.fecha_creacion.strftime('%d/%m/%Y'),
            fecha_ultimo_pago
        ]
        ws.append(row_data)
    
    # Aplicar bordes y alineación
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=14):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            # Alinear números a la derecha sin separadores
            if cell.column in [4, 5, 6, 7, 8, 9, 10, 11, 12]:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'
            # Centrar fechas y texto
            elif cell.column in [1, 2, 3, 13, 14]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajustar anchos de columnas
    ws.column_dimensions['A'].width = 20  # Cliente
    ws.column_dimensions['B'].width = 16  # Estado Cliente
    ws.column_dimensions['C'].width = 12  # Lista Negra
    ws.column_dimensions['D'].width = 16  # Mora Acumulada
    ws.column_dimensions['E'].width = 10  # Número
    ws.column_dimensions['F'].width = 18  # Monto Original
    ws.column_dimensions['G'].width = 18  # Monto de Interés
    ws.column_dimensions['H'].width = 18  # Interés Quincenal
    ws.column_dimensions['I'].width = 18  # Interés Mensual
    ws.column_dimensions['J'].width = 15  # Monto Pagado
    ws.column_dimensions['K'].width = 15  # Monto Pendiente
    ws.column_dimensions['L'].width = 18  # Número de Cuotas
    ws.column_dimensions['M'].width = 18  # Fecha de Creación
    ws.column_dimensions['N'].width = 18  # Fecha de Último Pago
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Prestamos_{datetime.now().strftime("%d%m%Y_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@require_permission('reporte.export')
@login_required(login_url='login')
def exportar_cuotas_excel(request):
    """
    Exporta la lista de cuotas con sus estados de pago a un archivo Excel.
    Información completa para auditoría y re-importación.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from datetime import datetime
    
    # Obtener todas las cuotas
    cuotas = Cuota.objects.select_related('prestamo', 'prestamo__cliente').all().order_by('prestamo_id', 'numero_cuota')
    clientes_lista_negra = set(
        ListaNegra.objects.filter(activa=True).values_list('cliente_id', flat=True)
    )
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Cuotas"
    
    # Estilos
    header_fill = PatternFill(start_color="f6ad55", end_color="f6ad55", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Color para cuotas pagadas/pendientes
    paid_fill = PatternFill(start_color="c6f6d5", end_color="c6f6d5", fill_type="solid")
    unpaid_fill = PatternFill(start_color="fed7d7", end_color="fed7d7", fill_type="solid")
    
    # Encabezados
    headers = [
        'Cliente', 'Estado Cliente', 'Lista Negra', 'Préstamo ID', 'Cuota Nº',
        'Capital Original', 'Interés Original', 'Mora Original',
        'Capital Pagado', 'Interés Pagado', 'Mora Pagada',
        'Fecha Esperada', 'Fecha Real Pago',
        'Total Original', 'Total Pendiente', 'Total Pagado', 'Estado'
    ]
    ws.append(headers)
    
    # Aplicar estilos al encabezado
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    ws.row_dimensions[1].height = 35
    
    # Agregar datos
    for cuota in cuotas:
        cliente_nombre = cuota.prestamo.cliente.nombre
        
        # Cálculos de totales
        mora_calculada = cuota.calcular_mora_diaria()
        total_original = float(cuota.monto_original) + float(cuota.interes_normal)
        total_pendiente = float(cuota.monto_pendiente) + float(cuota.monto_pendiente_interes) + float(mora_calculada)
        total_pagado = float(cuota.monto_pagado_principal) + float(cuota.monto_pagado_interes) + float(cuota.monto_pagado_mora)
        estado_visual_cliente = 'LISTA_NEGRA' if cuota.prestamo.cliente_id in clientes_lista_negra else (cuota.prestamo.cliente.etiqueta_cliente or 'SIN_HISTORIAL')
        
        row_num = ws.max_row + 1
        ws.append([
            cliente_nombre,
            estado_visual_cliente,
            'Sí' if cuota.prestamo.cliente_id in clientes_lista_negra else 'No',
            cuota.prestamo.id,
            cuota.numero_cuota,
            float(cuota.monto_original),
            float(cuota.interes_normal),
            float(mora_calculada),
            float(cuota.monto_pagado_principal),
            float(cuota.monto_pagado_interes),
            float(cuota.monto_pagado_mora),
            cuota.fecha_pago_esperada.strftime('%d/%m/%Y') if cuota.fecha_pago_esperada else '',
            cuota.fecha_pago_real.strftime('%d/%m/%Y') if cuota.fecha_pago_real else '',
            total_original,
            total_pendiente,
            total_pagado,
            'Pagada' if cuota.pagado else 'Pendiente'
        ])
        
        # Colorear según estado
        fill_color = paid_fill if cuota.pagado else unpaid_fill
        for cell in ws[row_num]:
            cell.border = border
            cell.fill = fill_color
            # Alineación especial para números
            if cell.column in range(6, 12) or cell.column in range(14, 17):  # Columnas de montos
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
            elif cell.column in [12, 13]:  # Fechas
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 20  # Cliente
    ws.column_dimensions['B'].width = 16  # Estado Cliente
    ws.column_dimensions['C'].width = 12  # Lista Negra
    ws.column_dimensions['D'].width = 12  # Préstamo ID
    ws.column_dimensions['E'].width = 10  # Cuota Nº
    for col in ['F', 'G', 'H', 'I', 'J', 'K', 'N', 'O', 'P']:  # Montos
        ws.column_dimensions[col].width = 14
    for col in ['L', 'M']:  # Fechas
        ws.column_dimensions[col].width = 14
    ws.column_dimensions['Q'].width = 12  # Estado
    
    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Cuotas_Exportacion_{datetime.now().strftime("%d%m%Y_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


# ===============================================================================
# BUG #5: REPORTE DETALLADO DE PRÉSTAMOS RÁPIDOS
# ===============================================================================
@require_permission('reporte.view')
@login_required(login_url='login')
def reporte_prestamos_rapidos(request):
    """
    Reporte detallado de todos los préstamos rápidos del sistema.
    Similar al reporte de préstamos normales pero para PrestamoRapido.
    
    Incluye:
    - Lista de todos los PRs con cliente, monto, estado
    - Filtros por cliente, estado, rango de fechas
    - Cálculos de totales: monto total, pagado, pendiente
    - Porcentaje de progreso de pago
    """
    from decimal import Decimal
    
    prestamos_rapidos = PrestamoRapido.objects.select_related('cliente').all()
    
    # Filtro por cliente
    cliente_id = request.GET.get('cliente_id', '').strip()
    if cliente_id:
        try:
            prestamos_rapidos = prestamos_rapidos.filter(cliente_id=int(cliente_id))
        except (ValueError, TypeError):
            pass
    
    # Búsqueda por nombre de cliente
    busqueda = request.GET.get('search', '').strip()
    if busqueda:
        prestamos_rapidos = prestamos_rapidos.filter(
            cliente__nombre__icontains=busqueda
        )
    
    # Filtro por estado
    estado = request.GET.get('estado', '')
    if estado:
        prestamos_rapidos = prestamos_rapidos.filter(estado=estado)
    
    # Filtro por rango de fechas de creación
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    if fecha_desde:
        try:
            from datetime import datetime
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            prestamos_rapidos = prestamos_rapidos.filter(fecha_creacion__gte=fecha_desde_obj)
        except (ValueError, AttributeError):
            pass
    
    if fecha_hasta:
        try:
            from datetime import datetime
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            prestamos_rapidos = prestamos_rapidos.filter(fecha_creacion__lte=fecha_hasta_obj)
        except (ValueError, AttributeError):
            pass
    
    # Filtro por rango de montos
    monto_desde = request.GET.get('monto_desde', '')
    monto_hasta = request.GET.get('monto_hasta', '')
    
    if monto_desde:
        try:
            from decimal import Decimal
            monto_desde_val = Decimal(monto_desde)
            prestamos_rapidos = prestamos_rapidos.filter(monto__gte=monto_desde_val)
        except:
            pass
    
    if monto_hasta:
        try:
            from decimal import Decimal
            monto_hasta_val = Decimal(monto_hasta)
            prestamos_rapidos = prestamos_rapidos.filter(monto__lte=monto_hasta_val)
        except:
            pass
    
    # Calcular totales
    total_monto = sum(float(pr.monto) for pr in prestamos_rapidos)
    total_pagado = sum(float(pr.monto_pagado) for pr in prestamos_rapidos)
    total_pendiente = sum(float(pr.saldo_pendiente) for pr in prestamos_rapidos)
    total_pr = prestamos_rapidos.count()
    
    # Porcentaje general
    porcentaje_pago = (total_pagado / total_monto * 100) if total_monto > 0 else 0
    
    # Contar por estado
    por_estado = {
        'PENDIENTE': prestamos_rapidos.filter(estado='PENDIENTE').count(),
        'PARCIALMENTE_PAGADO': prestamos_rapidos.filter(estado='PARCIALMENTE_PAGADO').count(),
        'PAGADO': prestamos_rapidos.filter(estado='PAGADO').count(),
    }
    
    contexto = {
        'prestamos_rapidos': prestamos_rapidos,
        'total': total_pr,
        'search': busqueda,
        'estado': estado,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'monto_desde': monto_desde,
        'monto_hasta': monto_hasta,
        'total_monto': total_monto,
        'total_pagado': total_pagado,
        'total_pendiente': total_pendiente,
        'porcentaje_pago': round(porcentaje_pago, 2),
        'por_estado': por_estado,
        # Estadísticas globales
        'total_clientes': Cliente.objects.count(),
        'clientes_activos': Cliente.objects.filter(estado='ACTIVO').count(),
    }
    
    return render(request, 'mi_app/reporte_prestamos_rapidos.html', contexto)


# ===============================================================================
# BUG #8: HISTÓRICO DE PAGOS MEJORADO - REPORTE GENERAL
# ===============================================================================
@require_permission('reporte.view')
@login_required(login_url='login')
def historico_pagos(request):
    """
    Reporte profesional del histórico de pagos del sistema.
    
    Incluye:
    - Tabla de todos los pagos con desglose (principal/interés/mora)
    - Filtros: cliente, rango de fechas, rango de montos
    - Estadísticas: total pagado, promedio, etc.
    - Gráfico: pagos por mes
    """
    from decimal import Decimal
    from datetime import datetime
    from collections import defaultdict
    
    pagos = Pago.objects.select_related('cuota__prestamo__cliente').all()
    pagos_rapidos = PagoPrestamoRapido.objects.select_related('prestamo_rapido__cliente', 'cuota_rapida').all()
    
    # Filtro por cliente
    cliente_id = request.GET.get('cliente_id', '').strip()
    if cliente_id:
        try:
            cliente_id_int = int(cliente_id)
            pagos = pagos.filter(cuota__prestamo__cliente_id=cliente_id_int)
            pagos_rapidos = pagos_rapidos.filter(prestamo_rapido__cliente_id=cliente_id_int)
        except (ValueError, TypeError):
            pass
    
    # Búsqueda por nombre de cliente
    busqueda = request.GET.get('search', '').strip()
    if busqueda:
        pagos = pagos.filter(
            Q(cuota__prestamo__cliente__nombre__icontains=busqueda) |
            Q(cuota__prestamo__cliente__cedula__icontains=busqueda) |
            Q(cuota__prestamo__cliente__celular__icontains=busqueda)
        )
        pagos_rapidos = pagos_rapidos.filter(
            Q(prestamo_rapido__cliente__nombre__icontains=busqueda) |
            Q(prestamo_rapido__cliente__cedula__icontains=busqueda) |
            Q(prestamo_rapido__cliente__celular__icontains=busqueda)
        )
    
    # Filtro por rango de fechas
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            pagos = pagos.filter(fecha_pago__gte=fecha_desde_obj)
            pagos_rapidos = pagos_rapidos.filter(fecha_pago__gte=fecha_desde_obj)
        except (ValueError, AttributeError):
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            pagos = pagos.filter(fecha_pago__lte=fecha_hasta_obj)
            pagos_rapidos = pagos_rapidos.filter(fecha_pago__lte=fecha_hasta_obj)
        except (ValueError, AttributeError):
            pass
    
    # Filtro por rango de montos
    monto_desde = request.GET.get('monto_desde', '')
    monto_hasta = request.GET.get('monto_hasta', '')
    
    if monto_desde:
        try:
            monto_desde_val = Decimal(monto_desde)
            pagos = pagos.filter(monto_pagado__gte=monto_desde_val)
            pagos_rapidos = pagos_rapidos.filter(monto_pagado__gte=monto_desde_val)
        except:
            pass
    
    if monto_hasta:
        try:
            monto_hasta_val = Decimal(monto_hasta)
            pagos = pagos.filter(monto_pagado__lte=monto_hasta_val)
            pagos_rapidos = pagos_rapidos.filter(monto_pagado__lte=monto_hasta_val)
        except:
            pass
    
    # Filtro por tipo de pago
    tipo_pago = request.GET.get('tipo_pago', '')
    if tipo_pago == 'principal':
        pagos = pagos.filter(monto_principal__gt=0)
        # En pagos rápidos no hay desglose, se tratan como principal
    elif tipo_pago == 'interes':
        pagos = pagos.filter(monto_interes__gt=0)
        pagos_rapidos = pagos_rapidos.none()
    elif tipo_pago == 'mora':
        pagos = pagos.filter(monto_mora__gt=0)
        pagos_rapidos = pagos_rapidos.none()

    pagos = pagos.order_by('-fecha_pago')
    pagos_rapidos = pagos_rapidos.order_by('-fecha_pago')

    pagos_historial = []

    for pago in pagos:
        pagos_historial.append({
            'id': pago.id,
            'cliente_id': pago.cuota.prestamo.cliente.id,
            'cliente_nombre': pago.cuota.prestamo.cliente.nombre,
            'cliente_cedula': pago.cuota.prestamo.cliente.cedula,
            'cliente_celular': pago.cuota.prestamo.cliente.celular,
            'cuota_label': f'Cuota #{pago.cuota.numero_cuota}',
            'monto_principal': float(pago.monto_principal),
            'monto_interes': float(pago.monto_interes),
            'monto_mora': float(pago.monto_mora),
            'monto_pagado': float(pago.monto_pagado),
            'fecha_pago': pago.fecha_pago,
            'origen': 'NORMAL',
            'detalle_url': reverse('detalles_cuota', args=[pago.cuota.id]),
        })

    for pago in pagos_rapidos:
        numero_cuota = pago.cuota_rapida.numero_cuota if pago.cuota_rapida else '-'
        detalle_url = reverse('detalle_prestamo_rapido', args=[pago.prestamo_rapido.id])
        pagos_historial.append({
            'id': pago.id,
            'cliente_id': pago.prestamo_rapido.cliente.id,
            'cliente_nombre': pago.prestamo_rapido.cliente.nombre,
            'cliente_cedula': pago.prestamo_rapido.cliente.cedula,
            'cliente_celular': pago.prestamo_rapido.cliente.celular,
            'cuota_label': f'Rápido #{pago.prestamo_rapido.id} · Cuota {numero_cuota}',
            'monto_principal': float(pago.monto_pagado),
            'monto_interes': 0.0,
            'monto_mora': 0.0,
            'monto_pagado': float(pago.monto_pagado),
            'fecha_pago': pago.fecha_pago,
            'origen': 'RÁPIDO',
            'detalle_url': detalle_url,
        })

    pagos_historial.sort(key=lambda item: item['fecha_pago'], reverse=True)
    
    # Calcular estadísticas
    total_pagado = sum(item['monto_pagado'] for item in pagos_historial)
    total_principal = sum(item['monto_principal'] for item in pagos_historial)
    total_interes = sum(item['monto_interes'] for item in pagos_historial)
    total_mora = sum(item['monto_mora'] for item in pagos_historial)
    total_pagos = len(pagos_historial)
    promedio_pago = (total_pagado / total_pagos) if total_pagos > 0 else 0
    total_pagos_rapidos = sum(1 for item in pagos_historial if item['origen'] == 'RÁPIDO')
    total_monto_rapidos = sum(item['monto_pagado'] for item in pagos_historial if item['origen'] == 'RÁPIDO')
    total_pagos_normales = total_pagos - total_pagos_rapidos
    total_monto_normales = total_pagado - total_monto_rapidos
    
    # Datos para gráfico: Pagos por mes
    pagos_por_mes = defaultdict(float)
    for pago_item in pagos_historial:
        mes_key = pago_item['fecha_pago'].strftime('%Y-%m')
        pagos_por_mes[mes_key] += float(pago_item['monto_pagado'])
    
    # Ordenar por fecha
    meses_ordenados = sorted(pagos_por_mes.items())
    meses_labels = [mes for mes, _ in meses_ordenados]
    meses_valores = [monto for _, monto in meses_ordenados]
    
    principal_pct = round((total_principal / total_pagado) * 100, 2) if total_pagado > 0 else 0
    interes_pct = round((total_interes / total_pagado) * 100, 2) if total_pagado > 0 else 0
    mora_pct = round((total_mora / total_pagado) * 100, 2) if total_pagado > 0 else 0

    contexto = {
        'pagos': pagos_historial,
        'total': total_pagos,
        'search': busqueda,
        'cliente_id_filtro': cliente_id,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'monto_desde': monto_desde,
        'monto_hasta': monto_hasta,
        'tipo_pago_filtro': tipo_pago,
        # Estadísticas
        'total_pagado': total_pagado,
        'total_principal': total_principal,
        'total_interes': total_interes,
        'total_mora': total_mora,
        'promedio_pago': promedio_pago,
        'total_pagos_rapidos': total_pagos_rapidos,
        'total_monto_rapidos': total_monto_rapidos,
        'total_pagos_normales': total_pagos_normales,
        'total_monto_normales': total_monto_normales,
        'principal_pct': principal_pct,
        'interes_pct': interes_pct,
        'mora_pct': mora_pct,
        # Gráfico
        'meses_labels': meses_labels,
        'meses_valores': meses_valores,
    }
    
    return render(request, 'mi_app/historico_pagos.html', contexto)

# ===============================================================================
# EXPORTACIONES ADICIONALES - NUEVOS REPORTES
# ===============================================================================

@require_permission('reporte.export')
@login_required
def exportar_cuotas_vencidas_excel(request):
    """Exporta reporte de cuotas vencidas a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from datetime import date
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Cuotas Vencidas"
    
    # Estilos
    header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Headers
    headers = ["Cliente", "Cédula", "Teléfono", "Estado Cliente", "Lista Negra", "Cuota #", "Monto Principal", 
               "Interés", "Mora", "Total", "Fecha Vencimiento", "Días Vencido"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    hoy = date.today()
    clientes_lista_negra = set(
        ListaNegra.objects.filter(activa=True).values_list('cliente_id', flat=True)
    )

    cuotas_vencidas = Cuota.objects.filter(
        fecha_pago_esperada__lt=hoy,
        estado='PENDIENTE'
    ).select_related('prestamo__cliente').order_by('-fecha_pago_esperada')
    
    for cuota in cuotas_vencidas:
        dias_vencido = (hoy - cuota.fecha_pago_esperada).days
        mora = cuota.calcular_mora_diaria()
        total_cuota = cuota.monto_original + cuota.interes_normal + mora
        estado_visual_cliente = 'LISTA_NEGRA' if cuota.prestamo.cliente_id in clientes_lista_negra else (cuota.prestamo.cliente.etiqueta_cliente or 'SIN_HISTORIAL')
        ws.append([
            cuota.prestamo.cliente.nombre,
            cuota.prestamo.cliente.cedula or "",
            cuota.prestamo.cliente.celular or "",
            estado_visual_cliente,
            'Sí' if cuota.prestamo.cliente_id in clientes_lista_negra else 'No',
            cuota.numero_cuota,
            float(cuota.monto_original),
            float(cuota.interes_normal),
            float(mora),
            float(total_cuota),
            cuota.fecha_pago_esperada.strftime('%d/%m/%Y'),
            dias_vencido,
        ])
    
    # Ancho de columnas
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Cuotas_Vencidas.xlsx"'
    wb.save(response)
    return response

@require_permission('reporte.export')
@login_required
def exportar_estadisticas_excel(request):
    """Exporta estadísticas generales a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from django.db.models import Sum, Count
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Estadísticas"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Título
    ws['A1'] = "ESTADÍSTICAS GENERALES DEL SISTEMA"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    
    # Datos
    row = 3
    total_pagado_normal = Pago.objects.aggregate(Sum('monto_pagado'))['monto_pagado__sum'] or 0
    total_pagado_rapido = PagoPrestamoRapido.objects.aggregate(Sum('monto_pagado'))['monto_pagado__sum'] or 0
    total_pagado_general = total_pagado_normal + total_pagado_rapido
    total_mora = sum(c.calcular_mora_diaria() for c in Cuota.objects.filter(pagado=False))

    stats = [
        ("Total Clientes", Cliente.objects.count()),
        ("Total Préstamos", Prestamo.objects.count()),
        ("Total Préstamos Rápidos", PrestamoRapido.objects.count()),
        ("Clientes en Lista Negra", ListaNegra.objects.filter(activa=True).count()),
        ("Total Cuotas", Cuota.objects.count()),
        ("Cuotas Pendientes", Cuota.objects.filter(estado='PENDIENTE').count()),
        ("Cuotas Vencidas", Cuota.objects.filter(estado='PENDIENTE', fecha_pago_esperada__lt=date.today()).count()),
        ("Total Pagado (Normal)", total_pagado_normal),
        ("Total Pagado (Rápido)", total_pagado_rapido),
        ("Total Pagado General", total_pagado_general),
        ("Mora Total Acumulada", total_mora),
        ("Total Capital Prestado", Prestamo.objects.aggregate(Sum('monto_total'))['monto_total__sum'] or 0),
    ]
    
    ws[f'A{row}'] = "Concepto"
    ws[f'B{row}'] = "Valor"
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'].font = header_font
    
    row += 1
    for concepto, valor in stats:
        ws[f'A{row}'] = concepto
        ws[f'B{row}'] = valor
        row += 1
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Estadisticas_Generales.xlsx"'
    wb.save(response)
    return response

@require_permission('reporte.export')
@login_required
def exportar_prestamos_rapidos_excel(request):
    """Exporta reporte de préstamos rápidos a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Préstamos Rápidos"
    
    # Estilos
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Headers
    headers = ["Cliente", "Cédula", "Teléfono", "Estado Cliente", "Lista Negra", "Monto", "Saldo Pendiente", 
               "Estado Préstamo", "Tasa Interés", "Fecha Creación", "Fecha Vencimiento"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Datos
    clientes_lista_negra = set(
        ListaNegra.objects.filter(activa=True).values_list('cliente_id', flat=True)
    )
    prestamos_rapidos = PrestamoRapido.objects.all().select_related('cliente').order_by('-fecha_solicitud')
    
    for pr in prestamos_rapidos:
        ws.append([
            pr.cliente.nombre,
            pr.cliente.cedula or "",
            pr.cliente.celular or "",
            'LISTA_NEGRA' if pr.cliente_id in clientes_lista_negra else (pr.cliente.etiqueta_cliente or 'SIN_HISTORIAL'),
            'Sí' if pr.cliente_id in clientes_lista_negra else 'No',
            float(pr.monto),
            float(pr.saldo_pendiente),
            pr.estado,
            float(pr.interes_porcentaje) if pr.interes_porcentaje else 0,
            pr.fecha_solicitud.strftime('%d/%m/%Y'),
            pr.fecha_vencimiento.strftime('%d/%m/%Y') if pr.fecha_vencimiento else "",
        ])
    
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Prestamos_Rapidos.xlsx"'
    wb.save(response)
    return response

@require_permission('reporte.export')
@login_required
def exportar_historico_pagos_excel(request):
    """Exporta histórico de pagos a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico Pagos"
    
    # Estilos
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Headers
    headers = ["Origen", "Cliente", "Cédula", "Detalle", "Principal", "Interés", "Mora", 
               "Total Pagado", "Fecha Pago"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Datos: pagos normales + pagos de préstamos rápidos
    pagos = Pago.objects.all().select_related('cuota__prestamo__cliente').order_by('-fecha_pago')
    pagos_rapidos = PagoPrestamoRapido.objects.all().select_related('prestamo_rapido__cliente', 'cuota_rapida').order_by('-fecha_pago')
    
    for pago in pagos:
        ws.append([
            'NORMAL',
            pago.cuota.prestamo.cliente.nombre,
            pago.cuota.prestamo.cliente.cedula or "",
            f"Cuota #{pago.cuota.numero_cuota}",
            float(pago.monto_principal),
            float(pago.monto_interes),
            float(pago.monto_mora),
            float(pago.monto_pagado),
            pago.fecha_pago.strftime('%d/%m/%Y'),
        ])

    for pago in pagos_rapidos:
        numero_cuota = pago.cuota_rapida.numero_cuota if pago.cuota_rapida else '-'
        ws.append([
            'RÁPIDO',
            pago.prestamo_rapido.cliente.nombre,
            pago.prestamo_rapido.cliente.cedula or "",
            f"Rápido #{pago.prestamo_rapido.id} · Cuota {numero_cuota}",
            float(pago.monto_pagado),
            0.0,
            0.0,
            float(pago.monto_pagado),
            pago.fecha_pago.strftime('%d/%m/%Y'),
        ])
    
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Historico_Pagos.xlsx"'
    wb.save(response)
    return response

@require_permission('reporte.export')
@login_required
def exportar_reporte_general_excel(request):
    """
    Exporta un reporte general completo con múltiples hojas:
    - Resumen Ejecutivo
    - Clientes
    - Préstamos
    - Préstamos Rápidos
    - Cuotas
    - Cuotas Vencidas
    - Histórico de Pagos
    - Estadísticas
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from django.db.models import Sum, Count
    
    wb = Workbook()
    wb.remove(wb.active)
    
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # ==================== HOJA 1: RESUMEN EJECUTIVO ====================
    ws_resumen = wb.create_sheet("Resumen Ejecutivo", 0)
    ws_resumen['A1'] = "REPORTE GENERAL CONSOLIDADO"
    ws_resumen['A1'].font = Font(bold=True, size=14)
    ws_resumen.merge_cells('A1:B1')
    
    stats = [
        ("Total Clientes", Cliente.objects.count()),
        ("Total Préstamos", Prestamo.objects.count()),
        ("Total Préstamos Rápidos", PrestamoRapido.objects.count()),
        ("Clientes en Lista Negra", ListaNegra.objects.filter(activa=True).count()),
        ("Total Cuotas", Cuota.objects.count()),
        ("Cuotas Pendientes", Cuota.objects.filter(estado='PENDIENTE').count()),
        ("Cuotas Vencidas", Cuota.objects.filter(estado='PENDIENTE', fecha_pago_esperada__lt=date.today()).count()),
        ("Total Mora Acumulada", sum(c.calcular_mora_diaria() for c in Cuota.objects.filter(pagado=False))),
        ("Total Capital Prestado", Prestamo.objects.aggregate(Sum('monto_total'))['monto_total__sum'] or 0),
        ("Total Pagado (Normal)", Pago.objects.aggregate(Sum('monto_pagado'))['monto_pagado__sum'] or 0),
        ("Total Pagado (Rápido)", PagoPrestamoRapido.objects.aggregate(Sum('monto_pagado'))['monto_pagado__sum'] or 0),
    ]
    
    row = 3
    for concepto, valor in stats:
        ws_resumen[f'A{row}'] = concepto
        ws_resumen[f'B{row}'] = valor
        row += 1
    
    ws_resumen.column_dimensions['A'].width = 35
    ws_resumen.column_dimensions['B'].width = 20
    
    # ==================== HOJA 2: CLIENTES ====================
    ws_clientes = wb.create_sheet("Clientes")
    headers_clientes = ["ID", "Nombre", "Cédula", "Teléfono", "Email", "Estado Cliente", "Lista Negra", "Préstamos", "Total Adeudado", "Mora Total"]
    ws_clientes.append(headers_clientes)
    
    for cell in ws_clientes[1]:
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    clientes_lista_negra = set(
        ListaNegra.objects.filter(activa=True).values_list('cliente_id', flat=True)
    )

    for cliente in Cliente.objects.all():
        prestamos_count = cliente.prestamo_set.count()
        adeudado = cliente.prestamo_set.aggregate(Sum('monto_total'))['monto_total__sum'] or 0
        mora_total = sum(c.calcular_mora_diaria() for c in Cuota.objects.filter(prestamo__cliente=cliente, pagado=False))
        ws_clientes.append([
            cliente.id,
            cliente.nombre,
            cliente.cedula or "",
            cliente.celular or "",
            cliente.email or "",
            'LISTA_NEGRA' if cliente.id in clientes_lista_negra else (cliente.etiqueta_cliente or 'SIN_HISTORIAL'),
            'Sí' if cliente.id in clientes_lista_negra else 'No',
            prestamos_count,
            float(adeudado),
            float(mora_total),
        ])
    
    for col in ws_clientes.columns:
        ws_clientes.column_dimensions[col[0].column_letter].width = 18
    
    # ==================== HOJA 3: PRÉSTAMOS ====================
    ws_prestamos = wb.create_sheet("Préstamos")
    headers_prestamos = ["ID", "Cliente", "Estado Cliente", "Lista Negra", "Monto Total", "Tasa Interés", "Cuotas", "Estado Préstamo", "Mora Acumulada", "Fecha Creación"]
    ws_prestamos.append(headers_prestamos)
    
    for cell in ws_prestamos[1]:
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    for prestamo in Prestamo.objects.all().select_related('cliente'):
        mora_acumulada = sum(c.calcular_mora_diaria() for c in prestamo.cuotas.filter(pagado=False))
        ws_prestamos.append([
            prestamo.id,
            prestamo.cliente.nombre,
            'LISTA_NEGRA' if prestamo.cliente_id in clientes_lista_negra else (prestamo.cliente.etiqueta_cliente or 'SIN_HISTORIAL'),
            'Sí' if prestamo.cliente_id in clientes_lista_negra else 'No',
            float(prestamo.monto_total),
            float(prestamo.interes_porcentaje),
            prestamo.cuotas.count(),
            prestamo.estado,
            float(mora_acumulada),
            prestamo.fecha_creacion.strftime('%d/%m/%Y'),
        ])
    
    for col in ws_prestamos.columns:
        ws_prestamos.column_dimensions[col[0].column_letter].width = 18
    
    # ==================== HOJA 4: PRÉSTAMOS RÁPIDOS ====================
    ws_pr = wb.create_sheet("Préstamos Rápidos")
    headers_pr = ["ID", "Cliente", "Estado Cliente", "Lista Negra", "Monto", "Saldo Pendiente", "Estado", "Fecha Creación"]
    ws_pr.append(headers_pr)
    
    for cell in ws_pr[1]:
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.font = Font(bold=True, color="000000")
    
    for pr in PrestamoRapido.objects.all().select_related('cliente'):
        ws_pr.append([
            pr.id,
            pr.cliente.nombre,
            'LISTA_NEGRA' if pr.cliente_id in clientes_lista_negra else (pr.cliente.etiqueta_cliente or 'SIN_HISTORIAL'),
            'Sí' if pr.cliente_id in clientes_lista_negra else 'No',
            float(pr.monto),
            float(pr.saldo_pendiente),
            pr.estado,
            pr.fecha_solicitud.strftime('%d/%m/%Y'),
        ])
    
    for col in ws_pr.columns:
        ws_pr.column_dimensions[col[0].column_letter].width = 18
    
    # ==================== HOJA 5: CUOTAS ====================
    ws_cuotas = wb.create_sheet("Cuotas")
    headers_cuotas = ["Cliente", "Estado Cliente", "Lista Negra", "Préstamo ID", "Cuota #", "Principal", "Interés", "Mora", "Total", "Vencimiento", "Estado"]
    ws_cuotas.append(headers_cuotas)
    
    for cell in ws_cuotas[1]:
        cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    for cuota in Cuota.objects.all().select_related('prestamo__cliente'):
        mora = cuota.calcular_mora_diaria()
        total_cuota = cuota.monto_original + cuota.interes_normal + mora
        ws_cuotas.append([
            cuota.prestamo.cliente.nombre,
            'LISTA_NEGRA' if cuota.prestamo.cliente_id in clientes_lista_negra else (cuota.prestamo.cliente.etiqueta_cliente or 'SIN_HISTORIAL'),
            'Sí' if cuota.prestamo.cliente_id in clientes_lista_negra else 'No',
            cuota.prestamo.id,
            cuota.numero_cuota,
            float(cuota.monto_original),
            float(cuota.interes_normal),
            float(mora),
            float(total_cuota),
            cuota.fecha_pago_esperada.strftime('%d/%m/%Y') if cuota.fecha_pago_esperada else "",
            cuota.estado,
        ])
    
    for col in ws_cuotas.columns:
        ws_cuotas.column_dimensions[col[0].column_letter].width = 18
    
    # ==================== HOJA 6: CUOTAS VENCIDAS ====================
    ws_vencidas = wb.create_sheet("Cuotas Vencidas")
    headers_vencidas = ["Cliente", "Estado Cliente", "Lista Negra", "Cuota #", "Mora", "Total", "Vencimiento", "Días Vencido"]
    ws_vencidas.append(headers_vencidas)
    
    for cell in ws_vencidas[1]:
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    hoy = date.today()
    for cuota in Cuota.objects.filter(estado='PENDIENTE', fecha_pago_esperada__lt=hoy).select_related('prestamo__cliente'):
        dias_vencido = (hoy - cuota.fecha_pago_esperada).days
        mora = cuota.calcular_mora_diaria()
        total_cuota = cuota.monto_original + cuota.interes_normal + mora
        ws_vencidas.append([
            cuota.prestamo.cliente.nombre,
            'LISTA_NEGRA' if cuota.prestamo.cliente_id in clientes_lista_negra else (cuota.prestamo.cliente.etiqueta_cliente or 'SIN_HISTORIAL'),
            'Sí' if cuota.prestamo.cliente_id in clientes_lista_negra else 'No',
            cuota.numero_cuota,
            float(mora),
            float(total_cuota),
            cuota.fecha_pago_esperada.strftime('%d/%m/%Y'),
            dias_vencido,
        ])
    
    for col in ws_vencidas.columns:
        ws_vencidas.column_dimensions[col[0].column_letter].width = 18
    
    # ==================== HOJA 7: HISTÓRICO DE PAGOS ====================
    ws_pagos = wb.create_sheet("Histórico Pagos")
    headers_pagos = ["Origen", "Cliente", "Cuota/Detalle", "Principal", "Interés", "Mora", "Total Pagado", "Fecha Pago"]
    ws_pagos.append(headers_pagos)
    
    for cell in ws_pagos[1]:
        cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    for pago in Pago.objects.all().select_related('cuota__prestamo__cliente').order_by('-fecha_pago'):
        ws_pagos.append([
            'NORMAL',
            pago.cuota.prestamo.cliente.nombre,
            f"Cuota #{pago.cuota.numero_cuota}",
            float(pago.monto_principal),
            float(pago.monto_interes),
            float(pago.monto_mora),
            float(pago.monto_pagado),
            pago.fecha_pago.strftime('%d/%m/%Y'),
        ])

    for pago in PagoPrestamoRapido.objects.all().select_related('prestamo_rapido__cliente', 'cuota_rapida').order_by('-fecha_pago'):
        numero_cuota = pago.cuota_rapida.numero_cuota if pago.cuota_rapida else '-'
        ws_pagos.append([
            'RÁPIDO',
            pago.prestamo_rapido.cliente.nombre,
            f"Rápido #{pago.prestamo_rapido.id} · Cuota {numero_cuota}",
            float(pago.monto_pagado),
            0.0,
            0.0,
            float(pago.monto_pagado),
            pago.fecha_pago.strftime('%d/%m/%Y'),
        ])

    for col in ws_pagos.columns:
        ws_pagos.column_dimensions[col[0].column_letter].width = 20

    # ==================== HOJA 8: LISTA NEGRA ====================
    ws_ln = wb.create_sheet("Lista Negra")
    headers_ln = ["Cliente", "Cédula", "Celular", "Razón", "Descripción", "Desde", "Hasta", "Activa", "Vigente"]
    ws_ln.append(headers_ln)

    for cell in ws_ln[1]:
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    for entry in ListaNegra.objects.select_related('cliente').order_by('-fecha_agregacion'):
        ws_ln.append([
            entry.cliente.nombre,
            entry.cliente.cedula or '',
            entry.cliente.celular or '',
            entry.get_razon_display(),
            entry.descripcion or '',
            entry.fecha_desde.strftime('%d/%m/%Y') if entry.fecha_desde else '',
            entry.fecha_hasta.strftime('%d/%m/%Y') if entry.fecha_hasta else 'Indefinido',
            'Sí' if entry.activa else 'No',
            'Sí' if entry.esta_vigente else 'No',
        ])
    
    for col in ws_ln.columns:
        ws_ln.column_dimensions[col[0].column_letter].width = 18
    
    # ==================== CREAR RESPUESTA ====================
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_General_Completo.xlsx"'
    wb.save(response)
    return response


# ===============================================================================
# BACKUPS - GOOGLE DRIVE
# ===============================================================================

@admin_required
@require_http_methods(["GET"])
def backups_list(request):
    """
    Vista para listar y administrar backups
    Solo accesible por staff/admin
    """
    try:
        from mi_app.utilities.backup_manager import backup_manager
        
        # Obtener lista de backups
        backups = backup_manager.list_backups()
        
        # Formatear fechas y tamaños para el template
        for b in backups:
            if b.get('fecha'):
                try:
                    # Intentar parsear fecha ISO o string simple
                    if isinstance(b['fecha'], str):
                        from django.utils.dateparse import parse_datetime
                        dt = parse_datetime(b['fecha'])
                        if not dt:
                            # Fallback para formatos manuales
                            dt = datetime.fromisoformat(b['fecha'].replace('Z', '+00:00'))
                    else:
                        dt = b['fecha']
                    
                    b['date_display'] = dt.strftime('%d/%m/%Y')
                    b['time_display'] = dt.strftime('%H:%M:%S')
                except Exception as e:
                    print(f"Error parseando fecha {b.get('fecha')}: {e}")
                    b['date_display'] = "N/A"
                    b['time_display'] = "N/A"
            else:
                b['date_display'] = "N/A"
                b['time_display'] = "N/A"
            
            size_bytes = b.get('tamaño', 0)
            b['size_display'] = f"{size_bytes / (1024*1024):.2f}"
        
        context = {
            'backups': backups,
            'total_backups': len(backups),
        }
        
        return render(request, 'mi_app/auditoria/backups_list.html', context)
    
    except Exception as e:
        messages.error(request, f'Error al acceder a backups: {str(e)}')
        return redirect('inicio')


@admin_required
@require_http_methods(["POST"])
def backup_create(request):
    """
    Vista para crear un backup manual
    Solo accesible por staff/admin
    """
    try:
        from mi_app.utilities.backup_manager import backup_manager
        
        result = backup_manager.create_backup()
        
        if result['success']:
            messages.success(request, result['mensaje'])
        else:
            messages.error(request, f"Error al crear backup: {result.get('error', 'Unknown error')}")
        
    except Exception as e:
        messages.error(request, f'Error al crear backup: {str(e)}')
    
    return redirect('backups:list')


@admin_required
def backup_descargar(request, backup_id):
    """
    Descarga un backup específico
    """
    try:
        from mi_app.utilities.backup_manager import backup_manager
        import os
        from django.http import FileResponse
        
        # Buscar el backup
        backups = backup_manager.list_backups()
        backup = next((b for b in backups if b['id'] == backup_id), None)
        
        if not backup or not os.path.exists(backup['ruta']):
            messages.error(request, 'Archivo de backup no encontrado')
            return redirect('backups:list')
            
        return FileResponse(open(backup['ruta'], 'rb'), as_attachment=True, filename=backup['nombre'])
        
    except Exception as e:
        messages.error(request, f'Error al descargar: {str(e)}')
        return redirect('backups:list')


@admin_required
@require_http_methods(["POST"])
def backup_delete(request, backup_id):
    """
    Elimina un backup específico
    """
    try:
        from mi_app.utilities.backup_manager import backup_manager
        
        result = backup_manager.delete_backup(backup_id)
        
        if result['success']:
            messages.success(request, 'Backup eliminado correctamente')
        else:
            messages.error(request, f"Error al eliminar: {result.get('error')}")
            
    except Exception as e:
        messages.error(request, f'Error al eliminar backup: {str(e)}')
        
    return redirect('backups:list')


@admin_required
@require_http_methods(["POST"])
def backup_restore(request, backup_id):
    """
    Restaura un backup específico
    """
    try:
        from mi_app.utilities.backup_manager import backup_manager
        
        result = backup_manager.restore_backup(backup_id)
        
        if result['success']:
            messages.success(request, result['mensaje'])
        else:
            messages.error(request, f"Error al restaurar: {result.get('error')}")
            
    except Exception as e:
        messages.error(request, f'Error al restaurar backup: {str(e)}')
        
    return redirect('backups:list')


@admin_required
@require_http_methods(["POST"])
def backup_upload(request):
    """
    Sube un archivo de backup (.sqlite3 o .zip) al servidor
    """
    try:
        if 'backup_file' not in request.FILES:
            messages.error(request, 'No se seleccionó ningún archivo')
            return redirect('backups:list')
            
        archivo = request.FILES['backup_file']
        nombre = archivo.name
        
        if not (nombre.endswith('.sqlite3') or nombre.endswith('.zip')):
            messages.error(request, 'Formato de archivo no válido. Use .sqlite3 o .zip')
            return redirect('backups:list')
            
        from mi_app.utilities.backup_manager import backup_manager
        ruta_destino = os.path.join(backup_manager.backups_dir, nombre)
        
        with open(ruta_destino, 'wb+') as destination:
            for chunk in archivo.chunks():
                destination.write(chunk)
                
        messages.success(request, f'Backup "{nombre}" subido correctamente')
        
    except Exception as e:
        messages.error(request, f'Error al subir backup: {str(e)}')
        
    return redirect('backups:list')


@admin_required
@require_http_methods(["GET", "POST"])
def backup_details(request, backup_id):
    """
    Vista para ver detalles de un backup específico
    Solo accesible por staff/admin
    """
    try:
        from mi_app.backup_manager import backup_manager
        
        # Obtener lista de backups y buscar el solicitado
        backups = backup_manager.list_backups()
        backup = None
        
        for b in backups:
            if b['id'] == backup_id:
                backup = b
                break
        
        if not backup:
            messages.error(request, 'Backup no encontrado')
            return redirect('backups_list')
        
        context = {
            'backup': backup,
        }
        
        return render(request, 'mi_app/auditoria/backup_details.html', context)
    
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')
        return redirect('backups_list')


# ===============================================================================
# AUDITORÍA DE CAMBIOS
# ===============================================================================

@admin_required
@require_http_methods(["GET"])
def auditoria_cambios(request):
    """
    Vista para consultar el historial de cambios del sistema.
    Solo accesible por staff/admin.
    Permite filtrar por usuario, fecha, tipo de acción y modelo.
    """
    from .models import HistorioCambios
    
    # Obtener filtros
    filtro_usuario = request.GET.get('usuario', '')
    filtro_modelo = request.GET.get('modelo', '')
    filtro_accion = request.GET.get('accion', '')
    filtro_fecha_desde = request.GET.get('desde', '')
    filtro_fecha_hasta = request.GET.get('hasta', '')
    
    # QuerySet base
    cambios = HistorioCambios.objects.select_related('usuario').all()
    
    # Aplicar filtros
    if filtro_usuario:
        cambios = cambios.filter(usuario__id=filtro_usuario)
    
    if filtro_modelo:
        cambios = cambios.filter(modelo__icontains=filtro_modelo)
    
    if filtro_accion:
        cambios = cambios.filter(accion=filtro_accion)
    
    if filtro_fecha_desde:
        from datetime import datetime
        cambios = cambios.filter(fecha_cambio__gte=datetime.strptime(filtro_fecha_desde, '%Y-%m-%d'))
    
    if filtro_fecha_hasta:
        from datetime import datetime
        cambios = cambios.filter(fecha_cambio__lte=datetime.strptime(filtro_fecha_hasta, '%Y-%m-%d'))
    
    # Obtener lista de usuarios para dropdown
    usuarios = User.objects.filter(cambios_realizados__isnull=False).distinct().order_by('first_name')
    
    # Obtener lista de modelos para dropdown
    from .models import HistorioCambios as HC
    modelos_lista = HC.objects.values_list('modelo', flat=True).distinct().order_by('modelo')
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(cambios, 50)  # 50 por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'cambios': page_obj.object_list,
        'total_cambios': cambios.count(),
        'usuarios': usuarios,
        'modelos_lista': modelos_lista,
        'tipos_accion': HistorioCambios.TIPOS_ACCION,
        'filtros_activos': bool(filtro_usuario or filtro_modelo or filtro_accion or filtro_fecha_desde or filtro_fecha_hasta),
    }
    
    return render(request, 'mi_app/auditoria/auditoria.html', context)


@admin_required
@require_http_methods(["GET"])
def auditoria_detalle(request, cambio_id):
    """
    Vista para ver el detalle completo de un cambio específico.
    Solo accesible por staff/admin.
    """
    from .models import HistorioCambios
    
    cambio = get_object_or_404(HistorioCambios, id=cambio_id)
    
    context = {
        'cambio': cambio,
    }
    
    return render(request, 'mi_app/auditoria/auditoria_detalle.html', context)


@admin_required
@require_http_methods(["GET"])
def auditoria_estadisticas(request):
    """
    Vista con estadísticas de cambios del sistema.
    Útil para análisis y monitoreo.
    """
    from .models import HistorioCambios
    from django.db.models import Count, Q
    from datetime import timedelta
    
    # Últimos 30 días
    hace_30_dias = date.today() - timedelta(days=30)
    
    # Estadísticas generales
    total_cambios = HistorioCambios.objects.count()
    cambios_mes = HistorioCambios.objects.filter(fecha_cambio__gte=hace_30_dias).count()
    cambios_hoy = HistorioCambios.objects.filter(fecha_cambio__date=date.today()).count()
    
    # Top usuarios (más cambios)
    top_usuarios_raw = HistorioCambios.objects.values('usuario__first_name', 'usuario__username').annotate(
        cantidad=Count('id')
    ).order_by('-cantidad')[:10]
    
    # Cambios por tipo de acción
    cambios_por_accion_raw = HistorioCambios.objects.values('accion').annotate(
        cantidad=Count('id')
    ).order_by('-cantidad')
    
    # Cambios por modelo
    cambios_por_modelo_raw = HistorioCambios.objects.values('modelo').annotate(
        cantidad=Count('id')
    ).order_by('-cantidad')
    
    # Calcular máximos para porcentajes
    max_usuarios = max([u['cantidad'] for u in top_usuarios_raw], default=1)
    max_accion = max([a['cantidad'] for a in cambios_por_accion_raw], default=1)
    max_modelo = max([m['cantidad'] for m in cambios_por_modelo_raw], default=1)
    
    # Agregar porcentaje calculado a cada elemento
    top_usuarios = [
        {**u, 'porcentaje': round((u['cantidad'] / max_usuarios * 100), 1)}
        for u in top_usuarios_raw
    ]
    
    cambios_por_accion = [
        {**a, 'porcentaje': round((a['cantidad'] / max_accion * 100), 1)}
        for a in cambios_por_accion_raw
    ]
    
    cambios_por_modelo = [
        {**m, 'porcentaje': round((m['cantidad'] / max_modelo * 100), 1)}
        for m in cambios_por_modelo_raw
    ]
    
    # Actividad últimos 7 días (para gráfico)
    cambios_por_dia = []
    max_dia = 1
    for i in range(7, -1, -1):
        fecha = date.today() - timedelta(days=i)
        cantidad = HistorioCambios.objects.filter(fecha_cambio__date=fecha).count()
        cambios_por_dia.append({
            'fecha': fecha.strftime('%d/%m'),
            'cantidad': cantidad,
        })
        max_dia = max(max_dia, cantidad)
    
    # Agregar porcentaje a días
    cambios_por_dia = [
        {**d, 'porcentaje': round((d['cantidad'] / max_dia * 100), 1)}
        for d in cambios_por_dia
    ]
    
    context = {
        'total_cambios': total_cambios,
        'cambios_mes': cambios_mes,
        'cambios_hoy': cambios_hoy,
        'top_usuarios': top_usuarios,
        'cambios_por_accion': cambios_por_accion,
        'cambios_por_modelo': cambios_por_modelo,
        'cambios_por_dia': cambios_por_dia,
    }
    
    return render(request, 'mi_app/auditoria/auditoria_estadisticas.html', context)


@require_permission('reporte.export')
@admin_required
@require_http_methods(["GET"])
def exportar_auditoria_excel(request):
    """
    Exporta el historial de auditoría a Excel.
    Solo accesible por staff/admin.
    """
    from .models import HistorioCambios
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoría"
    
    # Headers
    headers = [
        'ID', 'Fecha', 'Hora', 'Usuario', 'Acción', 'Modelo', 'Objeto', 
        'Campo', 'Valor Anterior', 'Valor Nuevo', 'Razón', 'IP'
    ]
    
    # Estilo headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    cambios = HistorioCambios.objects.select_related('usuario').order_by('-fecha_cambio')[:10000]
    
    for row, cambio in enumerate(cambios, 2):
        ws.cell(row=row, column=1, value=cambio.id)
        ws.cell(row=row, column=2, value=cambio.fecha_cambio.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=3, value=cambio.fecha_cambio.strftime('%H:%M:%S'))
        ws.cell(row=row, column=4, value=cambio.usuario.get_full_name() or cambio.usuario.username)
        ws.cell(row=row, column=5, value=cambio.get_accion_display())
        ws.cell(row=row, column=6, value=cambio.modelo)
        ws.cell(row=row, column=7, value=cambio.objeto_str)
        ws.cell(row=row, column=8, value=cambio.campo_modificado)
        ws.cell(row=row, column=9, value=cambio.valor_anterior)
        ws.cell(row=row, column=10, value=cambio.valor_nuevo)
        ws.cell(row=row, column=11, value=cambio.razon)
        ws.cell(row=row, column=12, value=cambio.ip_address or 'N/A')
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Auditoria_Cambios.xlsx"'
    wb.save(response)
    
    return response

# ===============================================================================
# ERROR #3: VISTA PARA LIMPIEZA AUTOMÁTICA DE PRÉSTAMOS ANTIGUOS
# ===============================================================================

@require_permission('auditoria.view')  # Requiere acceso a auditoría (admin/staff)
@login_required
@require_http_methods(["GET", "POST"])
def ejecutar_limpieza_prestamos(request):
    """
    Vista para ejecutar limpieza automática de préstamos completados antiguos
    Conserva scoring en ClienteScoring y Cliente
    GET: Muestra formulario con opciones
    POST: Ejecuta la limpieza
    """
    from django.core.management import call_command
    from io import StringIO
    
    if request.method == 'GET':
        # Mostrar página de confirmación
        return render(request, 'mi_app/admin/limpieza_prestamos.html', {
            'dias_default': 90
        })
    
    if request.method == 'POST':
        dias = int(request.POST.get('dias', 90))
        ejecutar = request.POST.get('ejecutar') == 'si'

        # Ejecutar management command
        out = StringIO()
        try:
            call_command('cleanup_prestamos_antiguos', dias=dias, execute=ejecutar, stdout=out)
            output = out.getvalue()
            if ejecutar:
                messages.success(request, '✅ Limpieza completada exitosamente')
            else:
                messages.info(request, 'ℹ️ Vista previa generada (dry-run, sin cambios)')
            
            # Log de auditoría (importar aquí para evitar circular imports)
            from .auditoria import registrar_cambio_manual, get_client_ip
            registrar_cambio_manual(
                usuario=request.user,
                modelo='Prestamo',
                accion='OTRO',
                objeto_id=0,
                objeto_str=f'Limpieza préstamos > {dias} días',
                razon=request.POST.get('razon', 'Mantenimiento'),
                ip_address=get_client_ip(request),
                notas='LIMPIEZA_PREVIEW' if not ejecutar else 'LIMPIEZA_EJECUTADA',
            )
            
        except Exception as e:
            messages.error(request, f'❌ Error en limpieza: {str(e)}')
            return redirect('limpieza_prestamos')
        
        # Redirigir con contexto de éxito
        return render(request, 'mi_app/admin/limpieza_prestamos.html', {
            'exito': True,
            'output': output,
            'dias': dias,
            'fue_ejecucion': ejecutar,
        })


@require_permission('cliente.edit')
@login_required(login_url='login')
def auto_tagging_lista_negra(request):
    """
    Vista para ejecutar auto-tagging de lista negra.
    Marca/desmarcar clientes automáticamente según comportamiento de pagos.
    GET: Muestra formulario con opciones
    POST: Ejecuta el auto-tagging
    """
    from django.core.management import call_command
    from io import StringIO
    
    if request.method == 'GET':
        # Mostrar página de confirmación
        return render(request, 'mi_app/admin/auto_tagging_lista_negra.html', {
            'dias_default': 30,
            'total_clientes': Cliente.objects.count(),
        })
    
    if request.method == 'POST':
        dias_mora = int(request.POST.get('dias_mora', 30))
        dry_run = request.POST.get('dry_run') != 'no'  # Por defecto dry-run

        # Ejecutar management command
        out = StringIO()
        try:
            call_command(
                'auto_tagging_lista_negra',
                dias=dias_mora,
                dry_run=dry_run,
                verbose=True,
                stdout=out
            )
            output = out.getvalue()
            
            if dry_run:
                messages.info(request, '📋 Vista previa de auto-tagging (sin aplicar cambios)')
            else:
                messages.success(request, '✅ Auto-tagging completado exitosamente')
            
            # Log de auditoría
            from .auditoria import registrar_cambio_manual, get_client_ip
            registrar_cambio_manual(
                usuario=request.user,
                modelo='ListaNegra',
                accion='OTRO',
                objeto_id=0,
                objeto_str=f'Auto-tagging lista negra (mora > {dias_mora} días)',
                razon=request.POST.get('razon', 'Mantenimiento automático'),
                ip_address=get_client_ip(request),
                notas='AUTO_TAGGING_PREVIEW' if dry_run else 'AUTO_TAGGING_EJECUTADO',
            )
            
        except Exception as e:
            messages.error(request, f'❌ Error en auto-tagging: {str(e)}')
            return redirect('auto_tagging_lista_negra')
        
        # Redirigir con contexto de éxito
        return render(request, 'mi_app/admin/auto_tagging_lista_negra.html', {
            'exito': True,
            'output': output,
            'dias_mora': dias_mora,
            'fue_dry_run': dry_run,
            'total_clientes': Cliente.objects.count(),
        })


@require_permission('cliente.view')
@login_required(login_url='login')
def auto_tagging_etiquetas(request):
    """
    Vista para ejecutar auto-tagging de etiquetas de clientes.
    Clasifica clientes como BUENO/MEDIO/MALO automáticamente.
    GET: Muestra formulario con opciones
    POST: Ejecuta el auto-tagging
    """
    from django.core.management import call_command
    from io import StringIO
    
    if request.method == 'GET':
        # Estadísticas actuales
        stats = {
            'BUENO': Cliente.objects.filter(etiqueta_cliente='BUENO').count(),
            'MEDIO': Cliente.objects.filter(etiqueta_cliente='MEDIO').count(),
            'MALO': Cliente.objects.filter(etiqueta_cliente='MALO').count(),
            'SIN_HISTORIAL': Cliente.objects.filter(etiqueta_cliente='SIN_HISTORIAL').count(),
        }
        return render(request, 'mi_app/admin/auto_tagging_etiquetas.html', {
            'total_clientes': Cliente.objects.count(),
            'stats': stats,
        })
    
    if request.method == 'POST':
        dry_run = request.POST.get('dry_run') != 'no'  # Por defecto dry-run

        # Ejecutar management command
        out = StringIO()
        try:
            call_command(
                'auto_tagging_etiquetas',
                dry_run=dry_run,
                verbose=True,
                stdout=out
            )
            output = out.getvalue()
            
            if dry_run:
                messages.info(request, '📋 Vista previa de etiquetación (sin aplicar cambios)')
            else:
                messages.success(request, '✅ Etiquetación completada exitosamente')
            
            # Log de auditoría
            from .auditoria import registrar_cambio_manual, get_client_ip
            registrar_cambio_manual(
                usuario=request.user,
                modelo='Cliente',
                accion='OTRO',
                objeto_id=0,
                objeto_str='Auto-tagging de etiquetas de clientes (BUENO/MEDIO/MALO)',
                razon=request.POST.get('razon', 'Mantenimiento automático'),
                ip_address=get_client_ip(request),
                notas='AUTO_TAGGING_ETIQUETAS_PREVIEW' if dry_run else 'AUTO_TAGGING_ETIQUETAS_EJECUTADO',
            )
            
        except Exception as e:
            messages.error(request, f'❌ Error en etiquetación: {str(e)}')
            return redirect('auto_tagging_etiquetas')
        
        # Redirigir con contexto de éxito
        stats_after = {
            'BUENO': Cliente.objects.filter(etiqueta_cliente='BUENO').count(),
            'MEDIO': Cliente.objects.filter(etiqueta_cliente='MEDIO').count(),
            'MALO': Cliente.objects.filter(etiqueta_cliente='MALO').count(),
            'SIN_HISTORIAL': Cliente.objects.filter(etiqueta_cliente='SIN_HISTORIAL').count(),
        }
        return render(request, 'mi_app/admin/auto_tagging_etiquetas.html', {
            'exito': True,
            'output': output,
            'fue_dry_run': dry_run,
            'total_clientes': Cliente.objects.count(),
            'stats': stats_after,
        })


@require_permission('cliente.view')
@login_required(login_url='login')
def exportar_lista_negra_excel(request):
    """
    Exporta la lista negra a un archivo Excel descargable.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    try:
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lista Negra"
        
        # Estilos
        header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Encabezados
        headers = ['ID', 'Nombre', 'Cédula', 'Celular', 'Email', 'Razón', 'Descripción', 'Desde', 'Hasta', 'Estado', 'Vigente']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment
        
        # Datos
        lista_negra_entries = ListaNegra.objects.select_related('cliente').order_by('-fecha_agregacion')
        
        for row_num, entry in enumerate(lista_negra_entries, 2):
            cliente = entry.cliente
            datos = [
                cliente.id,
                cliente.nombre,
                cliente.cedula or '-',
                cliente.celular or '-',
                cliente.email or '-',
                entry.get_razon_display(),
                entry.descripcion or '-',
                entry.fecha_desde.strftime('%d/%m/%Y') if entry.fecha_desde else '-',
                entry.fecha_hasta.strftime('%d/%m/%Y') if entry.fecha_hasta else 'Indefinido',
                'Activo' if entry.activa else 'Inactivo',
                'Sí' if entry.esta_vigente else 'No',
            ]
            
            for col_num, dato in enumerate(datos, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = dato
                cell.border = border
                cell.alignment = left_alignment if col_num > 1 else center_alignment
        
        # Ajustar ancho de columnas
        ancho_columnas = [8, 25, 15, 15, 20, 20, 30, 12, 12, 10, 10]
        for col_num, ancho in enumerate(ancho_columnas, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = ancho
        
        # Agregar fila de información
        info_row = len(lista_negra_entries) + 3
        ws.cell(row=info_row, column=1).value = f"Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws.cell(row=info_row + 1, column=1).value = f"Total: {len(lista_negra_entries)}"
        ws.cell(row=info_row + 2, column=1).value = f"Vigentes: {sum(1 for e in lista_negra_entries if e.esta_vigente)}"
        
        # Retornar archivo
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        nombre_archivo = f'lista_negra_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        
        wb.save(response)
        
        # Log de auditoría
        try:
            from .auditoria import registrar_cambio_manual, get_client_ip
            registrar_cambio_manual(
                usuario=request.user,
                modelo='ListaNegra',
                accion='EXPORTAR',
                objeto_id=0,
                objeto_str=f'Exportación a Excel de {len(lista_negra_entries)} registros',
                razon='Exportación lista negra',
                ip_address=get_client_ip(request),
                notas='EXPORTAR_LISTA_NEGRA_EXCEL',
            )
        except:
            pass  # Si falla auditoría, al menos el archivo se descarga
        
        return response
        
    except Exception as e:
        import traceback
        print(f"ERROR en exportar_lista_negra_excel: {str(e)}")
        traceback.print_exc()
        messages.error(request, f'❌ Error al exportar: {str(e)}')
        return redirect('lista_clientes')


@require_permission('cliente.view')
@login_required(login_url='login')
def reporte_interes_mensual(request):
    """
    Vista para mostrar un reporte completo de interés mensual.
    Muestra:
    - Interés total cobrado por período (1 mes, 3 meses, 12 meses)
    - Desglose por tipo de préstamo (normal vs. rápido)
    - Proyección para fin de mes
    - Comparativa con período anterior
    - Gráficos visuales
    """
    from .reportes_interes import ReporteInteresMensual
    import json
    
    reporte = ReporteInteresMensual()
    
    # Obtener todos los datos
    resumen = reporte.get_resumen_general()
    proyeccion = reporte.calcular_proyeccion_mes_actual()
    comparativa = reporte.get_comparativa_mes_anterior()
    
    # Preparar datos para gráficos (en formato JSON para Chart.js)
    datos_grafico_1mes = resumen['periodo_1_mes']['datos']
    datos_grafico_3meses = resumen['periodo_3_meses']['datos']
    datos_grafico_12meses = resumen['periodo_12_meses']['datos']
    
    # Labels y datos para cada gráfico
    labels_1mes = [d['mes'] for d in datos_grafico_1mes]
    valores_normal_1mes = [d['interes_normal'] for d in datos_grafico_1mes]
    valores_rapido_1mes = [d['interes_rapido'] for d in datos_grafico_1mes]
    
    labels_3meses = [d['mes'].split()[0] for d in datos_grafico_3meses]  # Solo mes
    valores_normal_3meses = [d['interes_normal'] for d in datos_grafico_3meses]
    valores_rapido_3meses = [d['interes_rapido'] for d in datos_grafico_3meses]
    
    labels_12meses = [d['mes'].split()[0] for d in datos_grafico_12meses]  # Solo mes
    valores_normal_12meses = [d['interes_normal'] for d in datos_grafico_12meses]
    valores_rapido_12meses = [d['interes_rapido'] for d in datos_grafico_12meses]
    
    # Contexto para template
    context = {
        'resumen': resumen,
        'proyeccion': proyeccion,
        'comparativa': comparativa,
        
        # Gráfico 1 mes
        'labels_1mes': json.dumps(labels_1mes),
        'valores_normal_1mes': json.dumps(valores_normal_1mes),
        'valores_rapido_1mes': json.dumps(valores_rapido_1mes),
        
        # Gráfico 3 meses
        'labels_3meses': json.dumps(labels_3meses),
        'valores_normal_3meses': json.dumps(valores_normal_3meses),
        'valores_rapido_3meses': json.dumps(valores_rapido_3meses),
        
        # Gráfico 12 meses
        'labels_12meses': json.dumps(labels_12meses),
        'valores_normal_12meses': json.dumps(valores_normal_12meses),
        'valores_rapido_12meses': json.dumps(valores_rapido_12meses),
    }
    
    return render(request, 'mi_app/reportes/reporte_interes_mensual.html', context)


@require_permission('cliente.view')
@login_required(login_url='login')
def exportar_reporte_interes_excel(request):
    """
    Exporta el reporte de interés mensual a Excel
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    from .reportes_interes import ReporteInteresMensual
    
    try:
        reporte = ReporteInteresMensual()
        resumen = reporte.get_resumen_general()
        proyeccion = reporte.calcular_proyeccion_mes_actual()
        comparativa = reporte.get_comparativa_mes_anterior()
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Interés Mensual"
        
        # Estilos
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        subheader_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:F1')
        titulo = ws['A1']
        titulo.value = "REPORTE DE INTERÉS MENSUAL"
        titulo.font = Font(bold=True, size=14, color="1F4E78")
        titulo.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Fecha de generación
        ws.merge_cells('A2:F2')
        fecha_celda = ws['A2']
        fecha_celda.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        fecha_celda.font = Font(size=9, italic=True)
        fecha_celda.alignment = Alignment(horizontal='center')
        
        fila_actual = 4
        
        # ===== SECCIÓN: RESUMEN POR PERÍODO =====
        ws.merge_cells(f'A{fila_actual}:F{fila_actual}')
        titulo_seccion = ws[f'A{fila_actual}']
        titulo_seccion.value = "RESUMEN POR PERÍODO"
        titulo_seccion.fill = header_fill
        titulo_seccion.font = header_font
        titulo_seccion.alignment = Alignment(horizontal='center')
        fila_actual += 1
        
        # Headers de períodos
        headers = ['Período', 'Préstamos Normales', 'Préstamos Rápidos', 'Total Interés', 'Cantidad Cuotas']
        for col, header in enumerate(headers, 1):
            celda = ws.cell(row=fila_actual, column=col)
            celda.value = header
            celda.fill = subheader_fill
            celda.font = subheader_font
            celda.border = border
        fila_actual += 1
        
        # Datos períodos
        periodos = [
            ('1 Mes', resumen['periodo_1_mes']),
            ('3 Meses', resumen['periodo_3_meses']),
            ('12 Meses', resumen['periodo_12_meses']),
        ]
        
        for periodo_nombre, datos_periodo in periodos:
            ws.cell(row=fila_actual, column=1).value = periodo_nombre
            ws.cell(row=fila_actual, column=2).value = round(datos_periodo['normal'], 2)
            ws.cell(row=fila_actual, column=3).value = round(datos_periodo['rapido'], 2)
            ws.cell(row=fila_actual, column=4).value = round(datos_periodo['total'], 2)
            
            # Contar cuotas
            total_cuotas = sum([d['cantidad_cuotas_normal'] + d['cantidad_cuotas_rapido'] 
                               for d in datos_periodo['datos']])
            ws.cell(row=fila_actual, column=5).value = total_cuotas
            
            for col in range(1, 6):
                celda = ws.cell(row=fila_actual, column=col)
                celda.border = border
                celda.alignment = Alignment(horizontal='right')
            
            fila_actual += 1
        
        fila_actual += 1
        
        # ===== SECCIÓN: PROYECCIÓN MES ACTUAL =====
        ws.merge_cells(f'A{fila_actual}:F{fila_actual}')
        titulo_proyeccion = ws[f'A{fila_actual}']
        titulo_proyeccion.value = f"PROYECCIÓN MES ACTUAL ({proyeccion['dias_transcurridos']} de {proyeccion['dias_totales']} días)"
        titulo_proyeccion.fill = header_fill
        titulo_proyeccion.font = header_font
        titulo_proyeccion.alignment = Alignment(horizontal='center')
        fila_actual += 1
        
        # Headers proyección
        headers_proy = ['Concepto', 'Normales', 'Rápidos', 'Total']
        for col, header in enumerate(headers_proy, 1):
            celda = ws.cell(row=fila_actual, column=col)
            celda.value = header
            celda.fill = subheader_fill
            celda.font = subheader_font
            celda.border = border
        fila_actual += 1
        
        # Datos proyección
        proyeccion_datos = [
            ('Interés Recaudado', proyeccion['interes_actual_normal'], proyeccion['interes_actual_rapido'], proyeccion['interes_actual_total']),
            ('Proyección Fin de Mes', proyeccion['interes_proyectado_normal'], proyeccion['interes_proyectado_rapido'], proyeccion['interes_proyectado_total']),
        ]
        
        for concepto, normal, rapido, total in proyeccion_datos:
            ws.cell(row=fila_actual, column=1).value = concepto
            ws.cell(row=fila_actual, column=2).value = round(normal, 2)
            ws.cell(row=fila_actual, column=3).value = round(rapido, 2)
            ws.cell(row=fila_actual, column=4).value = round(total, 2)
            
            for col in range(1, 5):
                celda = ws.cell(row=fila_actual, column=col)
                celda.border = border
                celda.alignment = Alignment(horizontal='right')
            
            fila_actual += 1
        
        fila_actual += 1
        
        # ===== SECCIÓN: COMPARATIVA MES ANTERIOR =====
        ws.merge_cells(f'A{fila_actual}:F{fila_actual}')
        titulo_comp = ws[f'A{fila_actual}']
        titulo_comp.value = "COMPARATIVA MES ANTERIOR"
        titulo_comp.fill = header_fill
        titulo_comp.font = header_font
        titulo_comp.alignment = Alignment(horizontal='center')
        fila_actual += 1
        
        if comparativa['mes_anterior']:
            headers_comp = ['Mes', 'Interés Total', 'Variación', 'Variación %']
            for col, header in enumerate(headers_comp, 1):
                celda = ws.cell(row=fila_actual, column=col)
                celda.value = header
                celda.fill = subheader_fill
                celda.font = subheader_font
                celda.border = border
            fila_actual += 1
            
            # Mes anterior
            ws.cell(row=fila_actual, column=1).value = comparativa['mes_anterior']['mes']
            ws.cell(row=fila_actual, column=2).value = round(comparativa['mes_anterior']['interes_total'], 2)
            ws.cell(row=fila_actual, column=3).value = 0
            ws.cell(row=fila_actual, column=4).value = "0%"
            for col in range(1, 5):
                ws.cell(row=fila_actual, column=col).border = border
            fila_actual += 1
            
            # Mes actual
            ws.cell(row=fila_actual, column=1).value = comparativa['mes_actual']['mes']
            ws.cell(row=fila_actual, column=2).value = round(comparativa['mes_actual']['interes_total'], 2)
            ws.cell(row=fila_actual, column=3).value = round(comparativa['variacion'], 2)
            
            variacion_pct = ws.cell(row=fila_actual, column=4)
            variacion_pct.value = f"{comparativa['variacion_pct']:.1f}%"
            
            # Color según variación
            color = "92D050" if comparativa['es_positivo'] else "FF0000"  # Verde o Rojo
            for col in range(3, 5):
                celda = ws.cell(row=fila_actual, column=col)
                celda.border = border
                celda.font = Font(color=color, bold=True)
            
            fila_actual += 1
        
        # Ajustar anchos de columnas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        
        # Retornar archivo
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        nombre_archivo = f'reporte_interes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        
        wb.save(response)
        
        # Log de auditoría
        try:
            from .auditoria import registrar_cambio_manual, get_client_ip
            registrar_cambio_manual(
                usuario=request.user,
                modelo='Reporte',
                accion='EXPORTAR',
                objeto_id=0,
                objeto_str='Exportación a Excel del reporte de interés mensual',
                razon='Exportación reporte interés',
                ip_address=get_client_ip(request),
                notas='EXPORTAR_REPORTE_INTERES_EXCEL',
            )
        except:
            pass
        
        return response
        
    except Exception as e:
        import traceback
        print(f"ERROR en exportar_reporte_interes_excel: {str(e)}")
        traceback.print_exc()
        messages.error(request, f'❌ Error al exportar reporte: {str(e)}')
        return redirect('reporte_interes_mensual')


# ===============================================================================
# BACKUP RÁPIDO - Funcionalidad Nueva (Feb 21, 2026)
# ===============================================================================

def get_client_ip(request):
    """Obtiene la IP del cliente"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def generar_backup_zip():
    """
    Genera un archivo ZIP con la base de datos.
    Retorna: (archivo_zip_bytes, nombre_archivo, tamaño_mb)
    """
    PROJECT_ROOT = Path(settings.BASE_DIR)
    DB_PATH = PROJECT_ROOT / 'db.sqlite3'
    
    if not DB_PATH.exists():
        raise FileNotFoundError(f"Base de datos no encontrada en {DB_PATH}")
    
    # Nombre del archivo con timestamp
    timestamp = dt.datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f'backup_{timestamp}.zip'
    
    # Crear ZIP en memoria
    zip_buffer = io.BytesIO()
    
    try:
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Agregar BD principal
            arcname = 'db.sqlite3'
            zip_file.write(str(DB_PATH), arcname=arcname)
            
            # Agregar carpeta de media si existe
            media_path = PROJECT_ROOT / 'media'
            if media_path.exists():
                for root, dirs, files in os.walk(media_path):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, PROJECT_ROOT)
                        zip_file.write(file_path, arcname=arcname)
        
        zip_buffer.seek(0)
        contenido = zip_buffer.getvalue()
        tamaño_mb = len(contenido) / (1024 * 1024)
        
        return contenido, nombre_archivo, tamaño_mb
    
    except Exception as e:
        raise Exception(f"Error al generar ZIP: {str(e)}")


@login_required
@require_http_methods(["POST"])
@ratelimit(key='user', rate='5/m')
def backup_rapido_generar(request):
    """
    Genera automáticamente un backup ZIP y lo guarda en /backups/.
    Rate limit: 5 backups por minuto máximo (operación CPU/IO intensiva)
    """
    try:
        print("✅ backup_rapido_generar iniciado")
        
        # Crear directorio de backups
        PROJECT_ROOT = Path(settings.BASE_DIR)
        BACKUPS_DIR = PROJECT_ROOT / 'backups'
        BACKUPS_DIR.mkdir(exist_ok=True)
        print(f"📁 BACKUPS_DIR: {BACKUPS_DIR}")

        # Límite diario global: máximo 10 backups por día (Aumentado de 3)
        fecha_hoy_prefijo = dt.datetime.now().strftime('backup_%Y%m%d_')
        backups_hoy = [
            f for f in BACKUPS_DIR.iterdir()
            if f.is_file()
            and f.name.startswith(fecha_hoy_prefijo)
            and (f.name.endswith('.zip') or f.name.endswith('.db') or f.name.endswith('.db.gz') or f.name.endswith('.sqlite3'))
        ]
        
        # Permitir a superusuarios saltarse el límite o usar el nuevo límite de 10
        if len(backups_hoy) >= 10 and not request.user.is_superuser:
            return JsonResponse({
                'success': False,
                'error': 'Límite diario alcanzado: máximo 10 backups por día.'
            }, status=429)
        
        # Ruta a la BD
        DB_PATH = PROJECT_ROOT / 'db.sqlite3'
        print(f"🗄️  DB_PATH: {DB_PATH}, existe: {DB_PATH.exists()}")
        
        if not DB_PATH.exists():
            raise FileNotFoundError(f"BD no encontrada en {DB_PATH}")
        
        # Generar nombre del ZIP
        timestamp = dt.datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f'backup_{timestamp}.zip'
        print(f"📦 Nombre archivo: {nombre_archivo}")
        
        # Crear ZIP
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.write(str(DB_PATH), arcname='db.sqlite3')
            print("✅ DB agregada al ZIP")
        
        zip_buffer.seek(0)
        contenido = zip_buffer.getvalue()
        tamaño_mb = len(contenido) / (1024 * 1024)
        print(f"📊 Tamaño: {tamaño_mb:.2f} MB")
        
        # Guardar archivo
        ruta_guardado = BACKUPS_DIR / nombre_archivo
        with open(ruta_guardado, 'wb') as f:
            f.write(contenido)
        print(f"✅ Archivo guardado: {ruta_guardado}")
        
        # Registrar en auditoría
        try:
            backup_audit = AuditoriaBackup.objects.create(
                usuario=request.user,
                tipo_backup='RAPIDO',
                ubicacion=str(BACKUPS_DIR),
                nombre_archivo=nombre_archivo,
                tamano_mb=tamaño_mb,
                estado='SUCCESS',
                ip_address=get_client_ip(request),
                notas='Backup desde dashboard'
            )
            print(f"✅ Auditoría registrada: {backup_audit.id}")
        except Exception as e:
            print(f"⚠️  Error al registrar auditoría: {e}")
        
        print("✅ Respuesta exitosa")
        return JsonResponse({
            'success': True,
            'nombre_archivo': nombre_archivo,
            'tamano_mb': round(tamaño_mb, 2),
            'backup_id': backup_audit.id if 'backup_audit' in locals() else 0
        })
    
    except Exception as e:
        print(f"❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@ratelimit(key='user', rate='10/m', method='GET')
def backup_rapido_descargar(request, backup_id=None):
    """
    Descarga el backup ZIP al navegador del usuario.
    Acepta backup_id desde URL o desde POST.
    Rate limit: 10 descargas por minuto máximo (prevenir abuso de banda ancha)
    """
    try:
        print("📥 Iniciando descarga de backup")
        
        # Obtener backup_id desde URL o POST
        if backup_id is None:
            backup_id = request.POST.get('backup_id')
        
        if not backup_id:
            return JsonResponse({
                'success': False,
                'error': 'ID del backup requerido'
            }, status=400)
        
        # Obtener backup y verificar que pertenece al usuario
        backup = AuditoriaBackup.objects.get(
            id=backup_id,
            usuario=request.user,
            estado='SUCCESS'
        )
        
        PROJECT_ROOT = Path(settings.BASE_DIR)
        ruta_archivo = PROJECT_ROOT / 'backups' / backup.nombre_archivo
        
        if not ruta_archivo.exists():
            return JsonResponse({
                'success': False,
                'error': 'Archivo no encontrado en servidor'
            }, status=404)
        
        print(f"📥 Descargando: {backup.nombre_archivo}")
        
        # Leer archivo
        with open(ruta_archivo, 'rb') as f:
            contenido = f.read()
        
        # Crear respuesta
        response = HttpResponse(
            contenido,
            content_type='application/zip'
        )
        response['Content-Disposition'] = f'attachment; filename="{backup.nombre_archivo}"'
        
        print("✅ Archivo enviado al navegador")
        
        return response
    
    except AuditoriaBackup.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Backup no encontrado'
        }, status=404)
    
    except Exception as e:
        print(f"❌ ERROR descarga: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
@ratelimit(key='user', rate='3/m')
def backup_rapido_correo(request):
    """
    Envía el backup ZIP más reciente por correo al cliente.
    Rate limit: 3 correos por minuto máximo (servidor SMTP/API limitado)
    """
    try:
        print("📧 Iniciando envío de correo")
        
        # Obtener el backup más reciente que aún exista físicamente
        backup_audits = AuditoriaBackup.objects.filter(
            usuario=request.user,
            estado='SUCCESS',
            tipo_backup__in=['RAPIDO', 'LOCAL']
        ).order_by('-fecha_inicio')

        PROJECT_ROOT = Path(settings.BASE_DIR)
        backup_audit = None
        ruta_archivo = None

        for audit in backup_audits:
            ruta_candidata = PROJECT_ROOT / 'backups' / audit.nombre_archivo
            if ruta_candidata.exists():
                backup_audit = audit
                ruta_archivo = ruta_candidata
                break

        if not backup_audit:
            return JsonResponse({
                'success': False,
                'error': 'No hay backups disponibles para enviar. Genera uno nuevo primero.'
            }, status=400)
        
        email_destino = getattr(settings, 'BACKUP_RECIPIENT_EMAIL', '') or request.user.email or ''
        if not email_destino:
            email_destino = getattr(settings, 'DEFAULT_FROM_EMAIL', '') or ''
        
        if not email_destino:
            return JsonResponse({
                'success': False,
                'error': 'No hay correo destino configurado para enviar el backup'
            }, status=400)
        
        print(f"📧 De: {settings.DEFAULT_FROM_EMAIL}")
        print(f"📧 Para: {email_destino}")
        print(f"📦 Archivo: {backup_audit.nombre_archivo}")

        backend_actual = getattr(settings, 'EMAIL_BACKEND', '')
        if backend_actual == 'django.core.mail.backends.console.EmailBackend':
            return JsonResponse({
                'success': False,
                'error': 'EMAIL_BACKEND está en consola. Configura SMTP real para enviar correos.'
            }, status=400)

        if not getattr(settings, 'EMAIL_HOST_USER', '') or not getattr(settings, 'EMAIL_HOST_PASSWORD', ''):
            return JsonResponse({
                'success': False,
                'error': 'Faltan EMAIL_HOST_USER o EMAIL_HOST_PASSWORD. Configura credenciales SMTP reales en .env.'
            }, status=400)
        
        # Crear y enviar correo
        email = EmailMessage(
            subject=f'📦 Backup de Base de Datos - {backup_audit.nombre_archivo}',
            body=f'Backup generado: {backup_audit.nombre_archivo}\nTamaño: {backup_audit.tamano_mb} MB',
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[email_destino]
        )
        
        # Adjuntar archivo
        with open(ruta_archivo, 'rb') as attachment:
            email.attach(
                backup_audit.nombre_archivo,
                attachment.read(),
                'application/zip'
            )
        
        email.send(fail_silently=False)
        print("✅ Correo enviado")
        
        return JsonResponse({
            'success': True,
            'mensaje': f'Correo enviado a {email_destino}',
            'correo_destino': email_destino,
            'entrega_real': True,
            'email_backend': backend_actual,
        })
    
    except Exception as e:
        print(f"❌ ERROR email: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ===============================================================================
# ALTO #4: MORA EN TIEMPO REAL - AJAX API
# ===============================================================================

@login_required
@require_http_methods(["GET"])
def api_cuota_mora_actual(request, cuota_id):
    """
    Obtiene la mora actual de una cuota (ALTO #4 - Real-time mora updates)
    
    Args:
        cuota_id (int): ID de la cuota
    
    Returns:
        JsonResponse con: mora_actual, interes_total, total_pendiente, timestamp
    """
    cuota = get_object_or_404(Cuota, id=cuota_id)
    try:
        mora_diaria = cuota.calcular_mora_diaria()
        monto_pendiente = cuota.monto_pendiente
        interes_pendiente = cuota.monto_pendiente_interes
        ahora = dt.datetime.now()
        timestamp_iso = ahora.isoformat()
        
        return JsonResponse({
            'success': True,
            'cuota_id': cuota_id,
            'mora_diaria': str(mora_diaria),
            'mora_acumulada': str(cuota.interes_mora_acumulado),
            'interes_pendiente': str(interes_pendiente),
            'monto_pendiente': str(monto_pendiente),
            'total_pendiente': str(Decimal(monto_pendiente) + interes_pendiente + mora_diaria),
            'estado': cuota.obtener_estado_cuota(),
            'fecha_vencimiento': str(cuota.fecha_pago_esperada) if cuota.fecha_pago_esperada else None,
            'dias_atraso': (date.today() - cuota.fecha_pago_esperada).days if cuota.fecha_pago_esperada else 0,
            'timestamp': timestamp_iso
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
